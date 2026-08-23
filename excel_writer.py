from __future__ import annotations

from numbers import Number
from pathlib import Path
from typing import Any
from collections import defaultdict
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
# Import
from datetime import date, timedelta
import calendar

from config import OUTPUT_DIR, OUTPUT_FILE, INPUT_DIR, PRODUCTS, AREAS
from processor import SalesProcessor


class ExcelWriter:
    def __init__(self) -> None:
        self.processor = SalesProcessor()
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.wb.calculation.fullCalcOnLoad = True
        self.wb.calculation.forceFullCalc = True
        self.wb.calculation.calcMode = "auto"

        self.dark_fill = PatternFill("solid", fgColor="1F4E78")
        self.light_fill = PatternFill("solid", fgColor="D9EAD3")
        self.area_fill = PatternFill("solid", fgColor="DDEBF7")
        self.white_bold = Font(bold=True, color="FFFFFF")
        self.bold = Font(bold=True)
        thin = Side(style="thin", color="A6A6A6")
        self.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.ddm_areas = self._load_ddm_areas()

    def _load_ddm_areas(self) -> dict[str, list[str]]:
        """Read Area -> DDM Name mapping from any workbook that has these headers."""
        result: dict[str, list[str]] = defaultdict(list)
        for path in sorted(INPUT_DIR.glob("*.xlsx")):
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
                for ws in wb.worksheets:
                    rows = ws.iter_rows(values_only=True)
                    header = next(rows, None)
                    if not header:
                        continue
                    lookup = {str(v).strip().lower(): i for i, v in enumerate(header) if v is not None}
                    area_i = lookup.get("area")
                    ddm_i = lookup.get("ddm name")
                    if area_i is None or ddm_i is None:
                        continue
                    for row in rows:
                        area = str(row[area_i]).strip().upper() if area_i < len(row) and row[area_i] is not None else ""
                        ddm = str(row[ddm_i]).strip() if ddm_i < len(row) and row[ddm_i] is not None else ""
                        if area in AREAS and ddm and area not in result[ddm]:
                            result[ddm].append(area)
                wb.close()
            except Exception:
                continue
        return dict(result)

    @staticmethod
    def _ddm_sheet_name(ddm: str) -> str:
        safe = "".join(ch if ch.isalnum() else "_" for ch in ddm).strip("_")
        return ("DDM_" + safe)[:31]

    @staticmethod
    def _area_list(area) -> list[str]:
        if isinstance(area, (list, tuple, set)):
            return list(area)
        return AREAS if area == "ALL" else [area]

    @staticmethod
    def _display_number(value: Any) -> Any:
        if isinstance(value, Number) and not isinstance(value, bool):
            return float(value) / 1000.0
        return value

    def _title(self, ws, text: str, end_col: int = 8) -> None:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        cell = ws.cell(1, 1, text)
        cell.fill = self.dark_fill
        cell.font = Font(bold=True, size=16, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 26

    def _write_table(
        self,
        ws,
        rows: list[dict[str, Any]],
        columns: list[Any],
        start_row: int,
        title: str | None = None,
        round_daily: bool = False,
    ) -> tuple[int, int, int]:
        if title:
            ws.cell(start_row, 1, title).font = Font(bold=True, size=12)
            start_row += 1
        header_row = start_row
        for col, name in enumerate(columns, 1):
            cell = ws.cell(header_row, col, str(name))
            cell.fill = self.dark_fill
            cell.font = self.white_bold
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        percent_columns = {"TY vs LY %", "TY vs Objective %"}
        numeric_columns = {
            "This Year Sales", "Last Year Sales", "Objective",
            "TY vs LY Variance", "TY vs Objective Variance",
        }
        for row_offset, record in enumerate(rows, 1):
            excel_row = header_row + row_offset
            for col, name in enumerate(columns, 1):
                value = record.get(name, "")
                cell = ws.cell(excel_row, col)
                if name in numeric_columns or isinstance(name, int):
                    shown = self._display_number(value)
                    cell.value = round(shown) if round_daily and isinstance(shown, Number) else shown
                else:
                    cell.value = value
                cell.border = self.border
                if name in percent_columns:
                    cell.number_format = "0.00%"
                elif name in numeric_columns or isinstance(name, int):
                    cell.number_format = "#,##0" if round_daily else "#,##0.00"
                if col > 1:
                    cell.alignment = Alignment(horizontal="center")
                # Highlight every Sunday and grey-out impossible dates.
                if isinstance(name, int):
                    month_name = str(record.get("Month", ""))
                    try:
                        month_num = list(calendar.month_name).index(month_name)
                        fy_start = self.processor.start_date.year if self.processor.start_date.month >= 7 else self.processor.start_date.year - 1
                        year = fy_start if month_num >= 7 else fy_start + 1
                        max_day = calendar.monthrange(year, month_num)[1]
                        if name > max_day:
                            cell.fill = PatternFill("solid", fgColor="E7E6E6")
                        elif date(year, month_num, name).weekday() == 6:
                            cell.fill = PatternFill("solid", fgColor="FCE4D6")
                    except (ValueError, TypeError):
                        pass
        return header_row + len(rows) + 1, header_row, header_row + len(rows)

    def _add_summary_total(self, ws, rows: list[dict[str, Any]], columns: list[str], last_row: int) -> int:
        total_row = last_row + 1
        totals = {name: 0.0 for name in columns}
        for row in rows:
            for name in ("This Year Sales", "Last Year Sales", "Objective"):
                totals[name] += float(row.get(name, 0) or 0)
        totals["TY vs LY Variance"] = totals["This Year Sales"] - totals["Last Year Sales"]
        totals["TY vs LY %"] = totals["TY vs LY Variance"] / totals["Last Year Sales"] if totals["Last Year Sales"] else 0
        totals["TY vs Objective Variance"] = totals["This Year Sales"] - totals["Objective"]
        totals["TY vs Objective %"] = totals["TY vs Objective Variance"] / totals["Objective"] if totals["Objective"] else 0

        for col, name in enumerate(columns, 1):
            cell = ws.cell(total_row, col)
            cell.fill = self.light_fill
            cell.font = self.bold
            cell.border = self.border
            if col == 1:
                cell.value = "Grand Total"
            else:
                cell.value = totals.get(name, "")
                if name.endswith("%"):
                    cell.number_format = "0.00%"
                else:
                    cell.value = self._display_number(cell.value)
                    cell.number_format = "#,##0.00"
        return total_row

    def _add_daily_total(self, ws, rows: list[dict[str, Any]], columns: list[Any], last_row: int) -> int:
        """Write visible numeric totals, not formulas requiring recalculation."""
        total_row = last_row + 1

        # Sum the exact displayed values. Daily and sales columns are displayed
        # in rounded thousand litres, so totals match the numbers on screen.
        totals: dict[Any, float] = {}
        numeric_names = set(range(1, 32)) | {
            "This Year Sales", "Last Year Sales", "TY vs LY Variance"
        }
        for name in numeric_names:
            total = 0.0
            for record in rows:
                raw = record.get(name, 0) or 0
                shown = self._display_number(raw)
                total += round(float(shown))
            totals[name] = total

        ty_total = totals.get("This Year Sales", 0.0)
        ly_total = totals.get("Last Year Sales", 0.0)
        totals["TY vs LY Variance"] = ty_total - ly_total
        totals["TY vs LY %"] = (ty_total - ly_total) / ly_total if ly_total else 0.0

        for col, name in enumerate(columns, 1):
            cell = ws.cell(total_row, col)
            cell.fill = self.light_fill
            cell.font = Font(bold=True, color="000000")
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if col == 1:
                cell.value = "Grand Total"
            elif name in ("Sales Group", "Month"):
                cell.value = ""
            elif name == "TY vs LY %":
                cell.value = totals[name]
                cell.number_format = "0.00%"
            elif name in numeric_names:
                cell.value = totals.get(name, 0)
                cell.number_format = "#,##0"
            else:
                cell.value = ""

        ws.row_dimensions[total_row].height = 24
        return total_row

    def _horizontal_area_links(self, ws, product: str | None, start_row: int, area_sheet: bool = False) -> int:
        ws.cell(start_row, 1, "AREA DETAILS — CLICK AN AREA").font = Font(bold=True, size=12)
        row = start_row + 1
        for col, area in enumerate(AREAS, 1):
            cell = ws.cell(row, col, area)
            target = f"Area_{area}" if area_sheet else f"{product}_{area}"
            cell.hyperlink = f"#'{target}'!A1"
            cell.fill = self.area_fill
            cell.font = Font(bold=True, color="0563C1", underline="single")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.border
            ws.column_dimensions[get_column_letter(col)].width = max(ws.column_dimensions[get_column_letter(col)].width or 0, 12)
        ws.row_dimensions[row].height = 24
        return row + 2

    def _add_variance_formatting(self, ws, ranges: list[str]) -> None:
        # Positive values display in green with an upward marker. Negative
        # values display in red with a downward marker.
        green = Font(color="008000", bold=True)
        red = Font(color="C00000", bold=True)
        for ref in ranges:
            ws.conditional_formatting.add(ref, CellIsRule(operator="greaterThan", formula=["0"], font=green))
            ws.conditional_formatting.add(ref, CellIsRule(operator="lessThan", formula=["0"], font=red))
            min_col, min_row, max_col, max_row = range_boundaries(ref)
            for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                    min_col=min_col, max_col=max_col):
                for cell in row:
                    current = cell.number_format or "#,##0.00"
                    if "%" in current:
                        cell.number_format = '[Green]"▲ "0.00%;[Red]"▼ "0.00%;0.00%'
                    else:
                        cell.number_format = '[Green]"▲ "#,##0.00;[Red]"▼ "#,##0.00;#,##0.00'

    @staticmethod
    def _safe_sheet_name(text: str, prefix: str = "Mat") -> str:
        import re
        clean=re.sub(r"[\/*?:\[\]]", " ", text).strip()
        return (prefix+"_"+clean)[:31]

    def _division_buttons(self, ws, area: str = "ALL", row: int = 3) -> None:
        """Top-level navigation between Division 10 Fuel and Division 20 Lubricant."""
        ws.cell(row, 1, "CATEGORY").font = Font(bold=True, size=11)
        fuel_target = "Dashboard" if area == "ALL" else f"Area_{area}"
        lubricant_target = "Lubricant_Dashboard" if area == "ALL" else f"Lubricant_Area_{area}"
        for start_col, label, target, fill in (
            (4, "FUEL", fuel_target, "DDEBF7"),
            (7, "LUBRICANTS", lubricant_target, "E2F0D9"),
        ):
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + 2)
            c = ws.cell(row, start_col, label)
            c.hyperlink = f"#'{target}'!A1"
            c.fill = PatternFill("solid", fgColor=fill)
            c.font = Font(bold=True, color="0563C1", underline="single")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = self.border

    def _material_sheet_name(self, group: str, area: str) -> str:
        # Area-specific sheets prevent a link from exposing preceding area sections.
        return self._safe_sheet_name(f"{area}_{group}", "LGrp")

    def _material_buttons(self, ws, area: str, start_row: int) -> int:
        groups = self.processor.lubricant_materials()
        ws.cell(start_row, 1, "LUBRICANT MAT GROUP TEXT").font = Font(bold=True, size=11)
        row = start_row + 1
        for i, group in enumerate(groups):
            r = row + i // 4
            c = 1 + (i % 4) * 3
            target = self._material_sheet_name(group, area)
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 2)
            cell = ws.cell(r, c, group)
            cell.hyperlink = f"#'{target}'!A1"
            cell.fill = PatternFill("solid", fgColor="E2F0D9")
            cell.font = Font(bold=True, color="0563C1", underline="single", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.border
            ws.row_dimensions[r].height = 34
        return row + (len(groups) + 3) // 4 + 1

    def _create_dashboard_data(self) -> None:
        ws = self.wb.create_sheet("Dashboard_Data")
        headers = ["Area", "As Of Date", "Product", "FY Current", "FY Last Benchmark", "FY Last Daily Avg", "FY Days", "FY Objective", "MTD Current", "MTD Last Benchmark", "MTD Last Daily Avg", "MTD Days", "MTD Objective"]
        ws.append(headers)
        for area in ["ALL"] + AREAS:
            source_area = None if area == "ALL" else area
            for record in self.processor.comparison_records(source_area):
                ws.append([
                    area, record["As Of Date"], record["Product"], record["FY Current"],
                    record["FY Last Calendar"], record["FY Last Weekday Adj"], record["FY Days"], record["FY Objective"],
                    record["MTD Current"], record["MTD Last Calendar"], record["MTD Last Weekday Adj"], record["MTD Days"], record["MTD Objective"],
                ])
        self.dashboard_data_last_row = ws.max_row

        start_col = 15
        selected_headers = ["Area", "As Of Date", "Month", "Product", "Current", "Last Benchmark", "Last Daily Avg", "Days", "Objective"]
        for i, h in enumerate(selected_headers, start_col):
            ws.cell(1, i, h)
        row = 2
        for area in ["ALL"] + AREAS:
            source_area = None if area == "ALL" else area
            for record in self.processor.selected_month_records(source_area):
                values = [area, record["As Of Date"], record["Month"], record["Product"], record["Current"], record["Last Calendar"], record["Last Weekday Adj"], record["Days"], record["Objective"]]
                for i, value in enumerate(values, start_col):
                    ws.cell(row, i, value)
                row += 1
        self.selected_data_last_row = row - 1

        quarter_start_col = 24
        quarter_headers = ["Area", "As Of Date", "Quarter", "Product", "Current", "Last Benchmark", "Last Daily Avg", "Days", "Objective"]
        for i, h in enumerate(quarter_headers, quarter_start_col):
            ws.cell(1, i, h)
        row = 2
        for area in ["ALL"] + AREAS:
            source_area = None if area == "ALL" else area
            for record in self.processor.selected_quarter_records(source_area):
                values = [area, record["As Of Date"], record["Quarter"], record["Product"], record["Current"], record["Last Calendar"], record["Last Weekday Adj"], record["Days"], record["Objective"]]
                for i, value in enumerate(values, quarter_start_col):
                    ws.cell(row, i, value)
                row += 1
        self.quarter_data_last_row = row - 1
        ws.sheet_state = "hidden"

    def _dashboard_table(self, ws, title_row: int, title: str, metric_prefix: str, area, asof_cell: str, selected: bool = False, month_cell: str | None = None, quarter: bool = False, quarter_cell: str | None = None) -> int:
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=10)
        ws.cell(title_row, 1, title).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(title_row, 1).fill = self.dark_fill
        ws.cell(title_row, 1).alignment = Alignment(horizontal="center", vertical="center")
        headers = [
            "Product", "TY_Sales", "LY_Sales", "Var_LY", "Var_LY%",
            "Obj", "TY_vs_Obj", "TY_vs_Obj%", "Avg/Day", "Working Days"
        ]
        header_row = title_row + 1
        area_values = self._area_list(area)
        def sumifs_formula(sum_range: str, area_range: str, other_criteria: str) -> str:
            parts = [f'SUMIFS({sum_range},{area_range},"{a}",{other_criteria})' for a in area_values]
            return "+".join(parts) if parts else "0"
        for c, h in enumerate(headers, 1):
            cell = ws.cell(header_row, c, h)
            cell.fill = self.dark_fill; cell.font = self.white_bold; cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for idx, product in enumerate(PRODUCTS, header_row + 1):
            product_cell = ws.cell(idx, 1, product)
            # A DDM dashboard receives a list of its assigned areas. It is a
            # final calculation view, so product names must not open the
            # division-wide product control sheets. Only normal Division/Area
            # dashboards keep product drill-down hyperlinks.
            is_ddm_view = not isinstance(area, str)
            if is_ddm_view:
                product_cell.font = Font(bold=True, color="000000")
                product_cell.hyperlink = None
            else:
                product_cell.font = Font(bold=True, color="0563C1", underline="single")
                target_sheet = product if area == "ALL" else f"{product}_{area}"
                product_cell.hyperlink = f"#'{target_sheet}'!A1"
            if quarter:
                other = (f"Dashboard_Data!$Y$2:$Y${self.quarter_data_last_row},{asof_cell},"
                         f"Dashboard_Data!$Z$2:$Z${self.quarter_data_last_row},{quarter_cell},"
                         f"Dashboard_Data!$AA$2:$AA${self.quarter_data_last_row},$A{idx}")
                ar = f"Dashboard_Data!$X$2:$X${self.quarter_data_last_row}"
                cur = f"({sumifs_formula(f'Dashboard_Data!$AB$2:$AB${self.quarter_data_last_row}', ar, other)})/1000"
                ly = f"({sumifs_formula(f'Dashboard_Data!$AC$2:$AC${self.quarter_data_last_row}', ar, other)})/1000"
                days = f"MAX({','.join([f'SUMIFS(Dashboard_Data!$AE$2:$AE${self.quarter_data_last_row},{ar},"{a}",{other})' for a in area_values])})"
                obj = f"({sumifs_formula(f'Dashboard_Data!$AF$2:$AF${self.quarter_data_last_row}', ar, other)})/1000"
            elif selected:
                other = (f"Dashboard_Data!$P$2:$P${self.selected_data_last_row},{asof_cell},"
                         f"Dashboard_Data!$Q$2:$Q${self.selected_data_last_row},{month_cell},"
                         f"Dashboard_Data!$R$2:$R${self.selected_data_last_row},$A{idx}")
                ar = f"Dashboard_Data!$O$2:$O${self.selected_data_last_row}"
                cur = f"({sumifs_formula(f'Dashboard_Data!$S$2:$S${self.selected_data_last_row}', ar, other)})/1000"
                ly = f"({sumifs_formula(f'Dashboard_Data!$T$2:$T${self.selected_data_last_row}', ar, other)})/1000"
                days = f"MAX({','.join([f'SUMIFS(Dashboard_Data!$V$2:$V${self.selected_data_last_row},{ar},"{a}",{other})' for a in area_values])})"
                obj = f"({sumifs_formula(f'Dashboard_Data!$W$2:$W${self.selected_data_last_row}', ar, other)})/1000"
            else:
                colmap = {"FY": ("D", "E", "G", "H"), "MTD": ("I", "J", "L", "M")}[metric_prefix]
                cc, lc, dc, oc = colmap
                other = (f"Dashboard_Data!$B$2:$B${self.dashboard_data_last_row},{asof_cell},"
                         f"Dashboard_Data!$C$2:$C${self.dashboard_data_last_row},$A{idx}")
                ar = f"Dashboard_Data!$A$2:$A${self.dashboard_data_last_row}"
                cur = f"({sumifs_formula(f'Dashboard_Data!${cc}$2:${cc}${self.dashboard_data_last_row}', ar, other)})/1000"
                ly = f"({sumifs_formula(f'Dashboard_Data!${lc}$2:${lc}${self.dashboard_data_last_row}', ar, other)})/1000"
                days = f"MAX({','.join([f'SUMIFS(Dashboard_Data!${dc}$2:${dc}${self.dashboard_data_last_row},{ar},"{a}",{other})' for a in area_values])})"
                obj = f"({sumifs_formula(f'Dashboard_Data!${oc}$2:${oc}${self.dashboard_data_last_row}', ar, other)})/1000"
            ws.cell(idx, 2, f"={cur}")
            ws.cell(idx, 3, f"={ly}")
            ws.cell(idx, 4, f"=B{idx}-C{idx}")
            ws.cell(idx, 5, f"=IFERROR(D{idx}/C{idx},0)")
            ws.cell(idx, 6, f"={obj}")
            ws.cell(idx, 7, f"=B{idx}-F{idx}")
            ws.cell(idx, 8, f"=IFERROR(G{idx}/F{idx},0)")
            ws.cell(idx, 9, f"=IFERROR(B{idx}/J{idx},0)")
            ws.cell(idx, 10, f"={days}")
            for c in range(1, 11):
                ws.cell(idx, c).border = self.border
                if c > 1: ws.cell(idx, c).alignment = Alignment(horizontal="center")
            for c in (2, 3, 4, 6, 7, 9): ws.cell(idx, c).number_format = '#,##0.00' if product == 'Lubricants' else '#,##0'
            for c in (5, 8): ws.cell(idx, c).number_format = '0.00%'
            ws.cell(idx, 10).number_format = '0'
        total = header_row + len(PRODUCTS) + 1
        ws.cell(total, 1, "Grand Total")
        for c in (2, 3, 4, 6, 7):
            ws.cell(total, c, f"=SUM({get_column_letter(c)}{header_row+1}:{get_column_letter(c)}{total-1})")
        ws.cell(total, 5, f"=IFERROR(D{total}/C{total},0)")
        ws.cell(total, 8, f"=IFERROR(G{total}/F{total},0)")
        ws.cell(total, 9, f"=IFERROR(B{total}/J{total},0)")
        ws.cell(total, 10, f"=MAX(J{header_row+1}:J{total-1})")
        for c in range(1, 11):
            ws.cell(total, c).fill = self.light_fill; ws.cell(total, c).font = self.bold; ws.cell(total, c).border = self.border
        for c in (2, 3, 4, 6, 7, 9): ws.cell(total, c).number_format = '#,##0.00'
        for c in (5, 8): ws.cell(total, c).number_format = '0.00%'
        return total

    def _build_performance_dashboard(self, ws, area: str = "ALL", ddm_name: str | None = None) -> None:
        areas_for_view = self._area_list(area)
        title = (f"DDM {ddm_name} — SALES DASHBOARD" if ddm_name else
                 ("Karachi Division Sales Dashboard" if area == "ALL" else f"AREA {area} — SALES DASHBOARD"))
        self._title(ws, title, 10)
        min_date = self.processor.start_date.date()
        fiscal_start_year = min_date.year if min_date.month >= 7 else min_date.year - 1
        fiscal_end = date(fiscal_start_year + 1, 6, 30)
        ws["A2"] = "Select As-of Date"; ws["A2"].font = self.bold
        dashboard_today = date.today() - timedelta(days=1)
        ws["B2"] = dashboard_today
        date_dv = DataValidation(type="date", operator="between", formula1=f"DATE({min_date.year},{min_date.month},{min_date.day})", formula2=f"DATE({dashboard_today.year},{dashboard_today.month},{dashboard_today.day})", allow_blank=False)
        ws.add_data_validation(date_dv); date_dv.add(ws["B2"])
        ws["B2"].number_format = "dd-mmm-yyyy"; ws["B2"].fill = PatternFill("solid", fgColor="FFF2CC"); ws["B2"].font = self.bold
        ws.merge_cells("C2:J2")
        ws["C2"] = ((f"DDM areas: {', '.join(areas_for_view)}. " if ddm_name else "") +
                    "Sales in KLs")
        ws["C2"].font = Font(italic=True, color="666666"); ws["C2"].alignment = Alignment(wrap_text=True)

        # DDM dashboards are focused views. Do not repeat the category and area
        # navigation block from the main dashboard; it made the DDM sheet look
        # like another full division dashboard. Keep only a simple return link.
        if ddm_name:
            ws["K2"] = "← Main Dashboard"
            ws["K2"].hyperlink = "#'Dashboard'!A1"
            ws["K2"].style = "Hyperlink"
            first_table_row = 4
        else:
            self._division_buttons(ws, area, 3)
            ws["A4"] = "AREA DASHBOARDS "; ws["A4"].font = Font(bold=True, size=11)
            for col, item in enumerate(AREAS, 2):
                cell = ws.cell(4, col, item); cell.hyperlink = f"#'Area_{item}'!A1"; cell.fill = self.area_fill
                cell.font = Font(bold=True, color="0563C1", underline="single"); cell.alignment = Alignment(horizontal="center"); cell.border = self.border
            if area == "ALL" and self.ddm_areas:
                ws["L3"] = "DDM DASHBOARDS — CLICK NAME"
                ws["L3"].font = Font(bold=True, size=11)
                for offset, ddm in enumerate(sorted(self.ddm_areas), 12):
                    c = ws.cell(4, offset, ddm)
                    c.hyperlink = f"#'{self._ddm_sheet_name(ddm)}'!A1"
                    c.fill = PatternFill("solid", fgColor="FFD966")
                    c.font = Font(bold=True, color="0563C1", underline="single")
                    c.alignment = Alignment(horizontal="center")
                    c.border = self.border
                    ws.column_dimensions[get_column_letter(offset)].width = max(12, len(ddm) + 3)
            if area != "ALL":
                ws["K2"] = "← Main Dashboard"; ws["K2"].hyperlink = "#'Dashboard'!A1"; ws["K2"].style = "Hyperlink"
            first_table_row = 6

        fy_total = self._dashboard_table(ws, first_table_row, "1. YTD_base", "FY", area, "$B$2")
        mtd_title = '="2. MTD_base "&TEXT($B$2,"mmmm")'
        mtd_total = self._dashboard_table(ws, fy_total + 2, mtd_title, "MTD", area, "$B$2")
        selected_title = mtd_total + 2
        ws.cell(selected_title, 1, "3. Previous Month Analysis").font = Font(bold=True, size=12)
        ws.cell(selected_title + 1, 1, "Select Month:").font = self.bold
        ws.cell(selected_title + 1, 2, date.today().strftime("%B"))
        dv = DataValidation(type="list", formula1='"July,August,September,October,November,December,January,February,March,April,May,June"', allow_blank=False)
        ws.add_data_validation(dv); dv.add(ws.cell(selected_title + 1, 2))
        selected_month_total = self._dashboard_table(ws, selected_title + 3, "", "SEL", area, "$B$2", selected=True, month_cell=f"$B${selected_title+1}")

        quarter_title = selected_month_total + 2
        ws.cell(quarter_title, 1, "4. QUARTER ANALYSIS").font = Font(bold=True, size=12)
        ws.cell(quarter_title + 1, 1, "Select Quarter:").font = self.bold
        current_month = date.today().month
        default_quarter = "Q1" if current_month in (7, 8, 9) else "Q2" if current_month in (10, 11, 12) else "Q3" if current_month in (1, 2, 3) else "Q4"
        ws.cell(quarter_title + 1, 2, default_quarter)
        qdv = DataValidation(type="list", formula1='"Q1,Q2,Q3,Q4"', allow_blank=False)
        ws.add_data_validation(qdv); qdv.add(ws.cell(quarter_title + 1, 2))
        quarter_total = self._dashboard_table(ws, quarter_title + 3, "Quarter Product Comparison", "QTR", area, "$B$2", quarter=True, quarter_cell=f"$B${quarter_title+1}")

        prev_row = quarter_total + 2
        ws.cell(prev_row,1,"5. PREVIOUS 6 COMPLETE MONTHS — AVERAGE SALES").font=Font(bold=True,size=12)
        ws.cell(prev_row+1,1,"Product"); ws.cell(prev_row+1,2,"Average Sales")
        for c in (1,2): ws.cell(prev_row+1,c).fill=self.dark_fill; ws.cell(prev_row+1,c).font=self.white_bold; ws.cell(prev_row+1,c).border=self.border
        for i,product in enumerate(PRODUCTS,prev_row+2):
            ws.cell(i,1,product); ws.cell(i,2,sum(self.processor.previous_six_month_average(product,a) for a in areas_for_view)/1000)
            ws.cell(i,2).number_format='#,##0.00'; ws.cell(i,1).border=self.border; ws.cell(i,2).border=self.border
        # Separate current-month average and trend tables requested for the dashboard.
        avg_start = prev_row + len(PRODUCTS) + 3
        ws.cell(avg_start, 1, "6. CURRENT MONTH PRODUCT-WISE DAILY AVERAGE").font = Font(bold=True, size=12)
        avg_headers = ["Product", "This Month Avg / Day", "Last Year Avg / Day", "Trend"]
        for c, h in enumerate(avg_headers, 1):
            cell = ws.cell(avg_start + 1, c, h); cell.fill = self.dark_fill; cell.font = self.white_bold; cell.border = self.border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        mtd_first_product = (fy_total + 2) + 2
        for offset, product in enumerate(PRODUCTS):
            r = avg_start + 2 + offset
            src = mtd_first_product + offset
            ws.cell(r, 1, product)
            ws.cell(r, 2, f'=I{src}')
            ws.cell(r, 3, f'=Dashboard_Data!K2*0+IFERROR(C{src}/J{src},0)')
            ws.cell(r, 4, f'=IF(B{r}>=C{r},"▲ Increasing","▼ Decreasing")')
            for c in range(1, 5): ws.cell(r, c).border = self.border
            for c in (2, 3): ws.cell(r, c).number_format = '#,##0.00'
        green_fill = PatternFill("solid", fgColor="E2F0D9")
        red_fill = PatternFill("solid", fgColor="FCE4D6")
        ws.conditional_formatting.add(f'D{avg_start+2}:D{avg_start+5}', FormulaRule(formula=[f'B{avg_start+2}>=C{avg_start+2}'], fill=green_fill, font=Font(color="008000", bold=True)))
        ws.conditional_formatting.add(f'D{avg_start+2}:D{avg_start+5}', FormulaRule(formula=[f'B{avg_start+2}<C{avg_start+2}'], fill=red_fill, font=Font(color="C00000", bold=True)))

        compare_start = avg_start
        ws.cell(compare_start, 6, "7. AREA-WISE LAST YEAR MONTH SALES AND THIS MONTH OBJECTIVE").font = Font(bold=True, size=12)
        ws.cell(compare_start + 1, 6, "Select Product").fill = self.light_fill
        ws.cell(compare_start + 1, 6).font = self.bold
        ws.cell(compare_start + 1, 6).border = self.border
        ws.cell(compare_start + 1, 7, "PMG")
        ws.cell(compare_start + 1, 7).fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(compare_start + 1, 7).font = self.bold
        ws.cell(compare_start + 1, 7).border = self.border
        product_selector = DataValidation(type="list", formula1='"HSD,PMG,R95,Lubricants"', allow_blank=False)
        ws.add_data_validation(product_selector)
        product_selector.add(ws.cell(compare_start + 1, 7))

        comp_header_row = compare_start + 2
        comp_headers = ["Area", "Last Year Full Month Sales", "This Month Objective"]
        for c, h in enumerate(comp_headers, 6):
            cell = ws.cell(comp_header_row, c, h); cell.fill = self.dark_fill; cell.font = self.white_bold; cell.border = self.border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Hidden helper table used by the product selector. Values are stored in
        # source units and displayed in thousands to match the dashboard.
        helper_col = 53  # BA
        helper_headers = ["Product", "Area", "Month", "LY Full Month", "Objective"]
        for offset, header in enumerate(helper_headers):
            ws.cell(1, helper_col + offset, header)
        helper_row = 2
        current_month_name = self.processor.end_date.strftime("%B") if self.processor.end_date else date.today().strftime("%B")
        last_month_year = (self.processor.end_date.year - 1) if self.processor.end_date else (date.today().year - 1)
        current_month_no = self.processor.end_date.month if self.processor.end_date else date.today().month
        for helper_product in PRODUCTS:
            for helper_area in AREAS:
                ly_value = self.processor.monthly_area_product_totals["last"].get(
                    (helper_product, helper_area, last_month_year, current_month_no), 0.0
                )
                objective_value = self.processor.objectives.get((helper_product, helper_area, current_month_name), 0.0)
                values = [helper_product, helper_area, current_month_name, ly_value, objective_value]
                for offset, value in enumerate(values):
                    ws.cell(helper_row, helper_col + offset, value)
                helper_row += 1
        helper_last = helper_row - 1

        for offset, item_area in enumerate(areas_for_view):
            r = comp_header_row + 1 + offset
            ws.cell(r, 6, item_area)
            ws.cell(r, 7, f'=SUMIFS($BD$2:$BD${helper_last},$BA$2:$BA${helper_last},$G${compare_start+1},$BB$2:$BB${helper_last},F{r})/1000')
            ws.cell(r, 8, f'=SUMIFS($BE$2:$BE${helper_last},$BA$2:$BA${helper_last},$G${compare_start+1},$BB$2:$BB${helper_last},F{r})/1000')
            for c in range(6, 9):
                ws.cell(r, c).border = self.border
                ws.cell(r, c).alignment = Alignment(horizontal="center")
            for c in (7, 8): ws.cell(r, c).number_format = '#,##0.00'
        for c in range(helper_col, helper_col + len(helper_headers)):
            ws.column_dimensions[get_column_letter(c)].hidden = True

        # Daily sales trend is intentionally shown only on the Analytics sheet.

        self._add_variance_formatting(ws,[f"D{first_table_row+2}:D{quarter_total}",f"E{first_table_row+2}:E{quarter_total}",f"G{first_table_row+2}:G{quarter_total}",f"H{first_table_row+2}:H{quarter_total}"])
        widths = [20, 18, 20, 18, 14, 18, 18, 16, 18, 14]
        for c, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = f"A{first_table_row}"; ws.sheet_view.showGridLines = False

    def _create_lubricant_data(self) -> None:
        ws = self.wb.create_sheet("Lubricant_Data")
        ws.append(["Area", "Period", "Date", "Mat Group Text", "Sales"])
        groups = self.processor.lubricant_materials()
        for period in ("last", "this"):
            source = self.processor.material_group_daily[period]
            for (group, area, day), value in sorted(source.items(), key=lambda x: (x[0][2], x[0][0], x[0][1])):
                ws.append([area, period, day, group, value])
                ws.append(["ALL", period, day, group, value])
        self.lubricant_data_last_row = ws.max_row
        for cell in ws[1]:
            cell.fill = self.dark_fill; cell.font = self.white_bold
        ws.sheet_state = "hidden"

    @staticmethod
    def _excel_date_expr(cell_ref: str) -> str:
        return cell_ref

    def _lubricant_table(self, ws, title_row: int, title: str, area: str, asof_cell: str,
                   mode: str, selector_cell: str | None = None) -> int:
        ws.cell(title_row, 1, title).font = Font(bold=True, size=12)
        header_row = title_row + 1
        headers = ["Product", "TY_Sales", "LY_Sales", "Var_LY", "Var_LY%",
                   "Obj", "TY_vs_Obj", "TY_vs_Obj%", "Avg/Day", "Working Days"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(header_row, c, h); cell.fill = self.dark_fill; cell.font = self.white_bold
            cell.border = self.border; cell.alignment = Alignment(horizontal="center", wrap_text=True)
        groups = self.processor.lubricant_materials()
        fiscal_start_year = self.processor.start_date.year if self.processor.start_date.month >= 7 else self.processor.start_date.year - 1
        fy_start = f"DATE({fiscal_start_year},7,1)"
        ly_fy_start = f"DATE({fiscal_start_year-1},7,1)"
        ly_fy_end = f"DATE({fiscal_start_year},6,30)"
        for i, group in enumerate(groups, header_row + 1):
            ws.cell(i, 1, group)
            if mode == "FY":
                cur_start, cur_end = fy_start, asof_cell
                ly_start, ly_end = ly_fy_start, ly_fy_end
            elif mode == "MTD":
                cur_start, cur_end = f"EOMONTH({asof_cell},-1)+1", asof_cell
                ly_start, ly_end = f"EDATE(EOMONTH({asof_cell},-1)+1,-12)", f"EOMONTH(EDATE({asof_cell},-12),0)"
            elif mode == "MONTH":
                month_map = {m:i for i,m in enumerate(["July","August","September","October","November","December","January","February","March","April","May","June"], 0)}
                # selector is converted through MATCH against fiscal month order.
                cur_start = f"EDATE({fy_start},MATCH({selector_cell},{{\"July\",\"August\",\"September\",\"October\",\"November\",\"December\",\"January\",\"February\",\"March\",\"April\",\"May\",\"June\"}},0)-1)"
                cur_end = f"MIN({asof_cell},EOMONTH({cur_start},0))"
                ly_start, ly_end = f"EDATE({cur_start},-12)", f"EOMONTH(EDATE({cur_start},-12),0)"
            else:  # QUARTER
                qoffset = f"(MATCH({selector_cell},{{\"Q1\",\"Q2\",\"Q3\",\"Q4\"}},0)-1)*3"
                cur_start = f"EDATE({fy_start},{qoffset})"
                cur_end = f"MIN({asof_cell},EOMONTH({cur_start},2))"
                ly_start, ly_end = f"EDATE({cur_start},-12)", f"EOMONTH(EDATE({cur_start},-12),2)"
            criteria_cur = f'Lubricant_Data!$A$2:$A${self.lubricant_data_last_row},"{area}",Lubricant_Data!$B$2:$B${self.lubricant_data_last_row},"this",Lubricant_Data!$D$2:$D${self.lubricant_data_last_row},$A{i},Lubricant_Data!$C$2:$C${self.lubricant_data_last_row},">="&{cur_start},Lubricant_Data!$C$2:$C${self.lubricant_data_last_row},"<="&{cur_end}'
            criteria_ly = f'Lubricant_Data!$A$2:$A${self.lubricant_data_last_row},"{area}",Lubricant_Data!$B$2:$B${self.lubricant_data_last_row},"last",Lubricant_Data!$D$2:$D${self.lubricant_data_last_row},$A{i},Lubricant_Data!$C$2:$C${self.lubricant_data_last_row},">="&{ly_start},Lubricant_Data!$C$2:$C${self.lubricant_data_last_row},"<="&{ly_end}'
            days = f'MAX(0,NETWORKDAYS.INTL({cur_start},{cur_end},11))'
            ly_days = f'({ly_end}-{ly_start}+1)'
            ws.cell(i, 2, f'=SUMIFS(Lubricant_Data!$E$2:$E${self.lubricant_data_last_row},{criteria_cur})/1000')
            ws.cell(i, 3, f'=IF({cur_end}<{cur_start},0,SUMIFS(Lubricant_Data!$E$2:$E${self.lubricant_data_last_row},{criteria_ly})/{ly_days}*{days}/1000)')
            ws.cell(i, 4, f'=B{i}-C{i}')
            ws.cell(i, 5, f'=IFERROR(D{i}/C{i},0)')
            for c in (6,7,8): ws.cell(i,c,"N/A")
            ws.cell(i, 9, f'=IFERROR(B{i}/J{i},0)')
            ws.cell(i, 10, f'={days}')
            for c in range(1,11): ws.cell(i,c).border=self.border
            for c in (2,3,4,9): ws.cell(i,c).number_format='#,##0.00'
            ws.cell(i,5).number_format='0.00%'; ws.cell(i,10).number_format='0'
        total = header_row + len(groups) + 1
        ws.cell(total,1,"Grand Total")
        for c in (2,3,4): ws.cell(total,c,f'=SUM({get_column_letter(c)}{header_row+1}:{get_column_letter(c)}{total-1})')
        ws.cell(total,5,f'=IFERROR(D{total}/C{total},0)')
        for c in (6,7,8): ws.cell(total,c,"N/A")
        ws.cell(total,9,f'=IFERROR(B{total}/J{total},0)'); ws.cell(total,10,f'=MAX(J{header_row+1}:J{total-1})')
        for c in range(1,11): ws.cell(total,c).fill=self.light_fill; ws.cell(total,c).font=self.bold; ws.cell(total,c).border=self.border
        for c in (2,3,4,9): ws.cell(total,c).number_format='#,##0.00'
        ws.cell(total,5).number_format='0.00%'
        self._add_variance_formatting(ws,[f'D{header_row+1}:E{total}'])
        return total

    def _build_lubricant_dashboard(self, ws, area: str = "ALL") -> None:
        label = "ALL AREAS" if area == "ALL" else f"AREA {area}"
        self._title(ws, f"LUBRICANTS PERFORMANCE DASHBOARD — {label}", 10)
        min_date=self.processor.start_date.date(); fiscal_start_year=min_date.year if min_date.month>=7 else min_date.year-1
        fiscal_end=date(fiscal_start_year+1,6,30)
        ws['A2']='As-of Date'; ws['A2'].font=self.bold
        dashboard_today=date.today(); ws['B2']=dashboard_today
        date_dv=DataValidation(type='date',operator='between',formula1=f'DATE({min_date.year},{min_date.month},{min_date.day})',formula2=f'DATE({dashboard_today.year},{dashboard_today.month},{dashboard_today.day})',allow_blank=False); ws.add_data_validation(date_dv); date_dv.add(ws['B2']); ws['B2'].number_format='dd-mmm-yyyy'; ws['B2'].fill=PatternFill('solid',fgColor='FFF2CC'); ws['B2'].font=self.bold
        self._division_buttons(ws, area, 3)
        ws['A4']='AREA LUBRICANT DASHBOARDS'; ws['A4'].font=Font(bold=True,size=11)
        for col,item in enumerate(AREAS,2):
            c=ws.cell(4,col,item); c.hyperlink=f"#'Lubricant_Area_{item}'!A1"; c.fill=self.area_fill; c.font=Font(bold=True,color='0563C1',underline='single'); c.alignment=Alignment(horizontal='center'); c.border=self.border
        nav_end=self._material_buttons(ws,area,5); first=nav_end+1
        fy=self._lubricant_table(ws,first,'1. FISCAL YEAR-TO-DATE MAT GROUP SUMMARY',area,'$B$2','FY')
        mtd=self._lubricant_table(ws,fy+2,'2. CURRENT MONTH-TO-DATE MAT GROUP SUMMARY',area,'$B$2','MTD')
        sr=mtd+2; ws.cell(sr,1,'3. SELECTED MONTH ANALYSIS').font=Font(bold=True,size=12); ws.cell(sr+1,1,'Select Month:').font=self.bold; ws.cell(sr+1,2,date.today().strftime('%B'))
        dv=DataValidation(type='list',formula1='"July,August,September,October,November,December,January,February,March,April,May,June"'); ws.add_data_validation(dv); dv.add(ws.cell(sr+1,2))
        sel=self._lubricant_table(ws,sr+3,'Selected Month Mat Group Comparison',area,'$B$2','MONTH',f'$B${sr+1}')
        qr=sel+2; ws.cell(qr,1,'4. SELECTED QUARTER ANALYSIS').font=Font(bold=True,size=12); ws.cell(qr+1,1,'Select Quarter:').font=self.bold
        cm=date.today().month; q='Q1' if cm in (7,8,9) else 'Q2' if cm in (10,11,12) else 'Q3' if cm in (1,2,3) else 'Q4'; ws.cell(qr+1,2,q)
        qdv=DataValidation(type='list',formula1='"Q1,Q2,Q3,Q4"'); ws.add_data_validation(qdv); qdv.add(ws.cell(qr+1,2))
        qt=self._lubricant_table(ws,qr+3,'Selected Quarter Mat Group Comparison',area,'$B$2','QUARTER',f'$B${qr+1}')
        pr=qt+2; ws.cell(pr,1,'5. PREVIOUS 6 COMPLETE MONTHS — AVERAGE SALES').font=Font(bold=True,size=12)
        ws.cell(pr+1,1,'Product'); ws.cell(pr+1,2,'Average Sales')
        for c in (1,2): ws.cell(pr+1,c).fill=self.dark_fill; ws.cell(pr+1,c).font=self.white_bold; ws.cell(pr+1,c).border=self.border
        for i,g in enumerate(self.processor.lubricant_materials(),pr+2):
            ws.cell(i,1,g); start=f'EOMONTH($B$2,-7)+1'; end=f'EOMONTH($B$2,-1)'
            crit=f'Lubricant_Data!$A$2:$A${self.lubricant_data_last_row},"{area}",Lubricant_Data!$B$2:$B${self.lubricant_data_last_row},"this",Lubricant_Data!$D$2:$D${self.lubricant_data_last_row},A{i},Lubricant_Data!$C$2:$C${self.lubricant_data_last_row},">="&{start},Lubricant_Data!$C$2:$C${self.lubricant_data_last_row},"<="&{end}'
            ws.cell(i,2,f'=SUMIFS(Lubricant_Data!$E$2:$E${self.lubricant_data_last_row},{crit})/6/1000'); ws.cell(i,2).number_format='#,##0.00'; ws.cell(i,1).border=self.border; ws.cell(i,2).border=self.border
        for c,w in enumerate([20,18,20,18,14,18,18,16,18,14],1): ws.column_dimensions[get_column_letter(c)].width=w
        ws.freeze_panes=f'A{first}'; ws.sheet_view.showGridLines=False

    def create_lubricant_dashboards(self) -> None:
        self._create_lubricant_data()
        self._build_lubricant_dashboard(self.wb.create_sheet('Lubricant_Dashboard',1),'ALL')
        for area in AREAS:
            self._build_lubricant_dashboard(self.wb.create_sheet(f'Lubricant_Area_{area}'),area)

    def _create_lubricant_liter_master_sheet(self) -> None:
        ws=self.wb.create_sheet("Lubricant_Product_Liters")
        self._title(ws,"LUBRICANT PRODUCT LITER MASTER — EDIT FOR FUTURE OPEN/TRANSIT FILES",4)
        ws.append(["Product","Liter","Usage Note"])
        for c in ws[2]: c.fill=self.dark_fill; c.font=self.white_bold; c.border=self.border
        for product,liters in sorted(self.processor.lubricant_liter_master.items()):
            ws.append([product,liters,"EA quantity × Liter; L quantity × 1"])
        for row in ws.iter_rows(min_row=3,max_row=ws.max_row,min_col=1,max_col=3):
            for cell in row: cell.border=self.border
        ws.column_dimensions["A"].width=58; ws.column_dimensions["B"].width=12; ws.column_dimensions["C"].width=34
        ws.freeze_panes="A3"; ws.sheet_view.showGridLines=False

    def _create_lubricant_pipeline_data(self) -> None:
        """Create hidden month-wise source data for the selectable pipeline table."""
        ws = self.wb.create_sheet("Lubricants_Pipeline_Data")
        headers = ["Month", "Area", "Execution", "Open", "Transit", "Total Pipeline",
                   "Last Year Adjusted", "Objective", "Achieved %", "TY vs LY %", "TY vs Objective %"]
        ws.append(headers)
        for month_start in self.processor.lubricant_pipeline_months():
            month_label = month_start.strftime("%B %Y")
            for rec in self.processor.lubricant_area_pipeline_rows(month_start):
                ws.append([
                    month_label, rec["Area"], rec["Execution"] / 1000,
                    rec["Open"] / 1000, rec["Transit"] / 1000,
                    rec["Total"] / 1000, rec["LY Adjusted"] / 1000,
                    rec["Objective"] / 1000, rec["Achieved %"],
                    rec["TY vs LY %"], rec["TY vs Objective %"],
                ])
        self.lubricant_pipeline_data_last_row = ws.max_row
        ws.sheet_state = "hidden"

    def _add_lubricant_area_pipeline_table(self, ws) -> int:
        """Add a selectable month-wise lubricant pipeline table and return its title row."""
        start=ws.max_row+3
        ws.cell(start,1,"Monthwise Lubricant").font=Font(bold=True,size=13)
        ws.cell(start+1,1,"Selected Month")
        default_month = self.processor.lubricant_pipeline_default_month().strftime("%B %Y")
        ws.cell(start+1,2,default_month)
        ws.cell(start+1,1).font=self.bold
        ws.cell(start+1,2).fill=PatternFill("solid",fgColor="FFF2CC")
        ws.cell(start+1,2).border=self.border
        months = [m.strftime("%B %Y") for m in self.processor.lubricant_pipeline_months()]
        month_dv = DataValidation(type="list", formula1='"' + ','.join(months) + '"', allow_blank=False)
        ws.add_data_validation(month_dv); month_dv.add(ws.cell(start+1,2))
        ws.cell(start+2,1,"Sales in KLs")
        ws.merge_cells(start_row=start+2,start_column=1,end_row=start+2,end_column=8)
        ws.cell(start+2,1).font=Font(italic=True,color="666666")

        # Keep this sheet focused on the selected month's position versus objective.
        # Last-year columns are intentionally omitted from the visible table.
        headers=["Area","Execution","Open Orders","Transit","This Month","Objective","Achieved %","This Month vs Objective","This Month vs Objective %"]
        hr=start+3
        for c,h in enumerate(headers,1):
            cell=ws.cell(hr,c,h); cell.fill=self.dark_fill; cell.font=Font(name="Calibri",size=12,bold=True,color="FFFFFF"); cell.border=self.border; cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.row_dimensions[hr].height = 34
        first_data = hr + 1
        # Visible columns B:G use source columns C,D,E,F,H,I respectively.
        # Achieved % is This Month divided by Objective and is retained as a
        # separate KPI from the objective variance percentage.
        source_columns = [3, 4, 5, 6, 8, 9]
        for offset, area in enumerate(AREAS):
            r = first_data + offset
            ws.cell(r,1,area)
            for c, source_col in enumerate(source_columns,2):
                src_letter=get_column_letter(source_col)
                ws.cell(r,c,f'=SUMIFS(Lubricants_Pipeline_Data!${src_letter}$2:${src_letter}${self.lubricant_pipeline_data_last_row},Lubricants_Pipeline_Data!$A$2:$A${self.lubricant_pipeline_data_last_row},$B${start+1},Lubricants_Pipeline_Data!$B$2:$B${self.lubricant_pipeline_data_last_row},$A{r})')
            ws.cell(r,8,f"=E{r}-F{r}")
            ws.cell(r,9,f"=IFERROR(H{r}/F{r},0)")
            for c in range(1,10):
                cell=ws.cell(r,c); cell.border=self.border; cell.alignment=Alignment(horizontal="center")
                cell.number_format="0.00%" if c in (7,9) else ("#,##0.00" if c>1 else "@")
        tr=first_data+len(AREAS); ws.cell(tr,1,"Total")
        for c in range(2,7):
            L=get_column_letter(c); ws.cell(tr,c,f"=SUM({L}{first_data}:{L}{tr-1})"); ws.cell(tr,c).number_format="#,##0.00"
        ws.cell(tr,7,f"=IFERROR(E{tr}/F{tr},0)")
        ws.cell(tr,8,f"=E{tr}-F{tr}")
        ws.cell(tr,9,f"=IFERROR(H{tr}/F{tr},0)")
        for c in range(1,10): ws.cell(tr,c).fill=self.light_fill; ws.cell(tr,c).font=self.bold; ws.cell(tr,c).border=self.border
        ws.cell(tr,7).number_format="0.00%"
        ws.cell(tr,8).number_format="#,##0.00"
        ws.cell(tr,9).number_format="0.00%"
        self._add_variance_formatting(ws,[f"H{first_data}:I{tr}"])
        for c,w in enumerate([10,18,16,14,16,16,15,24,24],1): ws.column_dimensions[get_column_letter(c)].width=max(ws.column_dimensions[get_column_letter(c)].width or 0,w)
        return start

    def _round_main_dashboard_display(self, ws) -> None:
        """Show all visible main-dashboard numeric values as whole numbers."""
        for row in ws.iter_rows(min_col=1, max_col=12):
            for cell in row:
                fmt = str(cell.number_format or "General")
                if "%" in fmt or any(token in fmt.lower() for token in ("dd", "mmm", "yy")):
                    continue
                value = cell.value
                if isinstance(value, Number) or (isinstance(value, str) and value.startswith("=")):
                    product_label = str(ws.cell(cell.row, 1).value or "")
                    cell.number_format = "#,##0.00" if product_label == "Lubricants" else "#,##0"

    def _create_lubricant_pipeline_sheet(self) -> None:
        """Create a dedicated selectable month-wise lubricant pipeline sheet."""
        self._create_lubricant_pipeline_data()
        ws = self.wb.create_sheet("Lubricants_Pipeline", 1)
        self._title(ws, "Monthwise Lubricant", 9)
        ws["A2"] = "← Back to Dashboard"
        ws["A2"].hyperlink = "#'Dashboard'!A1"
        ws["A2"].style = "Hyperlink"
        self._add_lubricant_area_pipeline_table(ws)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A7"

    def _add_main_top_outlet_selector(self, ws) -> None:
        """Dynamic Top-N outlet view on the main dashboard, based on Additional Qty."""
        rows = self.processor.top_outlets(None, limit=50)
        start_col = 14  # N
        ws.cell(2, start_col, "Top Outlets").fill = self.dark_fill
        ws.cell(2, start_col).font = self.white_bold
        ws.cell(2, start_col + 1, 5)
        ws.cell(2, start_col + 1).fill = PatternFill("solid", fgColor="FFF2CC")
        dv = DataValidation(type="list", formula1='"10,20,30,40,50"', allow_blank=False)
        ws.add_data_validation(dv); dv.add(ws.cell(2, start_col + 1))
        headers = ["Rank", "Cost Center", "Outlet Name", "Outlet Code", "Additional Qty (KL)"]
        for i, h in enumerate(headers, start_col):
            c = ws.cell(4, i, h); c.fill=self.dark_fill; c.font=self.white_bold; c.border=self.border; c.alignment=Alignment(horizontal="center")
        for rank in range(1, 51):
            rr = 4 + rank
            rec = rows[rank-1] if rank <= len(rows) else {}
            vals = [rank, rec.get("Cost Center", ""), rec.get("Outlet Name", ""), rec.get("Sold to Party", ""), rec.get("This Year Sales", 0)/1000]
            for j, val in enumerate(vals, start_col):
                c=ws.cell(rr,j, f'=IF(ROW()-4<=$O$2,{val if isinstance(val,(int,float)) else chr(34)+str(val).replace(chr(34),chr(34)*2)+chr(34)},"")')
                c.border=self.border
                if j == start_col + 4: c.number_format = "#,##0.00"
        for col, width in zip(range(start_col,start_col+5), [8,16,28,16,18]):
            ws.column_dimensions[get_column_letter(col)].width=width

    def _enforce_dashboard_variance_headers(self, ws) -> None:
        """Keep every variance/percentage-of-variance heading white on Dashboard."""
        for row in ws.iter_rows():
            for cell in row:
                text = str(cell.value or "").strip().casefold()
                variance_labels = {
                    "variance", "variance %", "var_ly", "var_ly%",
                    "ty_vs_obj", "ty_vs_obj%", "vs objective", "vs objective %",
                    "this month vs objective", "this month vs objective %"
                }
                if ("variance" in text or text in variance_labels) and not text.startswith("="):
                    cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
                    cell.fill = self.dark_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def create_dashboard(self) -> None:
        self._create_dashboard_data()
        ws = self.wb.create_sheet("Dashboard", 0)
        self._build_performance_dashboard(ws, "ALL")
        # Developer credit shown prominently on the MAIN Dashboard.
        ws.merge_cells("K1:N1")
        ws["K1"] = "Developed by: Rida Batool (NED) & M.Arham Shakeel (SSUET)"
        ws["K1"].font = Font(bold=True, italic=True, size=10, color="666666")
        ws["K1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = max(ws.row_dimensions[1].height or 15, 30)
        self._round_main_dashboard_display(ws)
        # Apply heading colors last so no earlier dashboard style can leave
        # Var_LY / TY_vs_Obj headings green.
        self._enforce_dashboard_variance_headers(ws)
        for ddm, ddm_areas in sorted(self.ddm_areas.items()):
            ddm_ws = self.wb.create_sheet(self._ddm_sheet_name(ddm))
            self._build_performance_dashboard(ddm_ws, ddm_areas, ddm_name=ddm)
            self._enforce_dashboard_variance_headers(ddm_ws)
        # Five compact dashboard navigation buttons fitted across A:L.
        navigation_buttons = (
            ("A5:B5", "MONTHLY MATRIX →", "Monthly_Area_Matrix", "70AD47"),
            ("C5:D5", "ANALYTICS →", "Analytics", "4472C4"),
            ("E5:F5", "OUTLET TY/LY →", "Outlet_TY_vs_LY", "5B9BD5"),
            ("G5:H5", "AREA 6M AVG →", "Area_Outlet_6M_Avg", "A5A5A5"),
            ("I5:J5", "LUBRICANTS →", "Lubricants_Pipeline", "ED7D31"),
        )
        for cell_range, label, target, fill in navigation_buttons:
            ws.merge_cells(cell_range)
            start_cell = cell_range.split(":")[0]
            button = ws[start_cell]
            button.value = label
            button.hyperlink = (f"#'{target.split('!', 1)[0]}'!{target.split('!', 1)[1]}"
                                if "!" in target else f"#'{target}'!A1")
            button.fill = PatternFill("solid", fgColor=fill)
            button.font = Font(bold=True, size=9, color="FFFFFF", underline="single")
            button.alignment = Alignment(horizontal="center", vertical="center")
            button.border = self.border
        ws.row_dimensions[5].height = 20

    def _daily_total_values(self, rows: list[dict[str, Any]]) -> dict[Any, float]:
        """Return totals using the same rounded thousand-litre values shown in Excel."""
        totals: dict[Any, float] = {}
        for day in range(1, 32):
            totals[day] = sum(round(float(self._display_number(r.get(day, 0) or 0))) for r in rows)

        totals["This Year Sales"] = sum(
            round(float(self._display_number(r.get("This Year Sales", 0) or 0))) for r in rows
        )
        totals["Last Year Sales"] = sum(
            round(float(self._display_number(r.get("Last Year Sales", 0) or 0))) for r in rows
        )
        totals["TY vs LY Variance"] = totals["This Year Sales"] - totals["Last Year Sales"]
        ly = totals["Last Year Sales"]
        totals["TY vs LY %"] = totals["TY vs LY Variance"] / ly if ly else 0.0
        totals["Objective"] = sum(round(float(self._display_number(r.get("Objective", 0) or 0))) for r in rows)
        totals["TY vs Objective Variance"] = totals["This Year Sales"] - totals["Objective"]
        obj = totals["Objective"]
        totals["TY vs Objective %"] = totals["TY vs Objective Variance"] / obj if obj else 0.0
        return totals

    def _write_daily_total_row(
        self,
        ws,
        excel_row: int,
        label: str,
        rows: list[dict[str, Any]],
        columns: list[Any],
        grand: bool = False,
    ) -> None:
        totals = self._daily_total_values(rows)
        fill = self.light_fill if grand else PatternFill("solid", fgColor="FFF2CC")

        for col, name in enumerate(columns, 1):
            cell = ws.cell(excel_row, col)
            cell.fill = fill
            cell.font = Font(bold=True, color="000000")
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if col == 1:
                cell.value = label
            elif name == "Sales Group":
                cell.value = ""
            elif name == "Month":
                # Keep each monthly total visible when Excel filters to one month.
                # A blank Month value causes the total row to be hidden by the filter.
                cell.value = "" if grand else label.removesuffix(" Total")
            elif name in {"TY vs LY %", "TY vs Objective %"}:
                cell.value = totals.get(name, 0)
                cell.number_format = "0.00%"
            elif isinstance(name, int) or name in {
                "This Year Sales", "Last Year Sales", "TY vs LY Variance", "Objective", "TY vs Objective Variance"
            }:
                cell.value = totals.get(name, 0)
                cell.number_format = "#,##0"
            else:
                cell.value = ""
        ws.row_dimensions[excel_row].height = 22 if not grand else 25

    def _create_control_data_sheet(self) -> None:
        """Create a simple hidden lookup sheet used by all month/day control sheets."""
        ws = self.wb.create_sheet("Control_Data")
        headers = ["Key", "Product", "Area", "Month", "Cost Center", "Outlet Name", "Sold to Party", "Objective", "Days In Month"]
        headers += [f"TY {d}" for d in range(1, 32)] + [f"LY {d}" for d in range(1, 32)]
        ws.append(headers)
        row_no = 2
        for product in PRODUCTS:
            for rec in self.processor.daily_detail(product):
                area = str(rec.get("Sales Group", ""))
                month = str(rec.get("Month", ""))
                cost_center = str(rec.get("Cost Center", ""))
                key = f"{product}|{area}|{month}|{cost_center}"
                try:
                    month_num = list(calendar.month_name).index(month)
                    fy_start = self.processor.start_date.year if self.processor.start_date.month >= 7 else self.processor.start_date.year - 1
                    year = fy_start if month_num >= 7 else fy_start + 1
                    dim = calendar.monthrange(year, month_num)[1]
                except Exception:
                    dim = 31
                values = [key, product, area, month, cost_center, rec.get("Outlet Name", ""), rec.get("Sold to Party", ""), rec.get("Objective", 0), dim]
                values += [rec.get(d, 0) for d in range(1, 32)]
                values += [rec.get(f"LY {d}", 0) for d in range(1, 32)]
                ws.append(values)
                row_no += 1
        self.control_data_last_row = ws.max_row
        ws.sheet_state = "hidden"

    def _write_daily_detail(self, ws, rows: list[dict[str, Any]], start_row: int, title: str) -> None:
        columns: list[Any] = ["Cost Center", "Outlet Name", "Sold to Party", "Sales Group", "Month"] + list(range(1, 32)) + ["TY_Sales", "LY_Sales", "Var_LY", "Var_LY%", "Obj", "TY_vs_Obj", "TY_vs_Obj%"]

        # Month and day controls. These are plain data-validation cells and standard
        # INDEX/MATCH formulas only, to remain compatible with desktop Excel.
        ws.cell(start_row, 1, "MONTH & DAY RANGE FILTER").fill = self.dark_fill
        ws.cell(start_row, 1).font = self.white_bold
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)

        control_row = start_row + 1
        ws.cell(control_row, 1, "Month").font = self.bold
        ws.cell(control_row, 2, self.processor.end_date.strftime("%B"))
        ws.cell(control_row, 4, "From Day").font = self.bold
        ws.cell(control_row, 5, 1)
        ws.cell(control_row, 6, "To Day").font = self.bold
        ws.cell(control_row, 7, self.processor.end_date.day)
        ws.cell(control_row, 9, "Sales in KLs")
        for c in (1,2,4,5,6,7):
            ws.cell(control_row,c).border = self.border
            ws.cell(control_row,c).fill = self.area_fill if c in (1,4,6) else PatternFill("solid", fgColor="FFF2CC")

        month_dv = DataValidation(type="list", formula1='"July,August,September,October,November,December,January,February,March,April,May,June"', allow_blank=False)
        day_dv = DataValidation(type="whole", operator="between", formula1="1", formula2="31", allow_blank=False)
        ws.add_data_validation(month_dv); month_dv.add(ws.cell(control_row,2))
        ws.add_data_validation(day_dv); day_dv.add(ws.cell(control_row,5)); day_dv.add(ws.cell(control_row,7))

        ws.cell(start_row + 3, 1, title).font = Font(bold=True, size=12)
        header = start_row + 4
        for col, name in enumerate(columns, 1):
            cell = ws.cell(header, col, str(name))
            cell.fill = self.dark_fill; cell.font = self.white_bold; cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # One row per outlet; month selection changes all daily and summary values.
        unique = {}
        for rec in rows:
            cc = str(rec.get("Cost Center", ""))
            if cc not in unique:
                unique[cc] = rec
        display_rows = sorted(unique.values(), key=lambda r: (str(r.get("Sales Group", "")), str(r.get("Outlet Name", ""))))
        data_last = self.control_data_last_row
        ty_range = f"Control_Data!$J$2:$AN${data_last}"
        ly_range = f"Control_Data!$AO$2:$BS${data_last}"
        key_range = f"Control_Data!$A$2:$A${data_last}"
        obj_range = f"Control_Data!$H$2:$H${data_last}"
        dim_range = f"Control_Data!$I$2:$I${data_last}"

        first_data = header + 1
        for i, rec in enumerate(display_rows):
            r = first_data + i
            cc = str(rec.get("Cost Center", "")); area = str(rec.get("Sales Group", ""))
            ws.cell(r,1,cc); ws.cell(r,2,rec.get("Outlet Name", "")); ws.cell(r,3,rec.get("Sold to Party", "")); ws.cell(r,4,area)
            ws.cell(r,5, f"=$B${control_row}")
            key_expr = f'"{ws.title.split("_")[0]}|"&$D{r}&"|"&$B${control_row}&"|"&$A{r}'
            match_expr = f'MATCH({key_expr},{key_range},0)'
            product_name = ws.title.split("_")[0]
            qty_fmt = "#,##0.00" if product_name == "Lubricants" else "#,##0"
            for day in range(1,32):
                c = 5 + day
                formula = f'=IF(OR({day}<$E${control_row},{day}>$G${control_row}),"",IFERROR(INDEX({ty_range},{match_expr},{day})/1000,0))'
                ws.cell(r,c,formula).number_format = qty_fmt
            ty_col, ly_col, var_col, pct_col, obj_col, objv_col, objp_col = 37,38,39,40,41,42,43
            ws.cell(r,ty_col,f'=SUM(F{r}:AJ{r})').number_format = qty_fmt
            ws.cell(r,ly_col,f'=IFERROR(SUM(INDEX({ly_range},{match_expr},$E${control_row}):INDEX({ly_range},{match_expr},$G${control_row}))/1000,0)').number_format = qty_fmt
            ws.cell(r,var_col,f'=AK{r}-AL{r}').number_format = qty_fmt
            ws.cell(r,pct_col,f'=IFERROR(AM{r}/AL{r},0)').number_format = "0.00%"
            ws.cell(r,obj_col,f'=IFERROR(INDEX({obj_range},{match_expr})/1000*($G${control_row}-$E${control_row}+1)/INDEX({dim_range},{match_expr}),0)').number_format = qty_fmt
            ws.cell(r,objv_col,f'=AK{r}-AO{r}').number_format = qty_fmt
            ws.cell(r,objp_col,f'=IFERROR(AP{r}/AO{r},0)').number_format = "0.00%"
            for c in range(1,len(columns)+1):
                ws.cell(r,c).border=self.border
                if c>1: ws.cell(r,c).alignment=Alignment(horizontal="center")

        last_data = first_data + len(display_rows) - 1
        total_row = last_data + 1
        for c,name in enumerate(columns,1):
            cell=ws.cell(total_row,c); cell.fill=self.light_fill; cell.font=Font(bold=True,color="000000"); cell.border=self.border
            if c==1:
                cell.value="Grand Total"
            elif 6 <= c <= 43:
                # When an area has no outlet rows, first_data > last_data.
                # A formula such as =SUM(F9:F8) is interpreted by Excel as a
                # range containing F9 itself, which creates a circular reference.
                # Empty sections therefore get numeric zero totals instead.
                if display_rows:
                    letter=get_column_letter(c); cell.value=f'=SUM({letter}{first_data}:{letter}{last_data})'
                else:
                    cell.value=0
                cell.number_format="0.00%" if c in (40,43) else "#,##0"
        # Recalculate percentages correctly at grand-total level only when data exists.
        if display_rows:
            ws.cell(total_row,40,f'=IFERROR(AM{total_row}/AL{total_row},0)').number_format="0.00%"
            ws.cell(total_row,43,f'=IFERROR(AP{total_row}/AO{total_row},0)').number_format="0.00%"
        else:
            ws.cell(total_row,40,0).number_format="0.00%"
            ws.cell(total_row,43,0).number_format="0.00%"

        ws.auto_filter.ref=f"A{header}:AQ{last_data}"
        self._add_variance_formatting(ws,[f"AM{first_data}:AN{total_row}",f"AP{first_data}:AQ{total_row}"])
        ws.freeze_panes=f"F{first_data}"
        ws.column_dimensions["A"].width=16; ws.column_dimensions["B"].width=38; ws.column_dimensions["C"].width=16; ws.column_dimensions["D"].width=12; ws.column_dimensions["E"].width=12
        for c in range(6,37): ws.column_dimensions[get_column_letter(c)].width=6
        for c in range(37,44): ws.column_dimensions[get_column_letter(c)].width=18
        ws.sheet_view.showGridLines=False

    def create_product_sheet(self, product: str) -> None:
        ws = self.wb.create_sheet(product)
        self._title(ws, f"{product} SALES DASHBOARD", 9)
        ws["A2"] = "← Back to Dashboard"
        ws["A2"].hyperlink = "#'Dashboard'!A1"
        ws["A2"].style = "Hyperlink"
        daily_start = self._horizontal_area_links(ws, product, 4)
        self._write_daily_detail(ws, self.processor.daily_detail(product), daily_start, "OUTLET-WISE DAILY SALES")

    def create_product_area_sheet(self, product: str, area: str) -> None:
        ws = self.wb.create_sheet(f"{product}_{area}")
        self._title(ws, f"{product} — AREA {area} DAILY SALES", 9)
        ws["A2"] = f"← Back to {product}"
        ws["A2"].hyperlink = f"#'{product}'!A1"
        ws["A2"].style = "Hyperlink"
        ws["C2"] = "← Back to Dashboard"
        ws["C2"].hyperlink = "#'Dashboard'!A1"
        ws["C2"].style = "Hyperlink"
        rows = self.processor.daily_detail(product, area)
        self._write_daily_detail(ws, rows, 4, f"{product} — {area} OUTLET-WISE DAILY SALES")

    def create_area_sheet(self, area: str) -> None:
        ws = self.wb.create_sheet(f"Area_{area}")
        self._build_performance_dashboard(ws, area)

    def create_material_sheet(self, material: str, area: str = "ALL") -> None:
        name = self._material_sheet_name(material, area)
        if name in self.wb.sheetnames:
            return
        ws = self.wb.create_sheet(name)
        area_title = "ALL AREAS" if area == "ALL" else f"AREA {area}"
        self._title(ws, f"LUBRICANT GROUP — {material} — {area_title}", 10)

        back_sheet = "Lubricant_Dashboard" if area == "ALL" else f"Lubricant_Area_{area}"
        ws["A2"] = f"← Back to {back_sheet}"
        ws["A2"].hyperlink = f"#'{back_sheet}'!A1"
        ws["A2"].style = "Hyperlink"

        # First show a dashboard-style Material Text summary for the selected
        # Mat Group Text and area. Material-level targets are not present in the
        # source targets workbook, so target columns are explicitly marked N/A.
        summary_rows = self.processor.lubricant_material_text_summary(
            material, None if area == "ALL" else area
        )
        fiscal_start = self.processor.start_date.date()
        fiscal_start = date(fiscal_start.year if fiscal_start.month >= 7 else fiscal_start.year - 1, 7, 1)
        fiscal_end = date(fiscal_start.year + 1, 6, 30)

        title_row = 4
        ws.cell(title_row, 1, f"1. {area_title} — MATERIAL TEXT SUMMARY").font = Font(bold=True, size=12)
        headers = [
            "Material Text", "Total Sales", "LY Sales", "Variance", "Variance %",
            "Objective / Target", "Vs Objective", "Vs Objective %",
            "Current Avg / Day", "Working Days",
        ]
        header_row = title_row + 1
        area_values = self._area_list(area)
        def sumifs_formula(sum_range: str, area_range: str, other_criteria: str) -> str:
            parts = [f'SUMIFS({sum_range},{area_range},"{a}",{other_criteria})' for a in area_values]
            return "+".join(parts) if parts else "0"
        for c, h in enumerate(headers, 1):
            cell = ws.cell(header_row, c, h)
            cell.fill = self.dark_fill
            cell.font = self.white_bold
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        first_summary_row = header_row + 1
        for ri, rec in enumerate(summary_rows, first_summary_row):
            ty = float(rec.get("This Year Sales", 0) or 0) / 1000
            ly = float(rec.get("Last Year Sales", 0) or 0) / 1000
            values = [rec.get("Material Text", ""), ty, ly, ty - ly, (ty - ly) / ly if ly else 0,
                      "N/A", "N/A", "N/A", None, None]
            for c, value in enumerate(values, 1):
                cell = ws.cell(ri, c, value)
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center", wrap_text=True)
            ws.cell(ri, 9, f'=IFERROR(B{ri}/J{ri},0)')
            ws.cell(ri, 10, f'=NETWORKDAYS.INTL(DATE({fiscal_start.year},7,1),MIN(TODAY(),DATE({fiscal_end.year},6,30)),11)')
            for c in (2, 3, 4, 9):
                ws.cell(ri, c).number_format = "#,##0.00"
            ws.cell(ri, 5).number_format = "0.00%"
            ws.cell(ri, 10).number_format = "0"

        total_row = first_summary_row + len(summary_rows)
        ws.cell(total_row, 1, "Grand Total")
        if summary_rows:
            ws.cell(total_row, 2, f"=SUM(B{first_summary_row}:B{total_row-1})")
            ws.cell(total_row, 3, f"=SUM(C{first_summary_row}:C{total_row-1})")
        else:
            ws.cell(total_row, 2, 0)
            ws.cell(total_row, 3, 0)
        ws.cell(total_row, 4, f"=B{total_row}-C{total_row}")
        ws.cell(total_row, 5, f"=IFERROR(D{total_row}/C{total_row},0)")
        for c in (6, 7, 8):
            ws.cell(total_row, c, "N/A")
        ws.cell(total_row, 10, f'=NETWORKDAYS.INTL(DATE({fiscal_start.year},7,1),MIN(TODAY(),DATE({fiscal_end.year},6,30)),11)')
        ws.cell(total_row, 9, f"=IFERROR(B{total_row}/J{total_row},0)")
        for c in range(1, 11):
            ws.cell(total_row, c).fill = self.light_fill
            ws.cell(total_row, c).font = self.bold
            ws.cell(total_row, c).border = self.border
        for c in (2, 3, 4, 9):
            ws.cell(total_row, c).number_format = "#,##0.00"
        ws.cell(total_row, 5).number_format = "0.00%"
        ws.cell(total_row, 10).number_format = "0"
        self._add_variance_formatting(ws, [f"D{first_summary_row}:D{total_row}", f"E{first_summary_row}:E{total_row}"])

        # Then show the full outlet/material detail table beneath the summary.
        detail_title_row = total_row + 3
        ws.cell(detail_title_row, 1, f"2. {area_title} — OUTLET & MATERIAL TEXT DETAIL").font = Font(bold=True, size=12)
        detail_cols = ["Cost Center", "Outlet Name", "Sold to Party", "Material Text", "This Year Sales", "Last Year Sales", "TY vs LY Variance", "TY vs LY %"]
        detail_header = detail_title_row + 1
        detail_rows = self.processor.lubricant_material_summary(material, None if area == "ALL" else area)
        for c, h in enumerate(detail_cols, 1):
            cell = ws.cell(detail_header, c, h)
            cell.fill = self.dark_fill
            cell.font = self.white_bold
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        first_detail_row = detail_header + 1
        for ri, rec in enumerate(detail_rows, first_detail_row):
            for c, h in enumerate(detail_cols, 1):
                value = rec.get(h, "")
                if h in ("This Year Sales", "Last Year Sales", "TY vs LY Variance") and isinstance(value, (int, float)):
                    value = value / 1000
                cell = ws.cell(ri, c, value)
                cell.border = self.border
                if c > 1:
                    cell.alignment = Alignment(horizontal="center")
                if h.endswith("%"):
                    cell.number_format = "0.00%"
                elif h in ("This Year Sales", "Last Year Sales", "TY vs LY Variance"):
                    cell.number_format = "#,##0.00"

        detail_total = first_detail_row + len(detail_rows)
        ws.cell(detail_total, 1, "Grand Total")
        if detail_rows:
            ws.cell(detail_total, 4, f"=SUM(D{first_detail_row}:D{detail_total-1})")
            ws.cell(detail_total, 5, f"=SUM(E{first_detail_row}:E{detail_total-1})")
        else:
            ws.cell(detail_total, 4, 0)
            ws.cell(detail_total, 5, 0)
        ws.cell(detail_total, 6, f"=D{detail_total}-E{detail_total}")
        ws.cell(detail_total, 7, f"=IFERROR(F{detail_total}/E{detail_total},0)")
        for c in range(1, 8):
            ws.cell(detail_total, c).fill = self.light_fill
            ws.cell(detail_total, c).font = self.bold
            ws.cell(detail_total, c).border = self.border
        for c in (4, 5, 6):
            ws.cell(detail_total, c).number_format = "#,##0.00"
        ws.cell(detail_total, 7).number_format = "0.00%"
        self._add_variance_formatting(ws, [f"F{first_detail_row}:F{detail_total}", f"G{first_detail_row}:G{detail_total}"])

        widths = [30, 42, 18, 18, 18, 16, 18, 16, 18, 14]
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = f"A{first_summary_row}"
        ws.sheet_view.showGridLines = False

    def create_analytics_sheet(self) -> None:
        ws = self.wb.create_sheet("Analytics", 1)
        self._title(ws, "SALES ANALYTICS", 18)
        ws["A2"] = "← Back to Dashboard"
        ws["A2"].hyperlink = "#'Dashboard'!A1"
        ws["A2"].style = "Hyperlink"
        ws["A3"] = "Sales in KL"
        ws["A3"].font = Font(italic=True, color="666666")

        months = ["July", "August", "September", "October", "November", "December",
                  "January", "February", "March", "April", "May", "June"]
        fy_start = self.processor.start_date.year if self.processor.start_date.month >= 7 else self.processor.start_date.year - 1

        # Separate This Year vs Last Year table and graph for every product.
        section_rows = [5, 24, 43, 62]
        for product, start_row in zip(PRODUCTS, section_rows):
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
            title_cell = ws.cell(start_row, 1, f"{product} — THIS YEAR VS LAST YEAR")
            title_cell.fill = self.dark_fill
            title_cell.font = Font(bold=True, size=12, color="FFFFFF")
            title_cell.alignment = Alignment(horizontal="center")
            title_cell.border = self.border

            header_row = start_row + 1
            for c, value in enumerate(("Month", "Current FY Sales (KL)", "Previous FY Sales (KL)"), 1):
                cell = ws.cell(header_row, c, value)
                cell.fill = self.dark_fill
                cell.font = self.white_bold
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center")

            ty = self.processor.monthly_sales(product, "this")
            ly = self.processor.monthly_sales(product, "last")
            for offset, month_name in enumerate(months, 1):
                row = header_row + offset
                month_no = list(calendar.month_name).index(month_name)
                year = fy_start if month_no >= 7 else fy_start + 1
                ws.cell(row, 1, calendar.month_abbr[month_no])
                ws.cell(row, 2, ty.get((year, month_no), 0) / 1000)
                ws.cell(row, 3, ly.get((year - 1, month_no), 0) / 1000)
                for c in range(1, 4):
                    ws.cell(row, c).border = self.border
                    ws.cell(row, c).alignment = Alignment(horizontal="center")
                ws.cell(row, 2).number_format = "#,##0.00"
                ws.cell(row, 3).number_format = "#,##0.00"

            total_row = header_row + 13
            ws.cell(total_row, 1, "Total")
            ws.cell(total_row, 2, f"=SUM(B{header_row + 1}:B{total_row - 1})")
            ws.cell(total_row, 3, f"=SUM(C{header_row + 1}:C{total_row - 1})")
            for c in range(1, 4):
                ws.cell(total_row, c).fill = PatternFill("solid", fgColor="E2F0D9")
                ws.cell(total_row, c).font = self.bold
                ws.cell(total_row, c).border = self.border
                ws.cell(total_row, c).alignment = Alignment(horizontal="center")
            ws.cell(total_row, 2).number_format = "#,##0.00"
            ws.cell(total_row, 3).number_format = "#,##0.00"

            chart = LineChart()
            chart.title = f"{product} Monthly Sales Trend (KL)"
            chart.y_axis.title = "Sales (KL)"
            chart.x_axis.title = "Month"
            chart.add_data(
                Reference(ws, min_col=2, max_col=3, min_row=header_row, max_row=header_row + 12),
                titles_from_data=True,
            )
            chart.set_categories(Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + 12))
            chart.height = 9.2
            chart.width = 18.5
            chart.legend.position = "r"
            chart.legend.overlay = False
            chart.visible_cells_only = False
            # Force all 12 fiscal-month labels (July to June) to remain visible.
            chart.x_axis.tickLblSkip = 1
            chart.x_axis.tickMarkSkip = 1
            chart.x_axis.tickLblPos = "low"
            chart.y_axis.majorGridlines = chart.y_axis.majorGridlines
            chart.y_axis.scaling.min = 0
            chart.y_axis.numFmt = '#,##0.0'
            chart.y_axis.delete = False
            chart.x_axis.delete = False
            for series in chart.series:
                series.marker.symbol = "circle"
                series.marker.size = 6
                series.graphicalProperties.line.width = 24000
            ws.add_chart(chart, f"E{start_row}")

        # Area-wise billed quantity table and chart (all products, current year).
        area_start = 82
        ws.merge_cells(start_row=area_start-1, start_column=1, end_row=area_start-1, end_column=2)
        ws.cell(area_start-1, 1, "AREA-WISE THIS YEAR SALES (KL)")
        ws.cell(area_start-1, 1).fill = self.dark_fill
        ws.cell(area_start-1, 1).font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        ws.cell(area_start-1, 1).alignment = Alignment(horizontal="center")
        ws.cell(area_start, 1, "Area")
        ws.cell(area_start, 2, "TY_Sales (KL)")
        for c in range(1, 3):
            ws.cell(area_start, c).fill = self.dark_fill
            ws.cell(area_start, c).font = self.white_bold
            ws.cell(area_start, c).border = self.border
        for idx, area in enumerate(AREAS, area_start + 1):
            qty = sum(self.processor.area_totals["this"].get((product, area), 0) for product in PRODUCTS)
            ws.cell(idx, 1, area)
            ws.cell(idx, 2, qty / 1000)
            ws.cell(idx, 2).number_format = "#,##0.00"
            ws.cell(idx, 1).border = self.border
            ws.cell(idx, 2).border = self.border

        area_chart = BarChart()
        area_chart.type = "bar"
        area_chart.style = 10
        area_chart.title = "Area Sales (KL)"
        area_chart.y_axis.title = "Area"
        area_chart.x_axis.title = "Sales (KL)"
        area_chart.add_data(Reference(ws, min_col=2, min_row=area_start, max_row=area_start + len(AREAS)), titles_from_data=True)
        area_chart.set_categories(Reference(ws, min_col=1, min_row=area_start + 1, max_row=area_start + len(AREAS)))
        area_chart.legend = None
        area_chart.height = 8.5
        area_chart.width = 17.5
        area_chart.x_axis.scaling.min = 0
        area_chart.x_axis.numFmt = '#,##0.0'
        area_chart.x_axis.delete = False
        area_chart.y_axis.delete = False
        area_chart.visible_cells_only = False
        ws.add_chart(area_chart, f"E{area_start}")

        # Date-wise total billed quantity (all products, current year).
        daily_totals = defaultdict(float)
        for (product, day), qty in self.processor.product_daily["this"].items():
            daily_totals[day] += qty
        daily_rows = sorted(daily_totals.items())
        daily_src_col = 25
        ws.cell(1, daily_src_col, "Date")
        ws.cell(1, daily_src_col + 1, "Sales in KL")
        ws.cell(1, daily_src_col + 2, "Sunday")
        for row_no, (day, qty) in enumerate(daily_rows, 2):
            ws.cell(row_no, daily_src_col, day)
            ws.cell(row_no, daily_src_col).number_format = "dd-mmm-yy"
            ws.cell(row_no, daily_src_col + 1, qty / 1000)
            ws.cell(row_no, daily_src_col + 1).number_format = "#,##0.0"
            # Sunday-only helper series used to put a clear red marker on the chart.
            ws.cell(row_no, daily_src_col + 2, qty / 1000 if day.weekday() == 6 else None)
            ws.cell(row_no, daily_src_col + 2).number_format = "#,##0.0"

        date_start = 98
        ws.cell(date_start, 1, "DATE-WISE SALES")
        ws.cell(date_start, 1).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=date_start, start_column=3, end_row=date_start, end_column=9)
        ws.cell(date_start, 3, "Date shown as dd-mmm-yy  |  Sunday = red marker")
        ws.cell(date_start, 3).font = Font(italic=True, color="9C0006")
        ws.cell(date_start, 3).alignment = Alignment(horizontal="left")
        date_chart = LineChart()
        date_chart.title = "Daily Sales Trend (KL) — Sundays marked in red"
        date_chart.y_axis.title = "Sales (KL)"
        date_chart.x_axis.title = "Date (dd-mmm)"
        if daily_rows:
            date_chart.add_data(Reference(ws, min_col=daily_src_col + 1, max_col=daily_src_col + 2, min_row=1, max_row=len(daily_rows) + 1), titles_from_data=True)
            date_chart.set_categories(Reference(ws, min_col=daily_src_col, min_row=2, max_row=len(daily_rows) + 1))
        date_chart.legend.position = "b"
        date_chart.legend.overlay = False
        date_chart.x_axis.number_format = "d-mmm"
        date_chart.x_axis.tickLblPos = "low"
        # Show readable daily dates without crowding: every third label is
        # displayed, while every daily point remains plotted.
        date_chart.x_axis.tickLblSkip = 2
        date_chart.x_axis.tickMarkSkip = 1
        if date_chart.series:
            date_chart.series[0].graphicalProperties.line.solidFill = "4472C4"
            date_chart.series[0].graphicalProperties.line.width = 24000
            date_chart.series[0].marker.symbol = "circle"
            date_chart.series[0].marker.size = 4
            date_chart.series[0].marker.graphicalProperties.solidFill = "4472C4"
            date_chart.series[0].marker.graphicalProperties.line.solidFill = "4472C4"
        if len(date_chart.series) > 1:
            date_chart.series[1].graphicalProperties.line.noFill = True
            date_chart.series[1].marker.symbol = "circle"
            date_chart.series[1].marker.size = 8
            date_chart.series[1].marker.graphicalProperties.solidFill = "FF0000"
            date_chart.series[1].marker.graphicalProperties.line.solidFill = "FF0000"
        date_chart.height = 9.0
        date_chart.width = 24.0
        date_chart.visible_cells_only = False
        ws.add_chart(date_chart, f"A{date_start + 1}")

        # Dynamic Top-N outlet tables. Each table has its own in-table selector.
        # Keep Top Outlets safely below the Daily Sales chart.
        start = 150
        sections = [("HSD", "TOP HSD OUTLETS"), ("PMG", "TOP PMG OUTLETS"),
                    ("R95", "TOP R95 OUTLETS"), (None, "TOP COMBINED OUTLETS")]
        selector_dv = DataValidation(type="list", formula1='"10,20,30,40,50"', allow_blank=False)
        ws.add_data_validation(selector_dv)
        helper_col = 30  # AD onward; hidden source data for dynamic tables.

        for idx, (product, title) in enumerate(sections):
            col = 1 + (idx % 2) * 6
            # Each table can display up to 50 outlets, so allow enough vertical room
            # between the first and second pair of outlet tables.
            row = start + (idx // 2) * 55
            selector_cell = ws.cell(row, col + 4)
            ws.cell(row, col, title)
            ws.cell(row, col).fill = self.dark_fill
            ws.cell(row, col).font = Font(bold=True, size=11, color="FFFFFF")
            ws.cell(row, col + 3, "Show Top")
            ws.cell(row, col + 3).fill = self.dark_fill
            ws.cell(row, col + 3).font = self.white_bold
            selector_cell.value = 10
            selector_cell.fill = PatternFill("solid", fgColor="FFF2CC")
            selector_cell.font = self.bold
            selector_cell.alignment = Alignment(horizontal="center")
            selector_dv.add(selector_cell)

            headers = ("Rank", "Cost Center", "Outlet Name", "Sold to Party", "Additional Qty (KL)")
            for offset, header in enumerate(headers):
                c = ws.cell(row + 1, col + offset, header)
                c.fill = self.dark_fill; c.font = self.white_bold; c.border = self.border
                c.alignment = Alignment(horizontal="center", wrap_text=True)

            records = self.processor.top_outlets(product, limit=50)
            source_start_col = helper_col + idx * 5
            for offset, header in enumerate(headers):
                ws.cell(1, source_start_col + offset, header)
            for rank in range(1, 51):
                rec = records[rank - 1] if rank <= len(records) else {}
                vals = [rank, rec.get("Cost Center", ""), rec.get("Outlet Name", ""),
                        rec.get("Sold to Party", ""), rec.get("This Year Sales", 0) / 1000]
                for offset, value in enumerate(vals):
                    ws.cell(1 + rank, source_start_col + offset, value)

                visible_row = row + 1 + rank
                selector_ref = f"${get_column_letter(col + 4)}${row}"
                for offset in range(5):
                    src = f"{get_column_letter(source_start_col + offset)}{1 + rank}"
                    ws.cell(visible_row, col + offset, f'=IF({rank}<={selector_ref},{src},"")')
                    ws.cell(visible_row, col + offset).border = self.border
                ws.cell(visible_row, col + 4).number_format = "#,##0.00"

        for c in range(helper_col, helper_col + len(sections) * 5):
            ws.column_dimensions[get_column_letter(c)].hidden = True

        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A5"
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        for c in range(5, 19):
            ws.column_dimensions[get_column_letter(c)].width = 12
        for c in range(daily_src_col, daily_src_col + 3):
            ws.column_dimensions[get_column_letter(c)].hidden = True

        # Developer credit at the bottom of the Analytics sheet.
        developer_row = start + 55 + 54
        ws.merge_cells(start_row=developer_row, start_column=1, end_row=developer_row, end_column=12)
        ws.cell(developer_row, 1, "Developed by: Rida Batool (NED) & M.Arham Shakeel (SSUET)")
        ws.cell(developer_row, 1).font = Font(bold=True, italic=True, color="666666")
        ws.cell(developer_row, 1).alignment = Alignment(horizontal="center")

    def create_monthly_area_matrix_sheet(self) -> None:
        """Interactive monthly area matrix with TY/LY, objective and required daily run-rate."""
        ws = self.wb.create_sheet("Monthly_Area_Matrix", 2)
        self._title(ws, "MONTHLY AREA-WISE SALES MATRIX (KL)", 38)
        ws["A2"] = "← Back to Dashboard"; ws["A2"].hyperlink = "#'Dashboard'!A1"; ws["A2"].style = "Hyperlink"
        ws["A3"] = "Select Product"; ws["B3"] = "PMG"
        ws["D3"] = "Select Month"; ws["E3"] = self.processor.end_date.strftime("%B") if self.processor.end_date else "July"
        ws["G3"] = "Select Area"; ws["H3"] = AREAS[0]
        for ref in ("A3", "D3", "G3"):
            ws[ref].fill=self.light_fill; ws[ref].font=self.bold; ws[ref].border=self.border; ws[ref].alignment=Alignment(horizontal="center")
        for ref in ("B3", "E3", "H3"):
            ws[ref].fill=PatternFill("solid",fgColor="FFF2CC"); ws[ref].font=self.bold; ws[ref].border=self.border; ws[ref].alignment=Alignment(horizontal="center")
        product_dv=DataValidation(type="list",formula1='"PMG,HSD,R95,MOGAS"',allow_blank=False)
        month_dv=DataValidation(type="list",formula1='"July,August,September,October,November,December,January,February,March,April,May,June"',allow_blank=False)
        area_dv=DataValidation(type="list",formula1='"'+",".join(AREAS)+'"',allow_blank=False)
        ws.add_data_validation(product_dv); product_dv.add(ws["B3"])
        ws.add_data_validation(month_dv); month_dv.add(ws["E3"])
        ws.add_data_validation(area_dv); area_dv.add(ws["H3"])

        # Guidance paragraph retained below the dropdown selectors.
        ws.merge_cells("A4:AK4")
        ws["A4"] = (
            "Select Product, Month and Area from the yellow dropdowns. Daily sales are shown in thousands of liters; "
            "the Last Year benchmark is adjusted for elapsed working days, and Per Day Sale Required is "
            "calculated as (Objective - Total Sales) / Remaining Working Days."
        )
        ws["A4"].fill = PatternFill("solid", fgColor="E2F0D9")
        ws["A4"].font = Font(italic=True, color="375623")
        ws["A4"].alignment = Alignment(wrap_text=True, vertical="center")
        ws["A4"].border = self.border
        ws.row_dimensions[4].height = 34

        header_row=6
        headers=["Area"]+list(range(1,32))+["Total TY","LY","Var TY vs LY %","TM Obj","TM vs Obj %","PDSR"]
        for c,h in enumerate(headers,1):
            cell=ws.cell(header_row,c,h)
            if isinstance(h,int): cell.value=f'=IF({h}>INDEX($AU$2:$AU$13,MATCH($E$3,$AT$2:$AT$13,0)),"",{h})'
            cell.fill=self.dark_fill; cell.font=self.white_bold; cell.border=self.border; cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

        # Hidden normalized TY/LY source.
        src_col=39  # AM
        for i,v in enumerate(["Product","Area","Month","Day","TY Qty","LY Qty"],src_col): ws.cell(1,i,v)
        ty=defaultdict(float); ly=defaultdict(float)
        for period,target in (("this",ty),("last",ly)):
            for (product,area,day),qty in self.processor.area_daily[period].items():
                if product in ("PMG","HSD","R95") and area in AREAS:
                    target[(product,area,day.strftime("%B"),day.day)]+=qty
                    if product in ("PMG","R95"): target[("MOGAS",area,day.strftime("%B"),day.day)]+=qty
        keys=sorted(set(ty)|set(ly))
        for r,key in enumerate(keys,2):
            vals=list(key)+[ty.get(key,0.0),ly.get(key,0.0)]
            for c,v in enumerate(vals,src_col): ws.cell(r,c,v)
        src_last=max(2,len(keys)+1)

        # Month lookup AT:AW.
        month_lookup_col=46
        months=["July","August","September","October","November","December","January","February","March","April","May","June"]
        fy_start=self.processor.start_date.year if self.processor.start_date.month>=7 else self.processor.start_date.year-1
        for i,v in enumerate(["Month","Days","Year","Month No"],month_lookup_col): ws.cell(1,i,v)
        for r,m in enumerate(months,2):
            mn=list(calendar.month_name).index(m); yr=fy_start if mn>=7 else fy_start+1
            ws.cell(r,month_lookup_col,m); ws.cell(r,month_lookup_col+1,calendar.monthrange(yr,mn)[1]); ws.cell(r,month_lookup_col+2,yr); ws.cell(r,month_lookup_col+3,mn)

        # Objective helper AX:BA.
        obj_col=50
        for i,v in enumerate(["Product","Area","Month","Objective"],obj_col): ws.cell(1,i,v)
        obj_rows=[]
        for (p,a,m),v in self.processor.objectives.items(): obj_rows.append((p,a,m,v))
        # MOGAS objective is PMG + R95.
        for a in AREAS:
            for m in months:
                obj_rows.append(("MOGAS",a,m,sum(self.processor.objectives.get((p,a,m),0) for p in ("PMG","R95"))))
        for r,vals in enumerate(obj_rows,2):
            for c,v in enumerate(vals,obj_col): ws.cell(r,c,v)
        obj_last=max(2,len(obj_rows)+1)

        first=7
        for r,area in enumerate(AREAS,first):
            ws.cell(r,1,area); ws.cell(r,1).font=self.bold; ws.cell(r,1).fill=self.area_fill; ws.cell(r,1).border=self.border; ws.cell(r,1).alignment=Alignment(horizontal="center")
            for day in range(1,32):
                c=day+1
                ws.cell(r,c,f'=IF({day}>INDEX($AU$2:$AU$13,MATCH($E$3,$AT$2:$AT$13,0)),"",SUMIFS($AQ$2:$AQ${src_last},$AM$2:$AM${src_last},$B$3,$AN$2:$AN${src_last},$A{r},$AO$2:$AO${src_last},$E$3,$AP$2:$AP${src_last},{day})/1000)')
                ws.cell(r,c).number_format='#,##0'; ws.cell(r,c).border=self.border; ws.cell(r,c).alignment=Alignment(horizontal="center")
            ws.cell(r,33,f'=SUM(B{r}:AF{r})')
            # LY full-month sales / LY non-Sunday working days × current elapsed non-Sunday working days.
            ws.cell(r,34,f'=IFERROR(SUMIFS($AR$2:$AR${src_last},$AM$2:$AM${src_last},$B$3,$AN$2:$AN${src_last},$A{r},$AO$2:$AO${src_last},$E$3)/NETWORKDAYS.INTL(DATE(INDEX($AV$2:$AV$13,MATCH($E$3,$AT$2:$AT$13,0)),INDEX($AW$2:$AW$13,MATCH($E$3,$AT$2:$AT$13,0)),1),EOMONTH(DATE(INDEX($AV$2:$AV$13,MATCH($E$3,$AT$2:$AT$13,0)),INDEX($AW$2:$AW$13,MATCH($E$3,$AT$2:$AT$13,0)),1),0),11)*MAX(0,NETWORKDAYS.INTL(DATE(INDEX($AV$2:$AV$13,MATCH($E$3,$AT$2:$AT$13,0)),INDEX($AW$2:$AW$13,MATCH($E$3,$AT$2:$AT$13,0)),1),MIN(TODAY(),EOMONTH(DATE(INDEX($AV$2:$AV$13,MATCH($E$3,$AT$2:$AT$13,0)),INDEX($AW$2:$AW$13,MATCH($E$3,$AT$2:$AT$13,0)),1),0)),11))/1000,0)')
            ws.cell(r,35,f'=IFERROR((AG{r}-AH{r})/AH{r},0)')
            ws.cell(r,36,f'=SUMIFS($BA$2:$BA${obj_last},$AX$2:$AX${obj_last},$B$3,$AY$2:$AY${obj_last},$A{r},$AZ$2:$AZ${obj_last},$E$3)/1000')
            ws.cell(r,37,f'=IFERROR((AG{r}-AJ{r})/AJ{r},0)')
            # Required per working day = (objective - sales) / remaining non-Sunday days.
            ws.cell(r,38,f'=IF(TODAY()>EOMONTH(DATE(INDEX($AV$2:$AV$13,MATCH($E$3,$AT$2:$AT$13,0)),INDEX($AW$2:$AW$13,MATCH($E$3,$AT$2:$AT$13,0)),1),0),0,IFERROR(MAX(0,AJ{r}-AG{r})/MAX(1,NETWORKDAYS.INTL(MAX(TODAY()+1,DATE(INDEX($AV$2:$AV$13,MATCH($E$3,$AT$2:$AT$13,0)),INDEX($AW$2:$AW$13,MATCH($E$3,$AT$2:$AT$13,0)),1)),EOMONTH(DATE(INDEX($AV$2:$AV$13,MATCH($E$3,$AT$2:$AT$13,0)),INDEX($AW$2:$AW$13,MATCH($E$3,$AT$2:$AT$13,0)),1),0),11)),0))')
            for c in range(33,39): ws.cell(r,c).number_format='#,##0'; ws.cell(r,c).border=self.border
            ws.cell(r,35).number_format='0.00%'
            ws.cell(r,37).number_format='0.00%'
        total=first+len(AREAS)
        ws.cell(total,1,"Total")
        for c in range(2,39):
            letter=get_column_letter(c); ws.cell(total,c,f'=SUM({letter}{first}:{letter}{total-1})'); ws.cell(total,c).number_format='#,##0.00'
        for c in range(1,39): ws.cell(total,c).fill=self.light_fill; ws.cell(total,c).font=self.bold; ws.cell(total,c).border=self.border
        ws.cell(total,35,f'=IFERROR((AG{total}-AH{total})/AH{total},0)'); ws.cell(total,35).number_format="0.00%"
        ws.cell(total,37,f'=IFERROR((AG{total}-AJ{total})/AJ{total},0)'); ws.cell(total,37).number_format="0.00%"
        self._add_variance_formatting(ws,[f"AI{first}:AI{total}", f"AK{first}:AK{total}"])

        # Dynamic Sunday highlighting for the selected month.  The rule uses
        # the fiscal-year lookup table, so changing E3 immediately highlights
        # the exact Sunday date columns for that month (including leap years).
        sunday_fill = PatternFill("solid", fgColor="FFD966")
        sunday_font = Font(bold=True, color="9C6500")
        for day in range(1, 32):
            day_col = get_column_letter(day + 1)  # Day 1=B ... Day 31=AF
            sunday_formula = (
                f'AND({day_col}${header_row}<>"",'
                f'WEEKDAY(DATE(INDEX($AV$2:$AV$13,MATCH($E$3,$AT$2:$AT$13,0)),'
                f'INDEX($AW$2:$AW$13,MATCH($E$3,$AT$2:$AT$13,0)),'
                f'{day_col}${header_row}),2)=7)'
            )
            ws.conditional_formatting.add(
                f"{day_col}{header_row}:{day_col}{total}",
                FormulaRule(formula=[sunday_formula], fill=sunday_fill, font=sunday_font, stopIfTrue=True),
            )

        # Dynamic graph for the area selected in H3.  The hidden helper range
        # contains daily TY and LY values and updates with all three dropdowns.
        graph_col = 54  # BB
        graph_headers = ["Day", "This Year", "Last Year", "This Year Sunday", "Last Year Sunday"]
        for offset, value in enumerate(graph_headers):
            ws.cell(1, graph_col + offset, value)
        for day in range(1, 32):
            rr = day + 1
            ws.cell(rr, graph_col, day)
            ws.cell(
                rr,
                graph_col + 1,
                f'=IF({day}>INDEX($AU$2:$AU$13,MATCH($E$3,$AT$2:$AT$13,0)),NA(),'
                f'INDEX($B${first}:$AF${total-1},MATCH($H$3,$A${first}:$A${total-1},0),{day}))',
            )
            ws.cell(
                rr,
                graph_col + 2,
                f'=IF({day}>INDEX($AU$2:$AU$13,MATCH($E$3,$AT$2:$AT$13,0)),NA(),'
                f'SUMIFS($AR$2:$AR${src_last},$AM$2:$AM${src_last},$B$3,'
                f'$AN$2:$AN${src_last},$H$3,$AO$2:$AO${src_last},$E$3,'
                f'$AP$2:$AP${src_last},{day})/1000)',
            )
            # Marker-only helper series: return values only when the selected
            # calendar date is Sunday. These formulas update with the month
            # dropdown, allowing Sunday points to remain highlighted in red.
            sunday_test = (
                f'WEEKDAY(DATE(INDEX($AV$2:$AV$13,MATCH($E$3,$AT$2:$AT$13,0)),'
                f'INDEX($AW$2:$AW$13,MATCH($E$3,$AT$2:$AT$13,0)),{day}),1)=1'
            )
            ws.cell(
                rr,
                graph_col + 3,
                f'=IF(OR({day}>INDEX($AU$2:$AU$13,MATCH($E$3,$AT$2:$AT$13,0)),NOT({sunday_test})),NA(),'
                f'{get_column_letter(graph_col + 1)}{rr})',
            )
            ws.cell(
                rr,
                graph_col + 4,
                f'=IF(OR({day}>INDEX($AU$2:$AU$13,MATCH($E$3,$AT$2:$AT$13,0)),NOT({sunday_test})),NA(),'
                f'{get_column_letter(graph_col + 2)}{rr})',
            )

        matrix_chart = LineChart()
        matrix_chart.title = "Daily Sales Trend (KL)"
        matrix_chart.y_axis.title = "Sales (KL)"
        matrix_chart.x_axis.title = "Day"
        matrix_chart.add_data(
            Reference(ws, min_col=graph_col + 1, max_col=graph_col + 4, min_row=1, max_row=32),
            titles_from_data=True,
        )
        matrix_chart.set_categories(Reference(ws, min_col=graph_col, min_row=2, max_row=32))

        # The last two series contain Sunday values only. Hide their connecting
        # lines and use larger red markers so Sundays stand out on both TY and LY.
        for sunday_series in matrix_chart.series[2:4]:
            # Keep marker-only Sunday series, but avoid the nested marker-line
            # XML that caused Excel to repair the generated workbook.
            sunday_series.graphicalProperties.line.noFill = True
            sunday_series.marker.symbol = "circle"
            sunday_series.marker.size = 8
            sunday_series.marker.graphicalProperties.solidFill = "FF0000"

        matrix_chart.height = 8
        matrix_chart.width = 18
        matrix_chart.legend.position = "b"
        matrix_chart.visible_cells_only = False
        ws.add_chart(matrix_chart, f"A{total + 3}")

        ws.freeze_panes="B7"; ws.sheet_view.showGridLines=False
        ws.column_dimensions["A"].width=13
        for c in range(2,33): ws.column_dimensions[get_column_letter(c)].width=9
        for c in range(33,39): ws.column_dimensions[get_column_letter(c)].width=20
        ws.column_dimensions["AL"].width = 24
        ws.row_dimensions[header_row].height=38
        for c in range(src_col, graph_col + 5):
            ws.column_dimensions[get_column_letter(c)].hidden=True

    def create_outlet_comparison_sheet(self) -> None:
        ws = self.wb.create_sheet("Outlet_TY_vs_LY", 2)
        self._title(ws, "OUTLET SALES COMPARISON", 12)
        ws["A2"] = "← Back to Dashboard"; ws["A2"].hyperlink = "#'Dashboard'!A1"; ws["A2"].style = "Hyperlink"
        outlets = sorted(self.processor.cost_center_info)
        months = ["July","August","September","October","November","December","January","February","March","April","May","June"]
        labels = [("A3","Select Outlet Name or Cost Center"),("D3","From Month"),("G3","To Month"),("J3","Product")]
        for ref,text in labels:
            ws[ref]=text; ws[ref].fill=self.light_fill; ws[ref].font=self.bold; ws[ref].border=self.border
        ws["E3"] = months[0]; ws["H3"] = months[-1]; ws["K3"] = "ALL"
        for ref in ("B3","E3","H3","K3"):
            ws[ref].fill=PatternFill("solid",fgColor="FFF2CC"); ws[ref].font=self.bold; ws[ref].border=self.border

        # Hidden normalized data and selector mapping.  Each latest outlet gets
        # Cost Center, Code and Name search entries, all resolving to one stable CC.
        helper=self.wb.create_sheet("Outlet_Comparison_Data")
        helper.append(["Cost Center","Outlet","Sold to Party","Product","Month","Month Index","TY Qty","LY Qty","","Selector","Selector Cost Center"])
        fy_start=self.processor.start_date.year if self.processor.start_date.month>=7 else self.processor.start_date.year-1
        for outlet in outlets:
            info=self.processor._display_info(outlet)
            sold=info.get("Outlet Code", self.processor.outlet_sold_to_party.get(outlet,""))
            name=info.get("Outlet Name",outlet)
            areas={a for period in ("last","this") for (p,a,o,y,m) in self.processor.outlet_monthly_totals[period] if o==outlet}
            for product in PRODUCTS:
                for mi,mname in enumerate(months,1):
                    mn=list(calendar.month_name).index(mname); year=fy_start if mn>=7 else fy_start+1
                    ty=sum(self.processor.outlet_monthly_totals["this"].get((product,a,outlet,year,mn),0) for a in areas)
                    ly=sum(self.processor.outlet_monthly_totals["last"].get((product,a,outlet,year-1,mn),0) for a in areas)
                    helper.append([outlet,name,sold,product,mname,mi,ty,ly])

        selector_rows=[]
        # One clean dropdown row per outlet. The visible text contains both
        # searchable fields without repeating Code/Name variants.
        clean_entries=[]
        for outlet in outlets:
            info=self.processor._display_info(outlet)
            name=str(info.get("Outlet Name", outlet) or outlet).strip()
            clean_entries.append((name.casefold(), str(outlet), f"{name} | {outlet}", outlet))
        for _name_key,_cc_key,label,cc in sorted(clean_entries):
            selector_rows.append((label,cc))
        for i,(label,cc) in enumerate(selector_rows,2):
            helper.cell(i,10,label); helper.cell(i,11,cc)
        ws["B3"] = selector_rows[0][0] if selector_rows else ""
        helper.sheet_state="hidden"

        dv=DataValidation(type="list",formula1=f"='Outlet_Comparison_Data'!$J$2:$J${max(2,len(selector_rows)+1)}")
        fm=DataValidation(type="list",formula1='"'+','.join(months)+'"'); tm=DataValidation(type="list",formula1='"'+','.join(months)+'"')
        pdv=DataValidation(type="list",formula1='"ALL,PMG,HSD,R95,Lubricants"')
        for obj,cell in ((dv,"B3"),(fm,"E3"),(tm,"H3"),(pdv,"K3")): ws.add_data_validation(obj); obj.add(ws[cell])

        ws["A5"]="Current Outlet Name"
        ws["B5"]='=IFERROR(INDEX(Outlet_Comparison_Data!$B:$B,MATCH(INDEX(Outlet_Comparison_Data!$K:$K,MATCH($B$3,Outlet_Comparison_Data!$J:$J,0)),Outlet_Comparison_Data!$A:$A,0)),"")'
        ws["D5"]="Current Outlet Code"
        ws["E5"]='=IFERROR(INDEX(Outlet_Comparison_Data!$C:$C,MATCH(INDEX(Outlet_Comparison_Data!$K:$K,MATCH($B$3,Outlet_Comparison_Data!$J:$J,0)),Outlet_Comparison_Data!$A:$A,0)),"")'
        ws["G5"]="Cost Center"
        ws["H5"]='=IFERROR(INDEX(Outlet_Comparison_Data!$K:$K,MATCH($B$3,Outlet_Comparison_Data!$J:$J,0)),"")'
        for ref in ("A5","D5","G5"):
            ws[ref].font=self.bold; ws[ref].fill=self.area_fill; ws[ref].border=self.border
        for ref in ("B5","E5","H5"): ws[ref].border=self.border

        hdr=7
        for c,h in enumerate(("Month","TY Sales","LY Sales","Variance (KL)","Variance %"),1):
            cell=ws.cell(hdr,c,h); cell.fill=self.dark_fill; cell.font=self.white_bold; cell.border=self.border
        last=helper.max_row
        selected_cc='INDEX(Outlet_Comparison_Data!$K:$K,MATCH($B$3,Outlet_Comparison_Data!$J:$J,0))'
        for i,mname in enumerate(months,1):
            rr=hdr+i; ws.cell(rr,1,mname
            )
            condition=f'AND({i}>=MATCH($E$3,{{"July","August","September","October","November","December","January","February","March","April","May","June"}},0),{i}<=MATCH($H$3,{{"July","August","September","October","November","December","January","February","March","April","May","June"}},0))'
            product_crit=f'IF($K$3="ALL","*",$K$3)'
            ws.cell(rr,2,f'=IF({condition},SUMIFS(Outlet_Comparison_Data!$G$2:$G${last},Outlet_Comparison_Data!$A$2:$A${last},{selected_cc},Outlet_Comparison_Data!$E$2:$E${last},A{rr},Outlet_Comparison_Data!$D$2:$D${last},{product_crit})/1000,NA())')
            ws.cell(rr,3,f'=IF({condition},SUMIFS(Outlet_Comparison_Data!$H$2:$H${last},Outlet_Comparison_Data!$A$2:$A${last},{selected_cc},Outlet_Comparison_Data!$E$2:$E${last},A{rr},Outlet_Comparison_Data!$D$2:$D${last},{product_crit})/1000,NA())')
            ws.cell(rr,4,f'=IFERROR(B{rr}-C{rr},0)'); ws.cell(rr,5,f'=IFERROR(D{rr}/C{rr},0)')
            for c in range(1,6): ws.cell(rr,c).border=self.border
            for c in (2,3,4): ws.cell(rr,c).number_format="#,##0.00"
            ws.cell(rr,5).number_format="0.00%"
        chart=LineChart(); chart.title="Monthly Sales Trend (KL)"; chart.y_axis.title="Sales (KL)"; chart.x_axis.title="Month"
        chart.add_data(Reference(ws,min_col=2,max_col=3,min_row=hdr,max_row=hdr+12),titles_from_data=True); chart.set_categories(Reference(ws,min_col=1,min_row=hdr+1,max_row=hdr+12)); chart.height=8; chart.width=16; chart.legend.position="b"; chart.visible_cells_only=False
        chart.x_axis.tickLblSkip=1; chart.x_axis.tickMarkSkip=1; chart.x_axis.tickLblPos="low"
        chart.y_axis.scaling.min=0; chart.y_axis.numFmt='#,##0.00'
        for series in chart.series:
            series.marker.symbol="circle"; series.marker.size=6; series.graphicalProperties.line.width=24000
        ws.add_chart(chart,"G7")
        ws.freeze_panes="A8"; ws.sheet_view.showGridLines=False
        for col,w in {"A":20,"B":42,"C":18,"D":20,"E":18,"G":20,"H":18,"J":14,"K":18}.items(): ws.column_dimensions[col].width=w

    def create_area_outlet_average_sheet(self) -> None:
        ws=self.wb.create_sheet("Area_Outlet_6M_Avg",3)
        self._title(ws,"AREA-WISE OUTLET AVERAGE — PREVIOUS 6 MONTHS",9)
        ws["A2"]="← Back to Dashboard"; ws["A2"].hyperlink="#'Dashboard'!A1"; ws["A2"].style="Hyperlink"
        cols=["Area","Cost Center","Outlet Name","Sold to Party"]+[f"{p} 6M Avg" for p in PRODUCTS]+["Total 6M Avg"]
        rows=self.processor.outlet_six_month_averages(); hdr=4
        for c,h in enumerate(cols,1):
            cell=ws.cell(hdr,c,h); cell.fill=self.dark_fill; cell.font=self.white_bold; cell.border=self.border; cell.alignment=Alignment(horizontal="center",wrap_text=True)
        for r,rec in enumerate(rows,hdr+1):
            for c,h in enumerate(cols,1):
                # Total is intentionally fuel-only: HSD + PMG + R95.
                # Lubricants remain visible as a separate column but are never
                # included in Total 6M Avg.
                if h == "Total 6M Avg":
                    val = sum(float(rec.get(f"{product} 6M Avg", 0) or 0) for product in ("HSD", "PMG", "R95"))
                else:
                    val = rec.get(h,"")
                val=val/1000 if "Avg" in h and isinstance(val,(int,float)) else val
                cell=ws.cell(r,c,val); cell.border=self.border
                if "Avg" in h:
                    if h == "Lubricants 6M Avg":
                        cell.number_format = "#,##0.00"
                    else:
                        cell.number_format = "#,##0"
        ws.auto_filter.ref=f"A{hdr}:I{hdr+len(rows)}"; ws.freeze_panes=f"D{hdr+1}"; ws.sheet_view.showGridLines=False
        widths=[10,18,32,16,16,16,16,18,18]
        for c,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(c)].width=w

    def create_unmapped_outlets_sheet(self) -> None:
        ws = self.wb.create_sheet("Unmapped_Outlets")
        self._title(ws, "UNMAPPED OUTLET CODES — ADD THESE TO OUTLET_MASTER", 5)
        cols = ["Cost Center", "Outlet Code", "Outlet Name", "Area", "Last Year Qty", "This Year Qty", "Rows", "Action Required"]
        for c, h in enumerate(cols, 1):
            cell = ws.cell(3, c, h); cell.fill = self.dark_fill; cell.font = self.white_bold; cell.border = self.border
        rows = list(self.processor.unmapped_outlets.values())
        for r, rec in enumerate(rows, 4):
            vals = [rec.get("Cost Center", ""), rec.get("Outlet Code", ""), rec.get("Outlet Name", ""), rec.get("Area", ""), rec.get("Last Year Qty", 0), rec.get("This Year Qty", 0), rec.get("Rows", 0), "Add code and permanent Cost Center to Input/Outlet_Master.xlsx"]
            for c, value in enumerate(vals, 1):
                ws.cell(r, c, value).border = self.border
        if not rows:
            ws.cell(4, 1, "All SAP outlet codes are mapped.")
        widths = [18, 16, 36, 12, 55]
        for i, width in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A4"

    def create_all(self) -> None:
        print("[1/4] Creating Dashboard...")
        self.create_dashboard()
        self._create_lubricant_pipeline_sheet()
        self._create_lubricant_liter_master_sheet()
        self.create_lubricant_dashboards()
        self.create_analytics_sheet()
        self.create_monthly_area_matrix_sheet()
        self.create_outlet_comparison_sheet()
        self.create_area_outlet_average_sheet()
        self.create_unmapped_outlets_sheet()
        for material in self.processor.lubricant_materials():
            self.create_material_sheet(material, "ALL")
            for area in AREAS:
                self.create_material_sheet(material, area)
        # Formatting is applied when sheets are created. Avoid a full-workbook
        # cell-by-cell normalization pass here because it is extremely slow on
        # Streamlit Community Cloud for large workbooks.
        self._create_control_data_sheet()
        print("[2/4] Creating product sheets...")
        for product in PRODUCTS:
            print(f"      Creating {product}...")
            self.create_product_sheet(product)
        print("[3/4] Creating product-area detail sheets...")
        for product in PRODUCTS:
            for area in AREAS:
                print(f"      Creating {product}_{area}...")
                self.create_product_area_sheet(product, area)
        print("[4/4] Creating area summary sheets...")
        for area in AREAS:
            print(f"      Creating Area_{area}...")
            self.create_area_sheet(area)

    @staticmethod
    def _formula_references(formula: str, current_sheet: str) -> set[tuple[str, str]]:
        """Return real A1 references used by a formula.

        Text inside quoted Excel strings is excluded, and both ends of a range
        inherit the same explicit sheet name.  This prevents product text such
        as ``"R95|"`` and the second endpoint of ``Control_Data!J2:AN500``
        from being mistaken for local worksheet cells.
        """
        refs: set[tuple[str, str]] = set()

        # Remove double-quoted Excel string literals while preserving formula
        # structure. Escaped quotes inside strings are represented by "".
        cleaned = re.sub(r'"(?:[^"]|"")*"', '""', formula)

        # Match a single reference or an A1 range.  The optional sheet prefix
        # applies to both range endpoints, exactly as Excel interprets it.
        pattern = re.compile(
            r"(?:(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_ .]*))!)?"
            r"\$?([A-Z]{1,3})\$?(\d+)"
            r"(?::\$?([A-Z]{1,3})\$?(\d+))?"
        )
        for match in pattern.finditer(cleaned):
            sheet = (match.group(1) or match.group(2) or current_sheet).strip()
            refs.add((sheet, f"{match.group(3)}{match.group(4)}"))
            if match.group(5) and match.group(6):
                refs.add((sheet, f"{match.group(5)}{match.group(6)}"))
        return refs

    def _remove_circular_formulas(self) -> int:
        """Detect formula dependency cycles and safely replace only cyclic cells.

        A generated dashboard should never contain a circular calculation. This
        final validation prevents Excel's circular-reference warning even if a
        later layout change accidentally introduces a self/indirect reference.
        """
        formulas: dict[tuple[str, str], str] = {}
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas[(ws.title, cell.coordinate)] = cell.value

        graph: dict[tuple[str, str], set[tuple[str, str]]] = {}
        for node, formula in formulas.items():
            graph[node] = {ref for ref in self._formula_references(formula, node[0]) if ref in formulas}

        state: dict[tuple[str, str], int] = {}
        stack: list[tuple[str, str]] = []
        cyclic: set[tuple[str, str]] = set()

        def visit(node: tuple[str, str]) -> None:
            state[node] = 1
            stack.append(node)
            for dep in graph.get(node, ()):
                if state.get(dep, 0) == 0:
                    visit(dep)
                elif state.get(dep) == 1:
                    try:
                        cyclic.update(stack[stack.index(dep):])
                    except ValueError:
                        cyclic.add(dep)
            stack.pop()
            state[node] = 2

        for node in graph:
            if state.get(node, 0) == 0:
                visit(node)

        for sheet_name, coordinate in cyclic:
            cell = self.wb[sheet_name][coordinate]
            cell.value = 0
            cell.number_format = cell.number_format or "#,##0"
        if cyclic:
            print(f"Removed {len(cyclic)} circular formula cell(s): " + ", ".join(f"{s}!{c}" for s,c in sorted(cyclic)[:20]))
        else:
            print("Circular-reference validation passed: no formula cycles found.")
        return len(cyclic)

    def _finalize_dashboard_formatting(self) -> None:
        """Standardize summary headings and dashboard button alignment."""
        exact_variance_headers = {"variance", "variance %", "vs objective", "vs objective %", "var_ly", "var_ly%", "ty_vs_obj", "ty_vs_obj%"}
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    # Display headings without technical underscores while preserving formulas/data.
                    if isinstance(cell.value, str) and "_" in cell.value and (cell.font.bold or cell.fill.fill_type == "solid"):
                        cell.value = cell.value.replace("_", " ")
                    text = str(cell.value or "").strip().casefold()
                    if text in exact_variance_headers:
                        cell.fill = self.dark_fill
                        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = self.border
        # Main dashboard navigation buttons use five equal two-column blocks.
        if "Dashboard" in self.wb.sheetnames:
            ws = self.wb["Dashboard"]
            for col in range(1, 11):
                ws.column_dimensions[get_column_letter(col)].width = max(ws.column_dimensions[get_column_letter(col)].width or 0, 11)
            ws.row_dimensions[5].height = 24

    @staticmethod
    def _short_heading_text(value: str) -> str:
        """Apply the requested short labels to visible headings only."""
        text = value.replace("_", " ")
        replacements = (
            ("This Year", "TY"), ("THIS YEAR", "TY"),
            ("Last Year", "LY"), ("LAST YEAR", "LY"),
            ("Previous FY", "LY"), ("Current FY", "TY"),
            ("Variance", "Var"), ("VARIANCE", "VAR"),
            ("Objective", "Obj"), ("OBJECTIVE", "OBJ"),
        )
        for old, new in replacements:
            text = text.replace(old, new)
        text = text.replace(" — CLICK TO OPEN", "").replace("— CLICK TO OPEN", "")
        text = text.replace(" — CLICK NAME", "").replace("— CLICK NAME", "")
        text = re.sub(r"\s+", " ", text).strip()
        return text

    def _apply_requested_heading_text(self) -> None:
        """Standardize all visible headings and replace explanatory notes."""
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str) or cell.value.startswith("="):
                        continue
                    is_heading = bool(cell.font.bold or cell.fill.fill_type == "solid" or (cell.font.sz or 0) >= 12)
                    is_note = bool(cell.font.italic and str(cell.font.color.type if cell.font.color else "") != "")
                    if is_heading:
                        cell.value = self._short_heading_text(cell.value)
                        # Main/section headings should not contain Additional Qty wording.
                        if (cell.font.sz or 0) >= 12:
                            cell.value = re.sub(r"\s*[—-]?\s*ADDITIONAL QTY\s*\(KLS?\)", "", cell.value, flags=re.I).strip(" —-")
                    elif is_note:
                        cell.value = "Sales in KL"

            # Apply the same abbreviations to chart titles and axis titles when possible.
            for chart in ws._charts:
                for obj in (getattr(chart, "title", None), getattr(getattr(chart, "x_axis", None), "title", None), getattr(getattr(chart, "y_axis", None), "title", None)):
                    try:
                        paras = obj.tx.rich.p
                        for para in paras:
                            for run in para.r:
                                if run.t:
                                    run.t = self._short_heading_text(run.t)
                    except Exception:
                        pass

    def _apply_requested_display_formats(self) -> None:
        """Apply product-specific rounding and stronger dashboard headings everywhere."""
        fuels = {"PMG", "HSD", "R95"}
        for ws in self.wb.worksheets:
            title_upper = ws.title.upper()
            sheet_product = None
            for product in ("PMG", "HSD", "R95", "LUBRICANTS"):
                if title_upper == product or title_upper.startswith(product + "_"):
                    sheet_product = product
                    break

            current_context = sheet_product
            for row in ws.iter_rows():
                row_product = None
                for cell in row[:6]:
                    text = str(cell.value or "").strip().upper()
                    if text in fuels or text == "LUBRICANTS":
                        row_product = text
                        break
                    for product in ("PMG", "HSD", "R95", "LUBRICANTS"):
                        if text.startswith(product + " —") or text.startswith(product + "-"):
                            row_product = product
                            break
                    if row_product:
                        break
                if row_product:
                    current_context = row_product

                for cell in row:
                    fmt = str(cell.number_format or "General")
                    if "%" in fmt or any(token in fmt.lower() for token in ("dd", "mmm", "yy")):
                        continue
                    value = cell.value
                    if not (isinstance(value, Number) or (isinstance(value, str) and value.startswith("="))):
                        continue
                    context = row_product or sheet_product or current_context
                    if context in fuels:
                        cell.number_format = "#,##0"
                    elif context == "LUBRICANTS" or "LUBRICANT" in title_upper:
                        cell.number_format = "#,##0.00"

            # Main/DDM dashboard table headings: larger, bold and fully visible.
            if ws.title == "Dashboard" or ws.title.startswith("DDM_"):
                for row in ws.iter_rows():
                    for cell in row:
                        text = str(cell.value or "").strip().casefold()
                        if text in {"ty sales","ly sales","var ly","var ly%","obj","ty vs obj","ty vs obj%","avg/day","working days"}:
                            cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
                            cell.fill = self.dark_fill
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            cell.border = self.border
                            ws.row_dimensions[cell.row].height = max(ws.row_dimensions[cell.row].height or 0, 32)

    def save(self) -> Path:
        """Save the generated dashboard efficiently for cloud deployment.

        The workbook already receives formatting while each sheet is created.
        The previous version performed several additional full-workbook scans
        here, including formula graph construction for every formula cell.
        Those scans are very expensive for large dashboards and can cause
        Streamlit Community Cloud timeouts.
        """
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        print("PROGRESS: Writing Excel workbook to disk...", flush=True)
        self.wb.calculation.iterate = False
        self.wb.save(OUTPUT_FILE)
        print(f"PROGRESS: Dashboard created: {OUTPUT_FILE}", flush=True)
        return OUTPUT_FILE
