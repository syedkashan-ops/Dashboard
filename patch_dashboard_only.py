from openpyxl import load_workbook
from excel_writer import ExcelWriter
from config import OUTPUT_FILE

writer = ExcelWriter()
writer.wb = load_workbook(OUTPUT_FILE)
for name in ('Dashboard', 'Dashboard_Data'):
    if name in writer.wb.sheetnames:
        del writer.wb[name]
writer.create_dashboard()
writer.wb.calculation.fullCalcOnLoad = True
writer.wb.calculation.forceFullCalc = True
writer.wb.calculation.calcMode = 'auto'
writer.wb.save(OUTPUT_FILE)
print(OUTPUT_FILE)
