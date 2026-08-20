from openpyxl import load_workbook
from config import OUTPUT_FILE, PRODUCTS, AREAS
from excel_writer import ExcelWriter

print('Loading existing workbook...')
base = load_workbook(OUTPUT_FILE)
print('Loading processor and rebuilding control sheets...')
writer = ExcelWriter()
writer.wb = base
# Remove old control/helper sheets and rebuild them safely.
for name in ['Control_Data'] + list(PRODUCTS) + [f'{p}_{a}' for p in PRODUCTS for a in AREAS]:
    if name in writer.wb.sheetnames:
        del writer.wb[name]
writer._create_control_data_sheet()
for product in PRODUCTS:
    writer.create_product_sheet(product)
for product in PRODUCTS:
    for area in AREAS:
        writer.create_product_area_sheet(product, area)
writer.wb.calculation.fullCalcOnLoad = True
writer.wb.calculation.forceFullCalc = True
writer.wb.calculation.calcMode = 'auto'
writer.wb.save(OUTPUT_FILE)
print(f'Patched: {OUTPUT_FILE}')
