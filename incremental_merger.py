from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from shutil import copy2, move
from typing import Any

from openpyxl import load_workbook

from config import BASE_DIR, INPUT_DIR


DAILY_INPUT_DIR = BASE_DIR / "Daily_Input"
ARCHIVE_DIR = DAILY_INPUT_DIR / "Processed"
BACKUP_DIR = BASE_DIR / "Backup"

EXCEL_PATTERNS = ("*.xlsx", "*.xlsm")

REQUIRED_SALES_HEADERS = {
    "Material",
    "Outlet Name",
    "Sales Group",
}

QUANTITY_HEADERS = (
    "Addtional Qty",
    "Additional Qty",
)

DATE_HEADERS = (
    "Billing Date",
    "Services Rendered On",
)


def _normalise_header(value: Any) -> str:
    return " ".join(
        str(value or "")
        .replace("\n", " ")
        .split()
    ).strip()


def _normalise_key_value(value: Any) -> Any:
    """
    Convert values into a stable representation for duplicate
    comparison without changing the original Excel values.
    """

    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, bool):
        return str(value).upper()

    if isinstance(value, float):

        if value.is_integer():
            return str(int(value))

        return f"{value:.8f}".rstrip("0").rstrip(".")

    if isinstance(value, int):
        return str(value)

    return str(value).strip().upper()


def _find_header(ws) -> tuple[int, list[str], dict[str, int]]:

    for row_no in range(
        1,
        min(ws.max_row or 15, 15) + 1,
    ):

        values = [
            _normalise_header(
                ws.cell(row_no, col).value
            )
            for col in range(
                1,
                (ws.max_column or 1) + 1,
            )
        ]

        mapping = {
            name: idx
            for idx, name in enumerate(values)
            if name
        }

        has_date = any(
            name in mapping
            for name in DATE_HEADERS
        )

        has_quantity = any(
            name in mapping
            for name in QUANTITY_HEADERS
        )

        if (
            has_date
            and REQUIRED_SALES_HEADERS.issubset(mapping)
            and has_quantity
        ):
            return row_no, values, mapping

    raise ValueError(
        f"Sales header was not found "
        f"in sheet '{ws.title}'."
    )


def _sales_files(folder: Path) -> list[Path]:

    files: list[Path] = []

    for pattern in EXCEL_PATTERNS:
        files.extend(
            folder.glob(pattern)
        )

    return sorted(
        p
        for p in files
        if (
            not p.name.startswith("~$")
            and p.is_file()
        )
    )


def _is_objective_workbook(path: Path) -> bool:

    try:

        wb = load_workbook(
            path,
            read_only=True,
            data_only=True,
        )

        try:

            names = {
                name.strip().lower()
                for name in wb.sheetnames
            }

            return (
                "hsd" in names
                and "pmg" in names
                and bool(
                    names.intersection(
                        {
                            "hobc",
                            "r95",
                            "r-95",
                        }
                    )
                )
            )

        finally:
            wb.close()

    except Exception:
        return False


def _workbook_date_range(
    path: Path,
) -> tuple[datetime, datetime] | None:

    from processor import (
        _fast_xlsx_rows,
        _to_date,
    )

    headers: dict[str, int] | None = None
    header_row = 0

    for row_no, row in _fast_xlsx_rows(path):

        if row_no > 15:
            break

        candidate = {
            _normalise_header(value): idx
            for idx, value in enumerate(row)
            if _normalise_header(value)
        }

        has_date = any(
            name in candidate
            for name in DATE_HEADERS
        )

        has_quantity = any(
            name in candidate
            for name in QUANTITY_HEADERS
        )

        if (
            has_date
            and REQUIRED_SALES_HEADERS.issubset(candidate)
            and has_quantity
        ):

            headers = candidate
            header_row = row_no
            break

    if not headers:
        return None

    date_index = headers.get(
        "Billing Date",
        headers.get(
            "Services Rendered On"
        ),
    )

    low: datetime | None = None
    high: datetime | None = None

    for row_no, row in _fast_xlsx_rows(path):

        if (
            row_no <= header_row
            or date_index is None
            or date_index >= len(row)
        ):
            continue

        value = _to_date(
            row[date_index]
        )

        if not value:
            continue

        low = (
            value
            if low is None or value < low
            else low
        )

        high = (
            value
            if high is None or value > high
            else high
        )

    return (
        (low, high)
        if low and high
        else None
    )


def find_current_year_master() -> Path:

    candidates: list[
        tuple[datetime, Path]
    ] = []

    for path in _sales_files(INPUT_DIR):

        if _is_objective_workbook(path):
            continue

        period = _workbook_date_range(
            path
        )

        if period:
            candidates.append(
                (
                    period[1],
                    path,
                )
            )

    if len(candidates) < 2:
        raise ValueError(
            "Two sales workbooks were not found "
            "in the Input folder."
        )

    candidates.sort(
        key=lambda item: item[0]
    )

    return candidates[-1][1]


def _dedupe_columns(
    headers: dict[str, int],
) -> list[int]:
    """
    Build a stable transaction identity.

    Do NOT use quantity or Net Value in the duplicate key.
    Those values may be represented differently between SAP
    exports even when the underlying transaction is the same.

    Preferred identity:
        Billing Doc
        Reference Item
        Material

    Additional fields are used only where available to
    distinguish unusual SAP line structures.
    """

    preferred = [
        "Billing Doc",
        "Reference Item",
        "Material",
        "Billing Date",
        "Sold to party",
    ]

    indices = [
        headers[name]
        for name in preferred
        if name in headers
    ]

    # Billing Document + Material is required for
    # transaction-based duplicate detection.
    if (
        "Billing Doc" in headers
        and "Material" in headers
        and len(indices) >= 2
    ):
        return indices

    # Fallback for unusual files.
    return list(
        range(len(headers))
    )


def _row_key(
    values: list[Any],
    key_indices: list[int],
) -> tuple[Any, ...]:

    return tuple(
        _normalise_key_value(
            values[i]
            if i < len(values)
            else None
        )
        for i in key_indices
    )


def merge_daily_sales() -> dict[str, Any]:
    """
    Append only genuinely new SAP transactions to the
    current-year master.

    Re-uploading the same file should result in:
        Added = 0
        Duplicates = all uploaded rows
    """

    DAILY_INPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    ARCHIVE_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    BACKUP_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    daily_files = _sales_files(
        DAILY_INPUT_DIR
    )

    if not daily_files:

        return {
            "master": None,
            "files": 0,
            "added": 0,
            "duplicates": 0,
            "message":
                "No daily SAP file found.",
        }

    master_path = (
        find_current_year_master()
    )

    print(
        f"\nIncremental update master: "
        f"{master_path.name}"
    )

    keep_vba = (
        master_path.suffix.lower()
        == ".xlsm"
    )

    master_wb = load_workbook(
        master_path,
        read_only=False,
        data_only=False,
        keep_vba=keep_vba,
    )

    master_ws = master_wb[
        master_wb.sheetnames[0]
    ]

    (
        header_row,
        master_headers,
        master_map,
    ) = _find_header(master_ws)

    key_indices = _dedupe_columns(
        master_map
    )

    existing: set[
        tuple[Any, ...]
    ] = set()

    # ---------------------------------------------
    # Load existing transaction identities
    # ---------------------------------------------

    for row in master_ws.iter_rows(
        min_row=header_row + 1,
        max_col=len(master_headers),
        values_only=True,
    ):

        values = list(row)

        if not any(
            value is not None
            for value in values
        ):
            continue

        existing.add(
            _row_key(
                values,
                key_indices,
            )
        )

    print(
        f"Existing transaction keys: "
        f"{len(existing):,}"
    )

    added = 0
    duplicates = 0
    processed: list[Path] = []

    try:

        for daily_path in daily_files:

            print(
                f"Reading daily SAP file: "
                f"{daily_path.name}"
            )

            wb = load_workbook(
                daily_path,
                read_only=True,
                data_only=False,
            )

            try:

                ws = wb[
                    wb.sheetnames[0]
                ]

                (
                    daily_header_row,
                    _,
                    daily_map,
                ) = _find_header(ws)

                missing = [
                    name
                    for name in master_map
                    if name not in daily_map
                ]

                missing_core = [
                    name
                    for name
                    in REQUIRED_SALES_HEADERS
                    if name not in daily_map
                ]

                if not any(
                    name in daily_map
                    for name in QUANTITY_HEADERS
                ):

                    missing_core.append(
                        "Addtional Qty / Additional Qty"
                    )

                if not any(
                    name in daily_map
                    for name in DATE_HEADERS
                ):

                    missing_core.append(
                        "Billing Date/"
                        "Services Rendered On"
                    )

                if missing_core:

                    raise ValueError(
                        f"'{daily_path.name}' "
                        f"is missing required columns: "
                        f"{', '.join(missing_core)}"
                    )

                file_added = 0
                file_duplicates = 0

                for row in ws.iter_rows(
                    min_row=daily_header_row + 1,
                    values_only=True,
                ):

                    if not any(
                        value is not None
                        for value in row
                    ):
                        continue

                    output_values = [
                        (
                            row[daily_map[name]]
                            if (
                                name in daily_map
                                and daily_map[name]
                                < len(row)
                            )
                            else None
                        )
                        for name
                        in master_headers
                    ]

                    key = _row_key(
                        output_values,
                        key_indices,
                    )

                    if key in existing:

                        duplicates += 1
                        file_duplicates += 1
                        continue

                    master_ws.append(
                        output_values
                    )

                    existing.add(key)

                    added += 1
                    file_added += 1

                print(
                    f"  Added "
                    f"{file_added:,}; "
                    f"skipped "
                    f"{file_duplicates:,} "
                    f"duplicate row(s)."
                )

                if missing:

                    print(
                        "  Optional columns "
                        "not present and left blank: "
                        f"{', '.join(missing)}"
                    )

                processed.append(
                    daily_path
                )

            finally:
                wb.close()

        # ---------------------------------------------
        # Save master only if genuinely new rows exist
        # ---------------------------------------------

        if added:

            timestamp = (
                datetime.now().strftime(
                    "%Y%m%d_%H%M%S"
                )
            )

            backup_path = (
                BACKUP_DIR
                / (
                    f"{master_path.stem}"
                    f"_before_{timestamp}"
                    f"{master_path.suffix}"
                )
            )

            copy2(
                master_path,
                backup_path,
            )

            master_wb.save(
                master_path
            )

            # Invalidate processor cache so the
            # dashboard sees newly merged rows
            # immediately in the same run.
            try:

                from processor import (
                    _FAST_ROW_CACHE
                )

                _FAST_ROW_CACHE.pop(
                    str(
                        master_path.resolve()
                    ),
                    None,
                )

            except Exception:
                pass

            print(
                f"Master updated: "
                f"{master_path.name} "
                f"(+{added:,} row(s))"
            )

            print(
                f"Backup created: "
                f"{backup_path.relative_to(BASE_DIR)}"
            )

        else:

            print(
                "Master already contains all "
                "imported rows. "
                "No new data was added."
            )

    finally:
        master_wb.close()

    # ---------------------------------------------
    # Archive processed imports
    # ---------------------------------------------

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    for source in processed:

        destination = (
            ARCHIVE_DIR
            / (
                f"{source.stem}_"
                f"{timestamp}"
                f"{source.suffix}"
            )
        )

        counter = 1

        while destination.exists():

            destination = (
                ARCHIVE_DIR
                / (
                    f"{source.stem}_"
                    f"{timestamp}_"
                    f"{counter}"
                    f"{source.suffix}"
                )
            )

            counter += 1

        move(
            str(source),
            str(destination),
        )

    return {
        "master": master_path,
        "files": len(processed),
        "added": added,
        "duplicates": duplicates,
        "message":
            "Incremental SAP merge completed.",
    }
