from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from zipfile import ZipFile
import re
import calendar
from xml.etree import ElementTree as etree
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.datetime import from_excel
from statistics import median
from typing import Any

from openpyxl import load_workbook

from config import (
    INPUT_DIR,
    OUTPUT_DIR,
    FISCAL_MONTHS,
    PRODUCTS,
    AREAS,
    OBJECTIVE_SHEET_ALIASES,
    PRODUCT_MATERIAL_ALIASES,
    LUBRICANT_MATERIAL_CODES,
    OUTLET_MASTER_FILE,
    LUBRICANT_LITER_MASTER_FILE,
)

EXCEL_PATTERNS = ("*.xlsx", "*.xlsm")
STANDARD_AREAS = {f"A{i:02d}" for i in range(1, 9)}
SUMMARY_COLUMNS = [
    "Product",
    "This Year Sales",
    "Last Year Sales",
    "Objective",
    "TY vs LY Variance",
    "TY vs LY %",
    "TY vs Objective Variance",
    "TY vs Objective %",
]


def _normalise(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().replace("\n", " ")
    return text[:-2] if text.endswith(".0") else text


def _outlet_key(value: Any) -> str:
    """Normalise outlet names so target names match sales names reliably."""
    text = _normalise(value).upper()
    return "".join(ch for ch in text if ch.isalnum())


def _normalise_code(value: Any) -> str:
    text = _normalise(value)
    if not text:
        return ""
    try:
        return str(int(float(text)))
    except (TypeError, ValueError):
        return text.upper().replace(" ", "")


def _to_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _to_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)):
        try:
            return from_excel(value)
        except (TypeError, ValueError, OverflowError):
            return None
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                pass
    return None



_FAST_ROW_CACHE: dict[str, list[tuple[int, list[Any]]]] = {}

def _fast_xlsx_rows(path: Path):
    """Stream and cache the first worksheet from xlsx."""
    cache_key = str(path.resolve())
    if cache_key in _FAST_ROW_CACHE:
        yield from _FAST_ROW_CACHE[cache_key]
        return
    cached_rows: list[tuple[int, list[Any]]] = []
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with ZipFile(path) as z:
        strings: list[str] = []
        if "xl/sharedStrings.xml" in z.namelist():
            with z.open("xl/sharedStrings.xml") as source:
                for _, element in etree.iterparse(source, events=("end",)):
                    if element.tag == ns + "si":
                        strings.append("".join(element.itertext()))
                        element.clear()
        sheet_names = sorted(
            name for name in z.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        if not sheet_names:
            return
        with z.open(sheet_names[0]) as source:
            for _, row in etree.iterparse(source, events=("end",)):
                if row.tag != ns + "row":
                    continue
                cells: dict[int, Any] = {}
                for cell in row.findall(ns + "c"):
                    ref = cell.get("r", "A1")
                    match = re.match(r"[A-Z]+", ref)
                    if not match:
                        continue
                    col = column_index_from_string(match.group()) - 1
                    cell_type = cell.get("t")
                    node = cell.find(ns + "v")
                    value: Any = None if node is None else node.text
                    if cell_type == "s" and value is not None:
                        value = strings[int(value)]
                    elif cell_type == "inlineStr":
                        value = "".join(cell.itertext())
                    elif cell_type == "b":
                        value = value == "1"
                    elif value is not None:
                        try:
                            value = float(value)
                        except ValueError:
                            pass
                    cells[col] = value
                if cells:
                    max_col = max(cells)
                    item = (int(row.get("r", "0")), [cells.get(i) for i in range(max_col + 1)])
                    cached_rows.append(item)
                    yield item
                row.clear()
    _FAST_ROW_CACHE[cache_key] = cached_rows


def _excel_files() -> list[Path]:
    files: list[Path] = []
    for pattern in EXCEL_PATTERNS:
        files.extend(INPUT_DIR.glob(pattern))
    return sorted({p.resolve() for p in files if not p.name.startswith("~$")})


def _objective_workbook_info(path: Path) -> tuple[str, dict[str, str]] | None:
    """Identify objective workbooks by content, not filename.

    Returns ("outlet", sheet_map) for the outlet-level layout containing
    Area/Name/Cost Center/Jul..Jun, or ("area", sheet_map) for the official
    area-summary layout containing Sales Area and monthly columns.
    """
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        lowered = {name.strip().lower(): name for name in wb.sheetnames}
        mapping: dict[str, str] = {}
        for product, aliases in OBJECTIVE_SHEET_ALIASES.items():
            actual = next((lowered[a.lower()] for a in aliases if a.lower() in lowered), None)
            if actual is None:
                return None
            mapping[product] = actual

        outlet_count = 0
        area_count = 0
        for sheet_name in set(mapping.values()):
            ws = wb[sheet_name]
            headers = {_normalise(ws.cell(1, c).value).casefold() for c in range(1, min(ws.max_column, 25) + 1)}
            if {"area", "name", "cost center"}.issubset(headers):
                outlet_count += 1
            if "sales area" in headers:
                area_count += 1
        if outlet_count == len(set(mapping.values())):
            return "outlet", mapping
        if area_count == len(set(mapping.values())):
            return "area", mapping
        return None
    finally:
        wb.close()


def _objective_sheet_map(path: Path) -> dict[str, str] | None:
    """Backward-compatible helper used by older project utilities."""
    info = _objective_workbook_info(path)
    return info[1] if info else None

def _find_sales_header(path: Path) -> tuple[str, int, dict[str, int]] | None:
    try:
        for row_number, row in _fast_xlsx_rows(path):
            if row_number > 12:
                break
            headers = {_normalise(value): idx for idx, value in enumerate(row)}
            if "Billing Date" in headers or "Services Rendered On" in headers:
                required = {"Material", "Outlet Name", "Sales Group"}
                has_qty = any(name in headers for name in ("Addtional Qty", "Additional Qty"))
                if required.issubset(headers) and has_qty:
                    return "__FAST_FIRST_SHEET__", row_number, headers
        return None
    except Exception:
        return None


def _sales_date_score(path: Path, sheet: str, header_row: int, headers: dict[str, int]) -> float:
    date_name = "Billing Date" if "Billing Date" in headers else "Services Rendered On"
    date_index = headers[date_name]
    ordinals: list[int] = []
    for row_number, row in _fast_xlsx_rows(path):
        if row_number <= header_row or date_index >= len(row):
            continue
        dt = _to_date(row[date_index])
        if dt:
            ordinals.append(dt.toordinal())
    if not ordinals:
        raise ValueError(f"No valid sales dates were found in '{path.name}'.")
    return float(median(ordinals))


def detect_input_files() -> tuple[dict[str, Any], dict[str, Any], Path, dict[str, str], Path, dict[str, str]]:
    files = _excel_files()
    if len(files) < 4:
        raise ValueError(
            "Put two sales workbooks, one outlet-objective workbook and one area-objective workbook in the Input folder. "
            f"Only {len(files)} Excel file(s) were found."
        )

    outlet_objectives: list[tuple[Path, dict[str, str]]] = []
    area_objectives: list[tuple[Path, dict[str, str]]] = []
    sales: list[dict[str, Any]] = []
    for path in files:
        objective_info = _objective_workbook_info(path)
        if objective_info:
            kind, mapping = objective_info
            if kind == "outlet":
                outlet_objectives.append((path, mapping))
            else:
                area_objectives.append((path, mapping))
            continue
        found = _find_sales_header(path)
        if found:
            sheet, header_row, headers = found
            print(f"Checking sales period: {path.name}")
            score = _sales_date_score(path, sheet, header_row, headers)
            sales.append({
                "path": path,
                "sheet": sheet,
                "header_row": header_row,
                "headers": headers,
                "score": score,
            })

    if len(outlet_objectives) != 1:
        names = ", ".join(p.name for p, _ in outlet_objectives) or "none"
        raise ValueError(
            "Exactly one OUTLET objective workbook is required. It must contain product sheets with "
            "Area, Name, Cost Center and Jul-Jun columns. The filename can be anything. "
            f"Detected: {names}"
        )
    if len(area_objectives) != 1:
        names = ", ".join(p.name for p, _ in area_objectives) or "none"
        raise ValueError(
            "Exactly one AREA objective workbook is required. It must contain PMG, HSD, HOBC/R95 and "
            "Lubes/LUBE sheets with a Sales Area summary. The filename can be anything. "
            f"Detected: {names}"
        )
    if len(sales) != 2:
        names = ", ".join(item["path"].name for item in sales) or "none"
        raise ValueError(f"Exactly two sales workbooks are required. Detected: {names}")

    sales.sort(key=lambda item: item["score"])
    outlet_path, outlet_map = outlet_objectives[0]
    area_path, area_map = area_objectives[0]
    return sales[0], sales[1], outlet_path, outlet_map, area_path, area_map


class SalesProcessor:
    """Sales processing implemented only with openpyxl and Python dictionaries."""

    def __init__(self) -> None:
        INPUT_DIR.mkdir(parents=True, exist_ok=True)
        last_info, this_info, outlet_objective_path, outlet_sheet_map, area_objective_path, area_sheet_map = detect_input_files()
        self.last_year_file = last_info["path"]
        self.this_year_file = this_info["path"]
        self.outlet_objective_file = outlet_objective_path
        self.outlet_objective_sheet_map = outlet_sheet_map
        self.area_objective_file = area_objective_path
        self.area_objective_sheet_map = area_sheet_map
        # Retain legacy attribute names for any external customisations.
        self.objective_file = area_objective_path
        self.objective_sheet_map = area_sheet_map

        self.start_date: datetime | None = None
        self.end_date: datetime | None = None

        # Load the permanent outlet master before sales. All outlet-level calculations
        # are keyed by Cost Center, while code/name remain display attributes.
        self.code_to_cost_center: dict[str, str] = {}
        self.cost_center_info: dict[str, dict[str, Any]] = {}
        # Keep every historical outlet-name alias for dynamic Targets FY26 -> Cost Center resolution.
        self.cost_center_alias_names: dict[str, set[str]] = defaultdict(set)
        self.unmapped_outlets: dict[str, dict[str, Any]] = {}
        # Outlet Mapping is the single source of truth. Rebuild the convenience
        # summary automatically so users never maintain the same outlet twice.
        self._sync_cost_center_summary(OUTLET_MASTER_FILE)
        self._read_outlet_master(OUTLET_MASTER_FILE)
        # Open Orders > Info is the user-maintained source for the latest outlet
        # display code/name.  Historical aliases from Outlet_Master remain valid,
        # while the newest Info row controls dropdown labels and report display.
        self._read_open_orders_outlet_info()

        # A00 is built from the CoCo/CoRos outlet-name section in the target file.
        # These outlets are removed from their original A01-A08 sales groups.
        self.a00_outlet_names = self._read_a00_outlets(outlet_objective_path, outlet_sheet_map)
        # Some official CoRos/COCO outlets can exist in Outlet_Master even when a
        # monthly objective workbook accidentally omits their row. Include every
        # outlet explicitly marked COCO in its maintained master name so it remains
        # visible in A00 with zero sales/objective rather than disappearing.
        master_coco_names = {
            _normalise(info.get("Outlet Name"))
            for info in self.cost_center_info.values()
            if re.search(r"\bCOCO\b", _normalise(info.get("Outlet Name")), flags=re.IGNORECASE)
        }
        self.a00_outlet_names.update(name for name in master_coco_names if name)
        self.a00_outlet_keys = {_outlet_key(name) for name in self.a00_outlet_names}
        # Resolve the current CoRos list dynamically to permanent Cost Centers.
        # Nothing is hard-coded: additions/removals in Targets FY26 take effect on every run.
        self.a00_cost_centers = self._resolve_a00_cost_centers()

        # Keep every outlet listed in Targets FY26 visible in A00, even when it
        # has no sales row in either year. Resolved outlets use their permanent
        # Cost Center; unresolved target names receive a stable report-only key.
        # This guarantees that all target CoRos rows appear with zero sales.
        resolved_target_keys = set()
        for cc in self.a00_cost_centers:
            for alias in self.cost_center_alias_names.get(cc, set()):
                resolved_target_keys.add(_outlet_key(alias))
            resolved_target_keys.add(_outlet_key(self.cost_center_info.get(cc, {}).get("Outlet Name", "")))
        for target_name in sorted(self.a00_outlet_names):
            if _outlet_key(target_name) in resolved_target_keys:
                continue
            synthetic_cc = f"A00::{target_name}"
            self.a00_cost_centers.add(synthetic_cc)
            self.cost_center_info.setdefault(synthetic_cc, {
                "Outlet Name": target_name,
                "Outlet Code": "",
                "Area": "A00",
            })
            self.cost_center_alias_names[synthetic_cc].add(target_name)

        print(f"Detected {len(self.a00_outlet_names)} A00/CoCo outlet name(s) and {len(self.a00_cost_centers)} Cost Center/report key(s) from target workbook.")
        for name in sorted(self.a00_outlet_names):
            print(f"  A00: {name}")

        # Totals: product and product/category.
        self.product_totals = {"last": defaultdict(float), "this": defaultdict(float)}
        self.area_totals = {"last": defaultdict(float), "this": defaultdict(float)}
        # Exact daily totals used by the automatic dashboard comparison engine.
        self.product_daily = {"last": defaultdict(float), "this": defaultdict(float)}
        self.area_daily = {"last": defaultdict(float), "this": defaultdict(float)}
        # Analytics and lubricant material drill-down stores.
        self.outlet_product_totals = {"last": defaultdict(float), "this": defaultdict(float)}
        self.outlet_monthly_totals = {"last": defaultdict(float), "this": defaultdict(float)}
        self.outlet_sold_to_party: dict[str, Any] = {}
        self.material_outlet_totals = {"last": defaultdict(float), "this": defaultdict(float)}
        # Daily and monthly Mat Group Text totals power the Division 20 Lubricant dashboards.
        self.material_group_daily = {"last": defaultdict(float), "this": defaultdict(float)}
        self.material_group_monthly = {"last": defaultdict(float), "this": defaultdict(float)}
        self.monthly_product_totals = {"last": defaultdict(float), "this": defaultdict(float)}
        self.monthly_area_product_totals = {"last": defaultdict(float), "this": defaultdict(float)}

        # Derived outlet master fields. Users only maintain Cost Center,
        # Sold to Party and Outlet Name. Area and Available In are inferred
        # from sales; DDM is inferred from TM Name.xlsx using the derived Area.
        self.observed_products = {"last": defaultdict(set), "this": defaultdict(set)}
        self.observed_area_counts = {"last": defaultdict(lambda: defaultdict(int)), "this": defaultdict(lambda: defaultdict(int))}

        # Current-year daily rows and last-year matching totals.
        # key = product, outlet, raw group, category, month, fiscal month number
        self.daily_this: dict[tuple[Any, ...], list[float]] = {}
        self.daily_last: dict[tuple[Any, ...], list[float]] = {}

        print(f"Reading Last Year: {self.last_year_file.name}")
        self._read_sales(last_info, "last")
        print(f"Reading This Year: {self.this_year_file.name}")
        self._read_sales(this_info, "this")

        # Complete Outlet Mapping automatically after sales have revealed each
        # outlet's Area and products. Then rebuild the summary from that source.
        self._sync_outlet_mapping_derived_fields(OUTLET_MASTER_FILE)
        self._sync_cost_center_summary(OUTLET_MASTER_FILE)

        if self.start_date is None or self.end_date is None:
            raise ValueError("No valid dates were found in the current-year sales workbook.")

        print(f"Reading Outlet Objectives: {outlet_objective_path.name}")
        self.outlet_objectives = self._read_outlet_objectives(outlet_objective_path, outlet_sheet_map)
        print(f"Reading Area Objectives: {area_objective_path.name}")
        self.objectives = self._read_area_objectives(area_objective_path, area_sheet_map)

        # Lubricant open orders and transit are optional snapshot files.
        self.lubricant_liter_master = self._read_lubricant_liter_master(LUBRICANT_LITER_MASTER_FILE)
        # Month-wise pipeline snapshots, keyed by (month_start, area).
        self.lubricant_open_by_month_area = defaultdict(float)
        self.lubricant_transit_by_month_area = defaultdict(float)
        self._read_lubricant_snapshots()
        self._daily_cache: dict[tuple[str, str | None], list[dict[str, Any]]] = {}
        self._period_cache: dict[tuple[str, str, date, date], float] = {}
        self._weekday_cache: dict[tuple[str, date, date], float] = {}
        self._working_dates_cache: dict[tuple[date, date], list[date]] = {}

        print(f"Last Year : {self.last_year_file.name}")
        print(f"This Year : {self.this_year_file.name}")
        print(f"Outlet Objective: {self.outlet_objective_file.name}")
        print(f"Area Objective  : {self.area_objective_file.name}")
        print(f"Report period: {self.start_date:%d-%b-%Y} to {self.end_date:%d-%b-%Y}")


    def _read_ddm_by_area(self) -> dict[str, str]:
        """Read Area -> DDM Name from Input/TM Name.xlsx."""
        path = INPUT_DIR / "TM Name.xlsx"
        if not path.exists():
            return {}
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb["Data"] if "Data" in wb.sheetnames else wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            first = next(rows, ())
            headers = {_normalise(v).casefold(): i for i, v in enumerate(first)}
            area_col = headers.get("area")
            ddm_col = headers.get("ddm name", headers.get("ddm"))
            if area_col is None or ddm_col is None:
                return {}
            result: dict[str, str] = {}
            for row in rows:
                if area_col >= len(row) or ddm_col >= len(row):
                    continue
                area = _normalise(row[area_col]).upper()
                ddm = _normalise(row[ddm_col])
                if area in AREAS and ddm:
                    result[area] = ddm
            return result
        finally:
            wb.close()

    def _derived_area_for_cost_center(self, cost_center: str) -> str:
        """Prefer current-year observed Area, then last year, then master."""
        for period in ("this", "last"):
            counts = self.observed_area_counts[period].get(cost_center, {})
            if counts:
                return sorted(counts.items(), key=lambda item: (-item[1], item[0]))[0][0]
        return _normalise(self.cost_center_info.get(cost_center, {}).get("Area"))

    def _derived_products_for_cost_center(self, cost_center: str) -> list[str]:
        """Prefer current-year products; use last year only when current is absent."""
        products = self.observed_products["this"].get(cost_center)
        if not products:
            products = self.observed_products["last"].get(cost_center, set())
        order = {name: i for i, name in enumerate(PRODUCTS)}
        return sorted(products, key=lambda name: (order.get(name, 999), name))

    def _sync_outlet_mapping_derived_fields(self, path: Path) -> None:
        """Auto-fill Area, DDM Name and Available In in Outlet Mapping.

        The only fields users need to enter are Cost Center, Sold to Party and
        Outlet Name. Derived fields are refreshed on every dashboard run.
        """
        if not path.exists():
            return
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = load_workbook(path)
        try:
            if "Outlet Mapping" not in wb.sheetnames:
                return
            ws = wb["Outlet Mapping"]
            first = [cell.value for cell in ws[1]]
            headers = {_normalise(v).casefold(): i + 1 for i, v in enumerate(first)}
            code_key = "sold to party" if "sold to party" in headers else "outlet code"
            required = {"cost center", code_key, "outlet name"}
            if not required.issubset(headers):
                return

            # Add derived columns if an older master does not contain them.
            for label, key in (("Area", "area"), ("DDM Name", "ddm name"), ("Available In", "available in")):
                if key not in headers:
                    col = ws.max_column + 1
                    ws.cell(1, col, label)
                    headers[key] = col
                    cell = ws.cell(1, col)
                    cell.fill = PatternFill("solid", fgColor="1F4E78")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ddm_by_area = self._read_ddm_by_area()
            changed = 0
            for row_num in range(2, ws.max_row + 1):
                code = _normalise_code(ws.cell(row_num, headers[code_key]).value)
                raw_cc = _normalise(ws.cell(row_num, headers["cost center"]).value)
                name = _normalise(ws.cell(row_num, headers["outlet name"]).value)
                if not code and not raw_cc and not name:
                    continue
                cost_center = raw_cc
                if not cost_center or cost_center.casefold() in {"#n/a", "n/a", "na", "none", "null", "-"}:
                    cost_center = self.code_to_cost_center.get(code, code or name)
                if not cost_center:
                    continue

                area = self._derived_area_for_cost_center(cost_center)
                products = self._derived_products_for_cost_center(cost_center)
                ddm = ddm_by_area.get(area, "")

                values = {
                    "area": area,
                    "ddm name": ddm,
                    "available in": ", ".join(products),
                }
                for key, value in values.items():
                    cell = ws.cell(row_num, headers[key])
                    if _normalise(cell.value) != value:
                        cell.value = value
                        changed += 1

            # Make the three user-input fields visually clear; derived columns
            # are labelled as automatic through the workbook Read Me sheet.
            ws.freeze_panes = "A2"
            if "Read Me" in wb.sheetnames:
                readme = wb["Read Me"]
                messages = {
                    "outlet mapping": "Enter only Cost Center, Sold to Party and Outlet Name. Area, DDM Name and Available In are generated automatically on every dashboard run.",
                    "cost center summary": "Automatically generated from Outlet Mapping on every dashboard run. Do not edit manually.",
                }
                found = set()
                for row in readme.iter_rows():
                    key = _normalise(row[0].value).casefold()
                    if key in messages:
                        row[1].value = messages[key]
                        found.add(key)
                for key, message in messages.items():
                    if key not in found:
                        readme.append([key.title(), message])

            wb.save(path)
            print(f"Auto-filled Outlet Mapping derived fields ({changed} cell update(s)).")
        finally:
            wb.close()


    def _sync_cost_center_summary(self, path: Path) -> None:
        """Rebuild ``Cost Center Summary`` from ``Outlet Mapping``.

        Users maintain only the detailed Outlet Mapping sheet.  The summary is
        a derived convenience view and is refreshed on every dashboard run.
        Existing sheet names are retained.
        """
        if not path.exists():
            return
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = load_workbook(path)
        try:
            if "Outlet Mapping" not in wb.sheetnames:
                return
            mapping = wb["Outlet Mapping"]
            first = next(mapping.iter_rows(min_row=1, max_row=1, values_only=True), ())
            headers = {_normalise(v).casefold(): i for i, v in enumerate(first)}
            code_header = "sold to party" if "sold to party" in headers else "outlet code"
            required = {"cost center", "outlet name", code_header}
            if not required.issubset(headers):
                return

            grouped: dict[str, dict[str, Any]] = {}
            order: list[str] = []
            for row in mapping.iter_rows(min_row=2, values_only=True):
                code = _normalise_code(row[headers[code_header]])
                raw_cc = _normalise(row[headers["cost center"]])
                name = _normalise(row[headers["outlet name"]])
                available = _normalise(row[headers["available in"]]) if "available in" in headers else ""
                if not code and not name and not raw_cc:
                    continue
                # Match dashboard identity logic: missing/#N/A Cost Center falls
                # back to Sold-to Party, preventing an ambiguous shared #N/A row.
                cc = raw_cc
                if not cc or cc.casefold() in {"#n/a", "n/a", "na", "none", "null", "-"}:
                    cc = code or raw_cc or name
                if not cc:
                    continue
                if cc not in grouped:
                    grouped[cc] = {"names": [], "codes": [], "available": []}
                    order.append(cc)
                rec = grouped[cc]
                if name and name not in rec["names"]:
                    rec["names"].append(name)
                if code and code not in rec["codes"]:
                    rec["codes"].append(code)
                for item in [x.strip() for x in available.split(",") if x.strip()]:
                    if item not in rec["available"]:
                        rec["available"].append(item)

            if "Cost Center Summary" in wb.sheetnames:
                summary = wb["Cost Center Summary"]
                summary.delete_rows(1, summary.max_row)
            else:
                summary = wb.create_sheet("Cost Center Summary")

            summary.append([
                "Cost Center", "All Historical Outlet Names",
                "All Historical Sold-to Parties", "Outlet Name Count",
                "Sold-to Count", "Available In"
            ])
            for cc in order:
                rec = grouped[cc]
                summary.append([
                    cc, "; ".join(rec["names"]), "; ".join(rec["codes"]),
                    len(rec["names"]), len(rec["codes"]), ", ".join(rec["available"])
                ])

            fill = PatternFill("solid", fgColor="1F4E78")
            white_bold = Font(color="FFFFFF", bold=True)
            thin = Side(style="thin", color="B7B7B7")
            for cell in summary[1]:
                cell.fill = fill
                cell.font = white_bold
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in summary.iter_rows():
                for cell in row:
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    if cell.row > 1:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
            widths = {"A": 18, "B": 48, "C": 34, "D": 18, "E": 15, "F": 28}
            for col, width in widths.items():
                summary.column_dimensions[col].width = width
            summary.freeze_panes = "A2"
            summary.auto_filter.ref = f"A1:F{max(summary.max_row, 1)}"

            if "Read Me" in wb.sheetnames:
                readme = wb["Read Me"]
                updated = False
                for row in readme.iter_rows():
                    if _normalise(row[0].value).casefold() == "cost center summary":
                        row[1].value = "Automatically generated from Outlet Mapping on every dashboard run. Do not edit manually."
                        updated = True
                        break
                if not updated:
                    readme.append(["Cost Center Summary", "Automatically generated from Outlet Mapping on every dashboard run. Do not edit manually."])

            wb.save(path)
            print(f"Refreshed Cost Center Summary automatically from Outlet Mapping ({len(order)} Cost Center row(s)).")
        finally:
            wb.close()


    def _read_outlet_master(self, path: Path) -> None:
        """Load historical outlet aliases, using Cost Center as the stable identity.

        The master may contain several rows for the same Cost Center because both
        Outlet Name and Sold-to Party can change. Every Sold-to Party remains an
        alias of that Cost Center. The last row for a Cost Center is treated as
        its current display name/code, so appending a new row updates the report.
        Where Cost Center is blank or #N/A, Sold-to Party itself is the identity.
        """
        if not path.exists():
            raise ValueError(
                f"Missing outlet master: {path.name}. Keep the mapping workbook in the Input folder."
            )
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            # Prefer the detailed mapping sheet when the supplied workbook also
            # contains summary/read-me sheets.
            ws = wb["Outlet Mapping"] if "Outlet Mapping" in wb.sheetnames else wb[wb.sheetnames[0]]
            headers = {_normalise(v).casefold(): i for i, v in enumerate(next(ws.iter_rows(values_only=True)))}
            code_header = "sold to party" if "sold to party" in headers else "outlet code"
            required = {"cost center", code_header, "outlet name"}
            if not required.issubset(headers):
                raise ValueError(
                    "Outlet_Master.xlsx must contain Cost Center, Outlet Name, and Sold to Party (or Outlet Code) columns."
                )
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = _normalise_code(row[headers[code_header]])
                raw_cost_center = _normalise(row[headers["cost center"]])
                name = _normalise(row[headers["outlet name"]])
                area = _normalise(row[headers["area"]]) if "area" in headers else ""
                current = _normalise(row[headers["is current"]]).casefold() if "is current" in headers else ""
                if not code:
                    continue
                # Cost Center is the permanent key. For blank/error values, use
                # Sold-to Party exactly as requested, rather than dropping sales.
                cost_center = raw_cost_center
                if not cost_center or cost_center.casefold() in {"#n/a", "n/a", "na", "none", "null", "-"}:
                    cost_center = code
                existing = self.code_to_cost_center.get(code)
                if existing and existing != cost_center:
                    raise ValueError(f"Sold-to Party {code} is mapped to more than one Cost Center: {existing}, {cost_center}.")
                self.code_to_cost_center[code] = cost_center
                if name:
                    self.cost_center_alias_names[cost_center].add(name)
                info = self.cost_center_info.setdefault(cost_center, {"Outlet Name": name or cost_center, "Outlet Code": code, "Area": area})
                # With no Is Current column, append-only maintenance is supported:
                # the newest/last row becomes the current label. Historical codes
                # remain mapped to the same stable Cost Center.
                if not current or current in {"yes", "y", "true", "1", "current", "active"}:
                    info.update({
                        "Outlet Name": name or info.get("Outlet Name", cost_center),
                        "Outlet Code": code,
                        "Area": area or info.get("Area", ""),
                    })
        finally:
            wb.close()
        if not self.code_to_cost_center:
            raise ValueError("Outlet_Master.xlsx contains no usable Sold-to Party mappings.")
        print(f"Loaded {len(self.code_to_cost_center)} Sold-to aliases across {len(self.cost_center_info)} stable outlet identities.")

    def _read_open_orders_outlet_info(self) -> None:
        """Overlay current outlet code/name from the Open Orders ``Info`` sheet.

        Supported columns are Cost Center, Outlet Name and either Outlet Code or
        Sold-to Party.  The existing Product/Liter-only Info layout is ignored.
        Rows are processed top-to-bottom, so appending an updated mapping makes
        that last row the current display record while all prior codes remain
        searchable aliases of the same stable Cost Center.
        """
        cost_aliases = ("cost center", "cost centre", "costcenter")
        name_aliases = ("outlet name", "customer name", "sold-to party name", "sold to party name")
        code_aliases = ("outlet code", "sold to party", "sold-to party", "customer code")

        for path in _excel_files():
            if "open" not in path.stem.casefold():
                continue
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
            except Exception:
                continue
            try:
                info_ws = next((x for x in wb.worksheets if x.title.casefold() == "info"), None)
                if info_ws is None:
                    continue
                rows = info_ws.iter_rows(values_only=True)
                first = next(rows, ())
                headers = {_normalise(v).casefold(): i for i, v in enumerate(first)}

                def find_col(aliases):
                    return next((headers[a] for a in aliases if a in headers), None)

                cc_col = find_col(cost_aliases)
                name_col = find_col(name_aliases)
                code_col = find_col(code_aliases)
                if cc_col is None or name_col is None or code_col is None:
                    # Product/Liter conversion sheets have no outlet mapping.
                    continue

                loaded = 0
                for row in rows:
                    if max(cc_col, name_col, code_col) >= len(row):
                        continue
                    cost_center = _normalise(row[cc_col])
                    name = _normalise(row[name_col])
                    code = _normalise_code(row[code_col])
                    if not cost_center or cost_center.casefold() in {"#n/a", "n/a", "na", "none", "null", "-"}:
                        cost_center = self.code_to_cost_center.get(code, code)
                    if not cost_center:
                        continue
                    if code:
                        existing = self.code_to_cost_center.get(code)
                        if existing and existing != cost_center:
                            print(f"Ignored conflicting Open Orders Info mapping for code {code}: {existing} vs {cost_center}")
                            continue
                        self.code_to_cost_center[code] = cost_center
                    if name:
                        self.cost_center_alias_names[cost_center].add(name)
                    info = self.cost_center_info.setdefault(cost_center, {
                        "Outlet Name": name or cost_center, "Outlet Code": code, "Area": ""
                    })
                    info.update({
                        "Outlet Name": name or info.get("Outlet Name", cost_center),
                        "Outlet Code": code or info.get("Outlet Code", ""),
                    })
                    loaded += 1
                if loaded:
                    print(f"Loaded {loaded} current outlet mapping row(s) from {path.name} > Info.")
            finally:
                wb.close()

    def _write_unmapped_validation_file(self) -> None:
        """Write a user-maintainable list of missing outlet-code mappings."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        path = OUTPUT_DIR / "Unmapped_Outlets.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Unmapped_Outlets"
        headers = [
            "Cost Center (fill)", "Outlet Code", "Outlet Name", "Area",
            "Last Year Qty", "This Year Qty", "Source Rows", "Action"
        ]
        ws.append(headers)
        for rec in sorted(self.unmapped_outlets.values(), key=lambda r: (str(r.get("Area", "")), str(r.get("Outlet Name", "")))):
            ws.append([
                "", rec.get("Outlet Code", ""), rec.get("Outlet Name", ""), rec.get("Area", ""),
                rec.get("Last Year Qty", 0.0), rec.get("This Year Qty", 0.0), rec.get("Rows", 0),
                "Fill Cost Center, then add this row to Input/Outlet_Master.xlsx",
            ])
        fill = PatternFill("solid", fgColor="1F4E78")
        thin = Side(style="thin", color="B7B7B7")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        widths = [20, 16, 38, 10, 16, 16, 12, 58]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        ws.freeze_panes = "A2"
        wb.save(path)

    def _display_info(self, cost_center: str) -> dict[str, Any]:
        return self.cost_center_info.get(cost_center, {"Outlet Name": cost_center, "Outlet Code": "", "Area": ""})

    @staticmethod
    def _product_for(material: str, material_text: str) -> str | None:
        material_upper = material.upper().strip()
        text_upper = material_text.upper().strip()

        # Lubricants are identified by the known material codes or text.
        if material in LUBRICANT_MATERIAL_CODES or "LUB" in text_upper:
            return "Lubricants"

        # Match both Material and Material Text. Some CoRos rows contain a
        # numeric material code while the product name is only in Material Text.
        product_aliases = {
            "HSD": ("HSD", "HIGH SPEED DIESEL", "DIESEL"),
            "PMG": ("PMG", "PREMIER MOTOR GASOLINE", "MOTOR GASOLINE", "PETROL"),
            "R95": ("R95", "R-95", "HOBC", "ALTRON X", "RON 95"),
        }
        for product, aliases in product_aliases.items():
            if material_upper in aliases or any(alias in text_upper for alias in aliases):
                return product

        # Retain aliases configured for exact Material values.
        for product in ("HSD", "PMG", "R95"):
            aliases = {alias.upper() for alias in PRODUCT_MATERIAL_ALIASES[product]}
            if material_upper in aliases:
                return product
        return None

    def _resolve_a00_cost_centers(self) -> set[str]:
        """Resolve Targets FY26 CoRos names to stable Cost Centers.

        Exact normalized-name matches are preferred. A conservative prefix fallback
        is used only when it resolves to one Cost Center, supporting truncated target
        labels without accidentally moving similarly named outlets.
        """
        alias_index: dict[str, set[str]] = defaultdict(set)
        for cost_center, aliases in self.cost_center_alias_names.items():
            for alias in aliases:
                key = _outlet_key(alias)
                if key:
                    alias_index[key].add(cost_center)

        resolved: set[str] = set()
        unresolved: list[str] = []
        for target_name in sorted(self.a00_outlet_names):
            target = _outlet_key(target_name)
            matches = set(alias_index.get(target, set()))
            if not matches and target:
                for alias_key, cost_centers in alias_index.items():
                    if alias_key.startswith(target) or target.startswith(alias_key):
                        matches.update(cost_centers)
            if len(matches) == 1:
                resolved.update(matches)
            elif len(matches) > 1:
                print(f"WARNING: Ambiguous CoRos target name not assigned automatically: {target_name} -> {sorted(matches)}")
            else:
                unresolved.append(target_name)
        for target_name in unresolved:
            print(f"WARNING: CoRos target name has no Cost Center match in Outlet_Master: {target_name}")
        return resolved

    def _is_a00_outlet(self, outlet: str) -> bool:
        key = _outlet_key(outlet)
        if not key:
            return False
        return any(key == target or key.startswith(target) or target.startswith(key)
                   for target in self.a00_outlet_keys)

    def _area_category(self, raw_group: str, outlet: str, cost_center: str = "") -> str:
        # Outlet membership takes priority over Sales Group. Therefore an A00 outlet
        # is removed from A01-A08 even when its source row still carries that group.
        if cost_center in self.a00_cost_centers or self._is_a00_outlet(outlet):
            return "A00"
        upper = raw_group.upper().replace(" ", "")
        if upper in STANDARD_AREAS:
            return upper
        # Unexpected/unassigned groups do not become A00 unless their outlet is
        # explicitly listed in the target workbook's CoCo section.
        return upper or "Unassigned"

    def _read_a00_outlets(self, path: Path, sheet_map: dict[str, str]) -> set[str]:
        """Read all A00/CoRos outlet names from either objective layout."""
        names: set[str] = set()
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_name in set(sheet_map.values()):
                ws = wb[sheet_name]
                headers = {_normalise(ws.cell(1, c).value).casefold(): c for c in range(1, ws.max_column + 1)}
                # New outlet-level layout: Area | Name | Code | Cost Center | Jul..Jun.
                if {"area", "name", "cost center"}.issubset(headers):
                    area_col, name_col = headers["area"], headers["name"]
                    max_col = max(area_col, name_col)
                    for values in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
                        area = _normalise(values[area_col - 1] if area_col - 1 < len(values) else None).upper()
                        name = _normalise(values[name_col - 1] if name_col - 1 < len(values) else None)
                        if area == "A00" and name and name.casefold() != "total":
                            names.add(name)
                    continue

                # Legacy target layout.
                for row_no in range(12, ws.max_row + 1):
                    value = _normalise(ws.cell(row_no, 2).value)
                    if not value:
                        continue
                    low = value.casefold()
                    if low == "total" or "wise-target" in low or "wise target" in low:
                        continue
                    has_monthly_target = any(
                        isinstance(ws.cell(row_no, col).value, (int, float))
                        for col in range(3, 15)
                    )
                    if has_monthly_target and value.upper() not in STANDARD_AREAS:
                        names.add(value)
        finally:
            wb.close()
        if not names:
            print("WARNING: No A00/CoRos outlet names were found in the objective workbook.")
        return names

    def _read_sales(self, info: dict[str, Any], period: str) -> None:
        path: Path = info["path"]
        header_row = int(info["header_row"])
        headers: dict[str, int] = info["headers"]
        date_name = "Billing Date" if "Billing Date" in headers else "Services Rendered On"

        qty_name = "Addtional Qty" if "Addtional Qty" in headers else "Additional Qty"
        required_names = [date_name, "Material", "Outlet Name", "Sales Group", qty_name]
        indices = {name: headers[name] for name in required_names}
        material_text_index = headers.get("Material Text")
        mat_group_text_index = headers.get("Mat Group Text")
        sold_to_party_index = headers.get("Sold to party")

        count = 0
        for row_number, row in _fast_xlsx_rows(path):
            if row_number <= header_row:
                continue
            count += 1
            dt = _to_date(row[indices[date_name]] if indices[date_name] < len(row) else None)
            if not dt:
                continue
            material = _normalise(row[indices["Material"]] if indices["Material"] < len(row) else None)
            material_text = _normalise(row[material_text_index] if material_text_index is not None and material_text_index < len(row) else None)
            mat_group_text = _normalise(row[mat_group_text_index] if mat_group_text_index is not None and mat_group_text_index < len(row) else None)
            product = self._product_for(material, material_text)
            if product is None:
                continue
            source_outlet = _normalise(row[indices["Outlet Name"]] if indices["Outlet Name"] < len(row) else None) or "Unknown Outlet"
            raw_group = _normalise(row[indices["Sales Group"]] if indices["Sales Group"] < len(row) else None) or "Unassigned"
            sold_to_party = _normalise_code(row[sold_to_party_index] if sold_to_party_index is not None and sold_to_party_index < len(row) else None)
            cost_center = self.code_to_cost_center.get(sold_to_party)
            if not cost_center:
                # Fallback rule: when no Cost Center mapping exists, use Sold-to
                # Party as the outlet identity. This keeps the row in all totals.
                # Add the code to Outlet_Master later to merge it with a permanent
                # Cost Center or to update its current display name.
                cost_center = sold_to_party or _outlet_key(source_outlet) or "UNKNOWN OUTLET"
                self.code_to_cost_center.setdefault(sold_to_party, cost_center)
                self.cost_center_info.setdefault(cost_center, {
                    "Outlet Name": source_outlet,
                    "Outlet Code": sold_to_party,
                    "Area": "",
                })
            info = self._display_info(cost_center)
            master_area = _normalise(info.get("Area"))
            # Dynamic Targets FY26 CoRos membership has priority over the historical
            # Area stored in Outlet_Master. Other outlets continue using master/source logic.
            if cost_center in self.a00_cost_centers:
                category = "A00"
            else:
                category = master_area if master_area in AREAS else self._area_category(raw_group, source_outlet, cost_center)
            outlet = cost_center
            self.observed_products[period][cost_center].add(product)
            if category in AREAS:
                self.observed_area_counts[period][cost_center][category] += 1
            sold_to_party_value: Any = int(sold_to_party) if sold_to_party.isdigit() else sold_to_party
            self.outlet_sold_to_party[outlet] = sold_to_party_value
            qty = _to_float(row[indices[qty_name]] if indices[qty_name] < len(row) else None)

            self.product_totals[period][product] += qty
            self.area_totals[period][(product, category)] += qty
            self.product_daily[period][(product, dt.date())] += qty
            self.area_daily[period][(product, category, dt.date())] += qty
            self.outlet_product_totals[period][(product, category, outlet)] += qty
            self.outlet_monthly_totals[period][(product, category, outlet, dt.year, dt.month)] += qty
            self.monthly_product_totals[period][(product, dt.year, dt.month)] += qty
            self.monthly_area_product_totals[period][(product, category, dt.year, dt.month)] += qty
            if product == "Lubricants":
                # Group lubricant drill-downs by the source workbook's exact
                # "Mat Group Text" field (for example DMAINSTREAM, DPREMIUM).
                # Fall back only when an older input file does not have that column.
                material_group = mat_group_text or material_text or material
                # Retain Material Text as a separate dimension so Mat Group Text
                # summaries can show the exact lubricant product description.
                material_description = material_text or material
                self.material_outlet_totals[period][(material_group, material_description, category, outlet)] += qty
                self.material_group_daily[period][(material_group, category, dt.date())] += qty
                self.material_group_monthly[period][(material_group, category, dt.year, dt.month)] += qty

            month = dt.strftime("%B")
            fiscal_month = FISCAL_MONTHS.index(month) + 1 if month in FISCAL_MONTHS else 99
            key = (product, outlet, category, month, fiscal_month)
            if period == "this":
                if self.start_date is None or dt < self.start_date:
                    self.start_date = dt
                if self.end_date is None or dt > self.end_date:
                    self.end_date = dt
                days = self.daily_this.setdefault(key, [0.0] * 31)
                days[dt.day - 1] += qty
            else:
                days = self.daily_last.setdefault(key, [0.0] * 31)
                days[dt.day - 1] += qty

            if count % 25000 == 0:
                print(f"  processed {count:,} rows...")

    def _read_outlet_objectives(self, path: Path, sheet_map: dict[str, str]) -> defaultdict[tuple[str, str, str], float]:
        """Read individual outlet objectives from the outlet-level workbook.

        Values in this workbook are in KL and are converted to litres internally.
        They are used only for outlet/control-sheet rows and are never summed to
        create area objectives, preventing conflicts with the official area file.
        """
        outlet_objectives: defaultdict[tuple[str, str, str], float] = defaultdict(float)
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            month_aliases = {m[:3].casefold(): m for m in FISCAL_MONTHS}
            month_aliases.update({m.casefold(): m for m in FISCAL_MONTHS})
            valid_areas = {area.casefold(): area for area in AREAS}
            for product, sheet_name in sheet_map.items():
                ws = wb[sheet_name]
                headers = {_normalise(ws.cell(1, c).value).casefold(): c for c in range(1, ws.max_column + 1)}
                if not {"area", "name", "cost center"}.issubset(headers):
                    raise ValueError(f"Outlet objective sheet '{sheet_name}' must contain Area, Name and Cost Center columns.")
                month_cols = {month_aliases[key]: col for key, col in headers.items() if key in month_aliases}
                if len(month_cols) < 12:
                    raise ValueError(f"Outlet objective sheet '{sheet_name}' must contain Jul through Jun columns.")
                max_col = max([headers["area"], headers["name"], headers["cost center"], headers.get("code", 0), *month_cols.values()])
                for values in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
                    def at(col: int | None):
                        return values[col - 1] if col and col - 1 < len(values) else None
                    raw_name = _normalise(at(headers["name"]))
                    if not raw_name or raw_name.casefold() in {"total", "grand total"}:
                        continue
                    cc = _normalise_code(at(headers["cost center"]))
                    if not cc:
                        continue
                    raw_area = _normalise(at(headers["area"]))
                    area = valid_areas.get(raw_area.casefold(), raw_area.upper())
                    code_col = headers.get("code")
                    code = _normalise_code(at(code_col)) if code_col else ""
                    self.cost_center_info.setdefault(cc, {"Outlet Name": raw_name, "Outlet Code": code, "Area": area})
                    self.cost_center_alias_names[cc].add(raw_name)
                    if code:
                        self.code_to_cost_center.setdefault(code, cc)
                    for month, col in month_cols.items():
                        outlet_objectives[(product, cc, month)] += _to_float(at(col)) * 1000.0
        finally:
            wb.close()
        if not outlet_objectives:
            raise ValueError("No outlet-wise objectives could be read from the outlet objective workbook.")
        return outlet_objectives

    def _read_area_objectives(self, path: Path, sheet_map: dict[str, str]) -> defaultdict[tuple[str, str, str], float]:
        """Read official area-wise objectives from the area-summary workbook.

        These values are already in litres and exclusively power Main, DDM, Area,
        analytics and aggregate objective calculations.
        """
        objectives: defaultdict[tuple[str, str, str], float] = defaultdict(float)
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            valid_areas = {area.casefold(): area for area in AREAS}
            month_aliases = {m[:3].casefold(): m for m in FISCAL_MONTHS}
            month_aliases.update({m.casefold(): m for m in FISCAL_MONTHS})
            for product, sheet_name in sheet_map.items():
                ws = wb[sheet_name]
                # Official layout: column B = Sales Area, columns C:N = Jul-Jun.
                month_names: list[str] = []
                for col in range(3, 15):
                    value = ws.cell(1, col).value
                    dt = _to_date(value)
                    month_names.append(dt.strftime("%B") if dt else month_aliases.get(_normalise(value).casefold(), ""))
                if len([m for m in month_names if m]) < 12:
                    raise ValueError(f"Area objective sheet '{sheet_name}' must contain twelve monthly columns from Jul to Jun.")
                found_rows = 0
                for row in range(2, ws.max_row + 1):
                    raw_area = _normalise(ws.cell(row, 2).value)
                    if not raw_area:
                        continue
                    folded = raw_area.casefold()
                    if folded in {"coros", "coro", "cocos", "coco"}:
                        area = "A00"
                    else:
                        area = valid_areas.get(folded)
                    if area is None:
                        # Stop after the official total/CoRos-wise detail section begins.
                        if folded in {"total", "coros wise-targets", "coros wise targets"}:
                            break
                        continue
                    found_rows += 1
                    for col, month in enumerate(month_names, start=3):
                        objectives[(product, area, month)] += _to_float(ws.cell(row, col).value)
                if not found_rows:
                    raise ValueError(f"No area rows were found in area objective sheet '{sheet_name}'.")
        finally:
            wb.close()
        if not objectives:
            raise ValueError("No area-wise objectives could be read from the area objective workbook.")
        return objectives

    def outlet_month_objective(self, product: str, cost_center: str, month: str, area: str | None = None) -> float:
        """Return outlet objective; fall back to zero when legacy file has no outlet detail."""
        return float(getattr(self, "outlet_objectives", {}).get((product, str(cost_center), month), 0.0))

    def objective_total(self, product: str | None = None, area: str | None = None) -> float:
        current_month = self.end_date.strftime("%B")
        current_pos = FISCAL_MONTHS.index(current_month) if current_month in FISCAL_MONTHS else 11
        allowed = set(FISCAL_MONTHS[: current_pos + 1])
        total = 0.0
        for (item_product, item_area, month), value in self.objectives.items():
            if month not in allowed:
                continue
            if product and item_product != product:
                continue
            if area and item_area.lower() != area.lower():
                continue
            total += value
        return total

    @staticmethod
    def _with_variance(row: dict[str, Any]) -> dict[str, Any]:
        ty = float(row["This Year Sales"])
        ly = float(row["Last Year Sales"])
        obj = float(row["Objective"])
        row["TY vs LY Variance"] = ty - ly
        row["TY vs LY %"] = (ty - ly) / ly if ly else 0.0
        row["TY vs Objective Variance"] = ty - obj
        row["TY vs Objective %"] = (ty - obj) / obj if obj else 0.0
        return row

    def product_summary(self) -> list[dict[str, Any]]:
        rows = []
        for product in PRODUCTS:
            rows.append(self._with_variance({
                "Product": product,
                "This Year Sales": self.product_totals["this"][product],
                "Last Year Sales": self.product_totals["last"][product],
                "Objective": self.objective_total(product=product),
            }))
        return rows

    def area_product_summary(self, area: str) -> list[dict[str, Any]]:
        rows = []
        for product in PRODUCTS:
            rows.append(self._with_variance({
                "Product": product,
                "This Year Sales": self.area_totals["this"][(product, area)],
                "Last Year Sales": self.area_totals["last"][(product, area)],
                "Objective": self.objective_total(product=product, area=area),
            }))
        return rows

    def daily_detail(self, product: str, area: str | None = None) -> list[dict[str, Any]]:
        cache_key = (product, area)
        if cache_key in self._daily_cache:
            return self._daily_cache[cache_key]

        rows: list[dict[str, Any]] = []

        # Use the union of both years and seed every target A00 outlet for all
        # fiscal months. This makes zero-sales outlets visible instead of hiding them.
        all_keys = set(self.daily_this) | set(self.daily_last)
        if area is None or area.casefold() == "a00":
            # Seed every Targets FY26 CoRos Cost Center/report key for every
            # fiscal month. Outlets without sales therefore remain visible as 0.
            for outlet_cc in sorted(self.a00_cost_centers):
                for fiscal_month, month in enumerate(FISCAL_MONTHS, 1):
                    all_keys.add((product, outlet_cc, "A00", month, fiscal_month))
        for key in all_keys:
            item_product, outlet, category, month, fiscal_month = key
            if item_product != product:
                continue
            if area and category.casefold() != area.casefold():
                continue

            days = self.daily_this.get(key, [0.0] * 31)
            last_days = self.daily_last.get(key, [0.0] * 31)
            last_total = float(sum(last_days))
            this_total = float(sum(days))
            row: dict[str, Any] = {
                "Cost Center": outlet,
                "Outlet Name": self._display_info(outlet).get("Outlet Name", outlet),
                "Sold to Party": self._display_info(outlet).get("Outlet Code", self.outlet_sold_to_party.get(outlet, "")),
                # Show/filter by the final dynamically classified area.
                "Sales Group": category,
                "Month": month,
                "Fiscal Month": fiscal_month,
            }
            for day in range(1, 32):
                row[day] = float(days[day - 1])
                row[f"LY {day}"] = float(last_days[day - 1])
            objective = self.outlet_month_objective(product, outlet, month, category)
            row["This Year Sales"] = this_total
            row["Last Year Sales"] = last_total
            row["TY vs LY Variance"] = this_total - last_total
            row["TY vs LY %"] = (this_total - last_total) / last_total if last_total else 0.0
            row["Objective"] = objective
            row["TY vs Objective Variance"] = this_total - objective
            row["TY vs Objective %"] = (this_total - objective) / objective if objective else 0.0
            rows.append(row)

        rows.sort(key=lambda r: (r["Fiscal Month"], str(r["Sales Group"]), str(r["Outlet Name"])))
        self._daily_cache[cache_key] = rows
        return rows


    def _read_lubricant_liter_master(self, path: Path) -> dict[str, float]:
        if not path.exists():
            return {}
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            headers = {_normalise(v).casefold(): i for i, v in enumerate(next(rows))}
            product_col = headers.get("product", headers.get("material description", 0))
            liter_col = headers.get("liter", headers.get("liters", 1))
            result = {}
            for row in rows:
                if product_col >= len(row) or liter_col >= len(row):
                    continue
                product = _normalise(row[product_col])
                liters = _to_float(row[liter_col])
                if product and product.casefold() != "grand total" and liters > 0:
                    result[product.casefold()] = liters
            return result
        finally:
            wb.close()

    def _snapshot_type(self, path: Path) -> str | None:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                for ws in wb.worksheets:
                    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                    h = {_normalise(v).casefold() for v in first}
                    if {"order quantity (item)", "sales unit", "sales group", "material description"}.issubset(h):
                        return "open"
                    if {"sales group", "item description", "delivery quantity"}.issubset(h):
                        return "transit"
            finally:
                wb.close()
        except Exception:
            return None
        return None

    def _read_lubricant_snapshots(self) -> None:
        """Load Open Orders and Transit month-wise from their own dated rows.

        Open quantities are converted to litres using the ``info`` sheet inside
        the Open Orders workbook: Order Quantity (Item) × product litres.
        No snapshot is carried into another month.
        """
        for path in _excel_files():
            kind = self._snapshot_type(path)
            if not kind:
                continue
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                # For Open Orders, the workbook's own info sheet is the source
                # of product-to-litre conversion factors.
                workbook_liters: dict[str, float] = {}
                if kind == "open":
                    info_ws = next((x for x in wb.worksheets if x.title.casefold() == "info"), None)
                    if info_ws is not None:
                        info_rows = info_ws.iter_rows(values_only=True)
                        info_headers = {_normalise(v).casefold(): i for i, v in enumerate(next(info_rows, ())) }
                        product_col = info_headers.get("product", 0)
                        liter_col = info_headers.get("liter", info_headers.get("liters", 1))
                        for info_row in info_rows:
                            if product_col >= len(info_row) or liter_col >= len(info_row):
                                continue
                            product = _normalise(info_row[product_col])
                            liters = _to_float(info_row[liter_col])
                            if product and liters > 0:
                                workbook_liters[product.casefold()] = liters

                ws = next((x for x in wb.worksheets if x.title.casefold() == "data"), wb[wb.sheetnames[0]])
                rows = ws.iter_rows(values_only=True)
                headers = {_normalise(v).casefold(): i for i, v in enumerate(next(rows))}
                date_header = "document date" if kind == "open" else "loading date"
                date_col = headers.get(date_header)
                if date_col is None:
                    print(f"Skipped lubricant {kind} file without {date_header}: {path.name}")
                    continue

                for row in rows:
                    area_col = headers.get("sales group")
                    area = _normalise(row[area_col]) if area_col is not None and area_col < len(row) else ""
                    if area not in AREAS:
                        continue
                    row_date = _to_date(row[date_col] if date_col < len(row) else None)
                    if row_date is None:
                        continue
                    month_start = date(row_date.year, row_date.month, 1)

                    if kind == "open":
                        desc = _normalise(row[headers["material description"]])
                        qty = _to_float(row[headers["order quantity (item)"]])
                        factor = workbook_liters.get(desc.casefold(), 0.0)
                        self.lubricant_open_by_month_area[(month_start, area)] += qty * factor
                    else:
                        desc = _normalise(row[headers["item description"]])
                        qty = _to_float(row[headers["delivery quantity"]])
                        # Transit retains the shared litre master unless its file
                        # later gains an internal info sheet.
                        factor = self.lubricant_liter_master.get(desc.casefold(), 0.0)
                        self.lubricant_transit_by_month_area[(month_start, area)] += qty * factor
            finally:
                wb.close()
            print(f"Loaded month-wise lubricant {kind}: {path.name}")

    def lubricant_area_pipeline_rows(self, selected_month: date | None = None) -> list[dict[str, Any]]:
        """Return month-wise lubricant pipeline rows.

        Execution comes from the selected month in the main sales data.
        Open Orders and Transit come only from rows dated in that same selected
        month. Missing months remain zero; values are never carried forward.
        """
        assert self.start_date and self.end_date
        latest_asof = self.end_date.date()
        today = date.today()
        selected_month = selected_month or today.replace(day=1)
        month_start = selected_month.replace(day=1)
        month_last = date(month_start.year, month_start.month, calendar.monthrange(month_start.year, month_start.month)[1])
        is_current_month = (month_start.year, month_start.month) == (today.year, today.month)
        asof = min(today, month_last) if is_current_month else month_last

        ly_start = month_start.replace(year=month_start.year - 1)
        ly_end = date(ly_start.year, ly_start.month, calendar.monthrange(ly_start.year, ly_start.month)[1])
        elapsed_days = self._non_sunday_day_count(month_start, asof)
        ly_days = self._non_sunday_day_count(ly_start, ly_end)
        rows=[]
        for area in AREAS:
            execution = self._period_sales("this", "Lubricants", month_start, asof, area)
            open_qty = self.lubricant_open_by_month_area.get((month_start, area), 0.0)
            transit = self.lubricant_transit_by_month_area.get((month_start, area), 0.0)
            total = execution + open_qty + transit
            ly_full = self._period_sales("last", "Lubricants", ly_start, ly_end, area)
            ly_adj = (ly_full / ly_days * elapsed_days) if ly_days else 0.0
            objective = self.objectives.get(("Lubricants", area, month_start.strftime("%B")), 0.0)
            rows.append({
                "Area": area, "Execution": execution, "Open": open_qty,
                "Transit": transit, "Total": total, "LY Adjusted": ly_adj,
                "Objective": objective,
                "Achieved %": total / objective if objective else 0.0,
                "TY vs LY %": (total-ly_adj)/ly_adj if ly_adj else 0.0,
                "TY vs Objective %": (total-objective)/objective if objective else 0.0,
            })
        return rows

    def lubricant_pipeline_months(self) -> list[date]:
        """Return all 12 month starts for the current July–June fiscal year.

        The selector must not stop at the latest month found in source data. This
        keeps the current month (and the remaining FY months) available even when
        their execution data has not arrived yet.
        """
        today = date.today()
        start_year = today.year if today.month >= 7 else today.year - 1
        current = date(start_year, 7, 1)
        months=[]
        for _ in range(12):
            months.append(current)
            current = date(current.year + (1 if current.month == 12 else 0),
                           1 if current.month == 12 else current.month + 1, 1)
        return months

    def lubricant_pipeline_default_month(self) -> date:
        """Default pipeline selector to the actual current calendar month."""
        today = date.today()
        return today.replace(day=1)

    def lubricant_materials(self) -> list[str]:
        """Return lubricant Mat Group Text groups appearing in either year."""
        return sorted({k[0] for period in ("last", "this") for k in self.material_outlet_totals[period] if k[0]})

    def lubricant_material_text_summary(self, material_group: str, area: str | None = None) -> list[dict[str, Any]]:
        """Return one summary row per Material Text inside a Mat Group Text group."""
        material_texts = sorted({k[1] for period in ("last", "this")
                                 for k in self.material_outlet_totals[period]
                                 if k[0] == material_group and (area is None or k[2] == area) and k[1]})
        rows: list[dict[str, Any]] = []
        for material_text in material_texts:
            ty = sum(v for (group, text, source_area, _outlet), v
                     in self.material_outlet_totals["this"].items()
                     if group == material_group and text == material_text
                     and (area is None or source_area == area))
            ly = sum(v for (group, text, source_area, _outlet), v
                     in self.material_outlet_totals["last"].items()
                     if group == material_group and text == material_text
                     and (area is None or source_area == area))
            rows.append({
                "Material Text": material_text,
                "This Year Sales": ty,
                "Last Year Sales": ly,
                "TY vs LY Variance": ty - ly,
                "TY vs LY %": (ty - ly) / ly if ly else 0.0,
            })
        rows.sort(key=lambda r: r["This Year Sales"], reverse=True)
        return rows

    def lubricant_material_summary(self, material_group: str, area: str | None = None) -> list[dict[str, Any]]:
        """Return outlet-level rows split by the exact Material Text description."""
        combinations = sorted({(k[3], k[1]) for period in ("last", "this")
                               for k in self.material_outlet_totals[period]
                               if k[0] == material_group and (area is None or k[2] == area)})
        rows = []
        for outlet, material_text in combinations:
            ty = sum(v for (group, text, source_area, source_outlet), v
                     in self.material_outlet_totals["this"].items()
                     if group == material_group and text == material_text and source_outlet == outlet
                     and (area is None or source_area == area))
            ly = sum(v for (group, text, source_area, source_outlet), v
                     in self.material_outlet_totals["last"].items()
                     if group == material_group and text == material_text and source_outlet == outlet
                     and (area is None or source_area == area))
            rows.append({"Cost Center": outlet, "Outlet Name": self._display_info(outlet).get("Outlet Name", outlet), "Sold to Party": self._display_info(outlet).get("Outlet Code", self.outlet_sold_to_party.get(outlet, "")), "Material Text": material_text,
                         "This Year Sales": ty, "Last Year Sales": ly,
                         "TY vs LY Variance": ty - ly,
                         "TY vs LY %": (ty - ly) / ly if ly else 0.0})
        rows.sort(key=lambda r: r["This Year Sales"], reverse=True)
        return rows

    def monthly_sales(self, product: str, period: str, area: str | None = None) -> dict[tuple[int,int], float]:
        source = self.monthly_product_totals[period] if area is None else self.monthly_area_product_totals[period]
        if area is None:
            return {(y,m):v for (p,y,m),v in source.items() if p==product}
        return {(y,m):v for (p,a,y,m),v in source.items() if p==product and a==area}

    def top_outlets(self, product: str | None = None, area: str | None = None, limit: int = 20) -> list[dict[str, Any]]:
        totals=defaultdict(float)
        for (p,a,o),v in self.outlet_product_totals["this"].items():
            if (product is None or p==product) and (area is None or a==area): totals[o]+=v
        rows=[{"Cost Center":o,"Outlet Name":self._display_info(o).get("Outlet Name",o),"Sold to Party":self._display_info(o).get("Outlet Code",self.outlet_sold_to_party.get(o,"")),"This Year Sales":v} for o,v in totals.items()]
        rows.sort(key=lambda r:r["This Year Sales"], reverse=True)
        return rows[:limit]

    def outlet_six_month_averages(self) -> list[dict[str, Any]]:
        """Return area/outlet six-month monthly averages by product."""
        assert self.end_date
        first = self.end_date.date().replace(day=1)
        months = []
        cursor = first
        for _ in range(6):
            cursor = (cursor - timedelta(days=1)).replace(day=1)
            months.append((cursor.year, cursor.month))
        outlets = sorted({(a, o) for period in ("last", "this") for (p, a, o, y, m) in self.outlet_monthly_totals[period]})
        rows = []
        for area, outlet in outlets:
            rec = {"Area": area, "Cost Center": outlet, "Outlet Name": self._display_info(outlet).get("Outlet Name", outlet), "Sold to Party": self._display_info(outlet).get("Outlet Code", self.outlet_sold_to_party.get(outlet, ""))}
            total = 0.0
            for product in PRODUCTS:
                vals=[]
                for y,m in months:
                    value = self.outlet_monthly_totals["this"].get((product,area,outlet,y,m))
                    if value is None:
                        value = self.outlet_monthly_totals["last"].get((product,area,outlet,y,m),0.0)
                    vals.append(value)
                avg=sum(vals)/6.0
                rec[f"{product} 6M Avg"] = avg
                if product in ("PMG", "HSD", "R95"):
                    total += avg
            rec["Total 6M Avg"] = total
            rows.append(rec)
        rows.sort(key=lambda r:(r["Area"], -r["Total 6M Avg"], r["Outlet Name"]))
        return rows

    def previous_six_month_average(self, product: str, area: str | None = None) -> float:
        assert self.end_date
        first=self.end_date.date().replace(day=1)
        months=[]
        cursor=first
        for _ in range(6):
            cursor=(cursor.replace(day=1)-timedelta(days=1)).replace(day=1)
            months.append((cursor.year,cursor.month))
        values=[]
        for y,m in months:
            val=self.monthly_sales(product,"this",area).get((y,m))
            if val is None: val=self.monthly_sales(product,"last",area).get((y,m),0.0)
            values.append(val)
        return sum(values)/6.0

    @staticmethod
    def _shift_year(d: date, years: int = -1) -> date:
        try:
            return d.replace(year=d.year + years)
        except ValueError:
            return d.replace(year=d.year + years, day=28)

    def _period_sales(self, period: str, product: str, start: date, end: date, area: str | None = None) -> float:
        if end < start:
            return 0.0
        key=(period,product,area,start,end)
        if key not in self._period_cache:
            total=0.0
            d=start
            if area is None:
                source=self.product_daily[period]
                while d<=end:
                    total += source.get((product,d),0.0)
                    d += timedelta(days=1)
            else:
                source=self.area_daily[period]
                while d<=end:
                    total += source.get((product,area,d),0.0)
                    d += timedelta(days=1)
            self._period_cache[key]=total
        return self._period_cache[key]

    @staticmethod
    def _non_sunday_day_count(start: date, end: date) -> int:
        """Count elapsed calendar days excluding Sundays, inclusive."""
        if end < start:
            return 0
        total = (end - start).days + 1
        sundays = sum(1 for i in range(total) if (start + timedelta(days=i)).weekday() == 6)
        return total - sundays

    @staticmethod
    def _days_in_period(start: date, end: date) -> int:
        return max(0, (end - start).days + 1)

    def _prorated_last_year_comparison(
        self,
        product: str,
        current_start: date,
        current_end: date,
        last_full_start: date,
        last_full_end: date,
        area: str | None = None,
    ) -> tuple[float, float, int, float, float]:
        """
        Compare current sales with prorated full last-year sales.

        Last Year MTD formula:

            Last Year Full Month Sale
            / Total Calendar Days in Last Year Month
            * Calendar Days Elapsed in Current Month

        Calendar days are used intentionally. Sundays are included.
        """

        elapsed_days = (current_end - current_start).days + 1

        current_sales = self._period_sales(
            "this",
            product,
            current_start,
            current_end,
            area,
        )

        last_full_sales = self._period_sales(
            "last",
            product,
            last_full_start,
            last_full_end,
            area,
        )

        last_period_days = (
            last_full_end - last_full_start
        ).days + 1

        last_daily_rate = (
            last_full_sales / last_period_days
            if last_period_days > 0
            else 0.0
        )

        last_adjusted = last_daily_rate * elapsed_days

        current_avg = (
            current_sales / elapsed_days
            if elapsed_days > 0
            else 0.0
        )

        return (
            current_sales,
            last_adjusted,
            elapsed_days,
            current_avg,
            last_daily_rate,
        )


    def _last_year_ytd_comparison(
        self,
        product: str,
        fiscal_start: date,
        asof: date,
        area: str | None = None,
    ) -> float:
        """
        Calculate Last Year YTD as:

        1. Actual full sales of all completed months of the
           current fiscal year using corresponding months
           from last year.

        PLUS

        2. Corresponding current month of last year on
           calendar-day MTD basis.

        Example: As of 25-Aug-2026

            July-2025 full sales
            +
            August-2025 full sale / 31 * 25

        Example: As of 25-Sep-2026

            July-2025 full sales
            +
            August-2025 full sales
            +
            September-2025 full sale / 30 * 25
        """

        current_month_start = asof.replace(day=1)

        # ---------------------------------------------
        # PART 1:
        # Full sales of completed months from last year
        # ---------------------------------------------

        if current_month_start > fiscal_start:

            completed_end = (
                current_month_start
                - timedelta(days=1)
            )

            ly_completed_start = self._shift_year(
                fiscal_start
            )

            ly_completed_end = self._shift_year(
                completed_end
            )

            completed_month_sales = self._period_sales(
                "last",
                product,
                ly_completed_start,
                ly_completed_end,
                area,
            )

        else:

            completed_month_sales = 0.0

        # ---------------------------------------------
        # PART 2:
        # Current month of last year on MTD basis
        # ---------------------------------------------

        elapsed_days = (
            asof - current_month_start
        ).days + 1

        ly_month_start = self._shift_year(
            current_month_start
        )

        ly_month_end = date(
            ly_month_start.year,
            ly_month_start.month,
            calendar.monthrange(
                ly_month_start.year,
                ly_month_start.month,
            )[1],
        )

        ly_month_full_sales = self._period_sales(
            "last",
            product,
            ly_month_start,
            ly_month_end,
            area,
        )

        ly_month_days = (
            ly_month_end - ly_month_start
        ).days + 1

        ly_month_mtd = (
            ly_month_full_sales
            / ly_month_days
            * elapsed_days
            if ly_month_days > 0
            else 0.0
        )

        return (
            completed_month_sales
            + ly_month_mtd
        )


    def _objective_for_month(
        self,
        product: str,
        month: str,
        area: str | None = None,
    ) -> float:

        return sum(
            value
            for (p, a, m), value
            in self.objectives.items()
            if p == product
            and m == month
            and (area is None or a == area)
        )


    def _objective_through_month(
        self,
        product: str,
        month: str,
        area: str | None = None,
    ) -> float:

        pos = (
            FISCAL_MONTHS.index(month)
            if month in FISCAL_MONTHS
            else len(FISCAL_MONTHS) - 1
        )

        allowed = set(
            FISCAL_MONTHS[:pos + 1]
        )

        return sum(
            value
            for (p, a, m), value
            in self.objectives.items()
            if p == product
            and m in allowed
            and (area is None or a == area)
        )

        def comparison_records(
        self,
        area: str | None = None,
    ) -> list[dict[str, Any]]:

        """Precompute dashboard values for every possible automatic as-of date."""

        assert self.start_date and self.end_date

        start = self.start_date.date()

        fiscal_start = date(
            start.year if start.month >= 7 else start.year - 1,
            7,
            1,
        )

        # Build records for the complete fiscal year.
        end = date(
            fiscal_start.year + 1,
            6,
            30,
        )

        records: list[dict[str, Any]] = []

        asof = fiscal_start

        while asof <= end:

            month_start = asof.replace(day=1)

            ly_month_year = asof.year - 1

            ly_month_start = date(
                ly_month_year,
                asof.month,
                1,
            )

            ly_month_end = date(
                ly_month_year,
                asof.month,
                calendar.monthrange(
                    ly_month_year,
                    asof.month,
                )[1],
            )

            for product in PRODUCTS:

                # -----------------------------------------
                # Current Year YTD actual sales
                # -----------------------------------------

                fy_ty = self._period_sales(
                    "this",
                    product,
                    fiscal_start,
                    asof,
                    area,
                )

                # -----------------------------------------
                # Last Year YTD
                #
                # Completed full months
                # +
                # Current month on MTD basis
                # -----------------------------------------

                fy_ly = self._last_year_ytd_comparison(
                    product,
                    fiscal_start,
                    asof,
                    area,
                )

                fy_days = (
                    asof - fiscal_start
                ).days + 1

                fy_ty_avg = (
                    fy_ty / fy_days
                    if fy_days > 0
                    else 0.0
                )

                fy_ly_avg = (
                    fy_ly / fy_days
                    if fy_days > 0
                    else 0.0
                )

                # -----------------------------------------
                # MTD Sales
                # -----------------------------------------

                (
                    m_ty,
                    m_ly,
                    m_days,
                    m_ty_avg,
                    m_ly_avg,
                ) = self._prorated_last_year_comparison(
                    product,
                    month_start,
                    asof,
                    ly_month_start,
                    ly_month_end,
                    area,
                )

                # -----------------------------------------
                # Current Month Full Objective
                # -----------------------------------------

                full_month_objective = (
                    self._objective_for_month(
                        product,
                        asof.strftime("%B"),
                        area,
                    )
                )

                # -----------------------------------------
                # MTD Objective
                #
                # Full month objective
                # / calendar days in month
                # * elapsed calendar days
                # -----------------------------------------

                days_in_month = calendar.monthrange(
                    asof.year,
                    asof.month,
                )[1]

                mtd_objective = (
                    full_month_objective
                    / days_in_month
                    * m_days
                    if days_in_month > 0
                    else 0.0
                )

                # -----------------------------------------
                # FY / YTD Objective
                #
                # Completed months = FULL objective
                # Current month = MTD prorated objective
                # -----------------------------------------

                completed_months_objective = (
                    self._objective_through_month(
                        product,
                        asof.strftime("%B"),
                        area,
                    )
                    - full_month_objective
                )

                fy_objective = (
                    completed_months_objective
                    + mtd_objective
                )

                # -----------------------------------------
                # Final dashboard record
                # -----------------------------------------

                records.append({
                    "As Of Date": asof,
                    "Product": product,

                    "FY Current": fy_ty,
                    "FY Last Calendar": fy_ly,
                    "FY Last Weekday Adj": fy_ly_avg,
                    "FY Days": fy_days,

                    "FY Objective": fy_objective,

                    "MTD Current": m_ty,
                    "MTD Last Calendar": m_ly,
                    "MTD Last Weekday Adj": m_ly_avg,
                    "MTD Days": m_days,

                    "MTD Objective": mtd_objective,
                })

            # Move to the next as-of date.
            asof += timedelta(days=1)

        return records                    

        
        

        
        

    def selected_quarter_records(self, area: str | None = None) -> list[dict[str, Any]]:
        """Precompute fiscal-quarter performance for each available as-of date."""
        import calendar
        assert self.start_date and self.end_date
        start = self.start_date.date()
        fiscal_start = date(start.year if start.month >= 7 else start.year - 1, 7, 1)
        # Build dashboard lookup rows through the fiscal-year end so TODAY() can
        # advance automatically even when the latest sales file is one or more
        # days behind. Dates without uploaded sales correctly show zero new sales.
        end = date(fiscal_start.year + 1, 6, 30)
        quarter_months = {
            "Q1": ("July", "August", "September"),
            "Q2": ("October", "November", "December"),
            "Q3": ("January", "February", "March"),
            "Q4": ("April", "May", "June"),
        }
        records: list[dict[str, Any]] = []
        asof = fiscal_start
        while asof <= end:
            for quarter, months in quarter_months.items():
                first_month_num = list(calendar.month_name).index(months[0])
                last_month_num = list(calendar.month_name).index(months[-1])
                qstart_year = fiscal_start.year if first_month_num >= 7 else fiscal_start.year + 1
                qend_year = fiscal_start.year if last_month_num >= 7 else fiscal_start.year + 1
                qstart = date(qstart_year, first_month_num, 1)
                qend = date(qend_year, last_month_num, calendar.monthrange(qend_year, last_month_num)[1])
                active = asof >= qstart
                effective = min(asof, qend)
                ly_start = date(qstart.year - 1, qstart.month, qstart.day)
                ly_end = date(qend.year - 1, qend.month, qend.day)
                for product in PRODUCTS:
                    if active:
                        ty, ly, days, ty_avg, ly_avg = self._prorated_last_year_comparison(
                            product, qstart, effective, ly_start, ly_end, area
                        )
                    else:
                        ty = ly = ty_avg = ly_avg = 0.0
                        days = 0
                    objective = sum(self._objective_for_month(product, month, area) for month in months)
                    records.append({
                        "As Of Date": asof,
                        "Quarter": quarter,
                        "Product": product,
                        "Current": ty,
                        "Last Calendar": ly,
                        "Last Weekday Adj": ly_avg,
                        "Days": days,
                        "Objective": objective,
                    })
            asof += timedelta(days=1)
        return records

    def selected_month_records(self, area: str | None = None) -> list[dict[str, Any]]:
        import calendar
        assert self.start_date and self.end_date
        start = self.start_date.date()
        fiscal_start = date(start.year if start.month >= 7 else start.year - 1, 7, 1)
        # Build dashboard lookup rows through the fiscal-year end so TODAY() can
        # advance automatically even when the latest sales file is one or more
        # days behind. Dates without uploaded sales correctly show zero new sales.
        end = date(fiscal_start.year + 1, 6, 30)
        records: list[dict[str, Any]] = []
        asof = fiscal_start
        while asof <= end:
            for month_name in FISCAL_MONTHS:
                month_num = list(calendar.month_name).index(month_name)
                year = fiscal_start.year if month_num >= 7 else fiscal_start.year + 1
                mstart = date(year, month_num, 1)
                mend = date(year, month_num, calendar.monthrange(year, month_num)[1])
                active = asof >= mstart
                effective = min(asof, mend)
                ly_year = year - 1
                ly_start = date(ly_year, month_num, 1)
                ly_end = date(ly_year, month_num, calendar.monthrange(ly_year, month_num)[1])
                for product in PRODUCTS:
                    if active:
                        ty, ly, days, ty_avg, ly_avg = self._prorated_last_year_comparison(
                            product, mstart, effective, ly_start, ly_end, area
                        )
                    else:
                        ty = ly = ty_avg = ly_avg = 0.0
                        days = 0
                    records.append({
                        "As Of Date": asof,
                        "Month": month_name,
                        "Product": product,
                        "Current": ty,
                        "Last Calendar": ly,
                        "Last Weekday Adj": ly_avg,
                        "Days": days,
                        "Objective": self._objective_for_month(product, month_name, area),
                    })
            asof += timedelta(days=1)
        return records

